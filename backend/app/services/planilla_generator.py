"""
Service: Planilla Generator
Generates the monthly teacher payroll Excel file from scratch using openpyxl.

Design Decisions:
  - Generate from scratch (NOT clone template) — gives full control over layout.
  - One row per teacher × subject × group combination.
  - Two output sheets: "Planilla" (summary) and "Detalle" (granular slot view).
  - Payment rate: configurable from admin UI (default 70 Bs/academic hour).
  - Supports payment_overrides with row keys "teacher_ci:designation_id"
    and teacher-total keys {teacher_ci: float} for admin adjustments.
  - Freeze panes at row 7, col 4 (so identity cols + headers always visible).

Payment Model C:
  - Base pay = designation.monthly_hours × HOURLY_RATE (from app_settings)
  - Deduct ONLY hours where status=ABSENT
  - Teachers without ANY biometric record get full pay (0 absences assumed)
  - Payable hours = max(0, monthly_hours - absent_hours)

Column Layout:
  A(1)–P(16)  : Identity columns (Semestre, Nombre, CI, ..., Banco)
  Q(17)–AU(47): Days 1–31 of the month (always 31 columns; empty for non-existent days)
  AV(48)      : Hrs Pagables (payable_hours = base - absent)
  AW(49)      : Grupo
  AX(50)      : Materia
  AY(51)      : Pago por Hora (from app_settings HOURLY_RATE)
  AZ(52)      : Hrs Asignadas (base_monthly_hours from designation)
  BA(53)      : Hrs Descontadas (absent_hours)
  BB(54)      : Hrs Asistidas (attended hours, informational)
  BC(55)      : Total Pagable (payable_hours, verification)
  BD(56)      : Total Pago Calculado (payable_hours × HOURLY_RATE)
  BE(57)      : Pago Ajustado (admin override, NULL = use BD)
  BF(58)      : Observaciones

Month name mapping (Spanish):
  1=Enero, 2=Febrero, 3=Marzo, 4=Abril, 5=Mayo, 6=Junio,
  7=Julio, 8=Agosto, 9=Septiembre, 10=Octubre, 11=Noviembre, 12=Diciembre
"""
from __future__ import annotations

import calendar
import logging
import os
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.attendance import AttendanceRecord
from app.models.designation import Designation
from app.models.planilla import PlanillaOutput
from app.models.teacher import Teacher
from app.services import app_settings_service
from app.services.monetary_snapshot import build_calculation_snapshot
from app.services.payment_overrides import (
    calculate_override_total,
    get_teacher_override_allocations,
    normalize_payment_overrides,
    validate_payment_override_targets,
)

if TYPE_CHECKING:
    from app.schemas.planilla import ExcludedDaySchema

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Fallback rate when a row is constructed outside a DB context (tests, legacy
# callers).  The live value comes from ``app_settings_service.get_hourly_rate``
# and is injected through ``_build_planilla_data``/``_build_row``.
RATE_PER_HOUR: float = 70.0  # Bs per academic hour

# Spanish month names
MONTH_NAMES: dict[int, str] = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# Day-of-week letter codes (Spanish abbreviations)
# Python weekday(): 0=Mon, 1=Tue, 2=Wed, 3=Thu, 4=Fri, 5=Sat, 6=Sun
WEEKDAY_LETTERS: dict[int, str] = {
    0: "L",   # Lunes
    1: "M",   # Martes
    2: "M",   # Miércoles
    3: "J",   # Jueves
    4: "V",   # Viernes
    5: "S",   # Sábado
    6: "D",   # Domingo
}

# Column indices (1-based) for identity fields
COL_SEMESTRE = 1         # A
COL_NOMBRE = 2           # B
COL_CI = 3               # C
COL_EMAIL = 4            # D
COL_PHONE = 5            # E
COL_MATERIA = 6          # F
COL_GRUPO = 7            # G
COL_TIPO_DOCENTE = 8     # H  Externo/Permanente/Titular
COL_GENERO = 9           # I
COL_SAP = 10             # J
COL_FACTURA = 11         # K
COL_CUENTA = 12          # L
COL_NIVEL_ACAD = 13      # M
COL_PROFESION = 14       # N
COL_ESPECIALIDAD = 15    # O
COL_BANCO = 16           # P

# Day columns: Q(17) = day 1 ... AU(47) = day 31
DAY_COL_START = 17       # Q
DAY_COL_END = 47         # AU  (DAY_COL_START + 30)

# Summary columns (after days)
COL_TOTAL_HORAS = 48     # AV
COL_GRUPO_RESUMEN = 49   # AW
COL_MATERIA_RESUMEN = 50 # AX
COL_PAGO_HORA = 51       # AY
COL_HRS_TEORIA = 52      # AZ
COL_HRS_PRACT_INT = 53   # BA
COL_HRS_PRACT_EXT = 54   # BB
COL_TOTAL_HRS_CHECK = 55 # BC
COL_PAGO_CALCULADO = 56  # BD  (bruto)
COL_RETENCION_AMT = 57   # BE  (-13% RC-IVA si aplica)
COL_PAGO_NETO = 58       # BF  (final_payment = calculado - retención)
COL_PAGO_AJUSTADO = 59   # BG
COL_OBSERVACIONES = 60   # BH

TOTAL_COLS = 60  # BH

# Row layout in the worksheet
ROW_TITLE = 1
ROW_UNIVERSITY = 2
ROW_EMPTY = 3
ROW_SECTION_HEADERS = 4   # Merged section labels: DATOS DOCENTE | ASISTENCIA ... | RESUMEN
ROW_COL_HEADERS = 5       # Actual column names + day numbers
ROW_WEEKDAY = 6           # Day-of-week letters (L/M/M/J/V/S/D) under day columns
DATA_ROW_START = 7        # First data row

# ---------------------------------------------------------------------------
# Colors / Styles
# ---------------------------------------------------------------------------

COLOR_HEADER_BG = "1F4E79"          # Dark blue — main header
COLOR_SECTION_BG = "2E75B6"         # Medium blue — section headers
COLOR_COL_HEADER_BG = "BDD7EE"      # Light blue — column name headers
COLOR_WEEKDAY_BG = "DEEAF1"         # Very light blue — day-of-week row
COLOR_DAY_CLASS = "E2EFDA"          # Light green — day with classes
COLOR_DAY_ABSENT = "FCE4D6"         # Light red/orange — absent day
COLOR_DAY_LATE = "FFF2CC"           # Light yellow — late day
COLOR_DAY_WEEKEND = "F2F2F2"        # Light gray — weekend/no schedule
COLOR_DAY_EXCLUDED = "FF6B6B"       # Red — admin-excluded day
COLOR_SUMMARY_BG = "EDEDED"         # Light gray — summary columns header
COLOR_TOTAL_ROW = "FFE699"          # Amber — totals row
COLOR_WHITE = "FFFFFF"
COLOR_OUTSIDE_CUTOFF = "EFEFEF"     # Light gray — day outside cutoff window
COLOR_SPACER = "D9D9D9"             # Slightly darker gray — spacer column between month blocks

THIN_SIDE = Side(border_style="thin", color="B8B8B8")
MEDIUM_SIDE = Side(border_style="medium", color="595959")

THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)
MEDIUM_BORDER = Border(
    left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE, bottom=MEDIUM_SIDE
)

# ---------------------------------------------------------------------------
# Data Transfer Objects
# ---------------------------------------------------------------------------


@dataclass
class MonthBlock:
    """Describes one block of day columns in the Excel.

    When the payroll period fits within a single calendar month (legacy
    behavior), there is ONE block covering days 1..days_in_month with
    ``active_start=1`` and ``active_end=days_in_month``.

    When the cutoff crosses months (e.g. Mar 21 → Apr 20), there are TWO
    blocks — one per calendar month — and ``active_start``/``active_end``
    mark the sub-range that holds real attendance data. Days OUTSIDE that
    range are rendered grayed out to signal "outside cutoff period".
    """

    month: int
    year: int
    month_name: str
    days_in_month: int     # total calendar days in this month
    col_start: int         # Excel column index (1-based) for day 1 of this block
    active_start: int      # first day (1-based) within the cutoff window
    active_end: int        # last day (1-based) within the cutoff window


@dataclass
class PlanillaRow:
    """One row in the planilla = one teacher × one subject × one group."""

    # Identity
    teacher_ci: str
    designation_id: int
    teacher_name: str
    email: Optional[str]
    phone: Optional[str]
    subject: str
    semester: str
    group_code: str
    teacher_type: Optional[str]        # Externo/Permanente/Titular
    gender: Optional[str]
    sap_code: Optional[str]
    invoice_retention: Optional[str]
    account_number: Optional[str]
    academic_level: Optional[str]
    profession: Optional[str]
    specialty: Optional[str]
    bank: Optional[str]
    nit: Optional[str] = None

    # Daily hours: {date: academic_hours} — keyed by full date to support
    # cross-month attendance windows (e.g. Mar 21 → Apr 20).
    daily_hours: dict[date, int] = field(default_factory=dict)

    # Status per day: {date: status_string} — for background coloring
    daily_status: dict[date, str] = field(default_factory=dict)

    # Model C payment fields
    base_monthly_hours: int = 0        # From designation.monthly_hours (assigned load)
    absent_hours: int = 0              # Hours deducted for ABSENT slots
    payable_hours: int = 0             # base_monthly_hours - absent_hours (effective pay)

    # Legacy totals (kept for backward compat and Excel day-column sums)
    total_hours: int = 0               # payable_hours — used for payment calculation
    total_theory_hours: int = 0        # base_monthly_hours shown in summary
    total_practice_internal_hours: int = 0   # absent_hours shown as deductions
    total_practice_external_hours: int = 0   # attended hours (informational)

    # Payment
    rate_per_hour: float = RATE_PER_HOUR
    calculated_payment: float = 0.0

    # RC-IVA 13% retention
    has_retention: bool = False           # True if teacher has invoice_retention = "RETENCION"
    retention_rate: float = 0.0           # 0.13 if has_retention, else 0.0
    retention_amount: float = 0.0         # calculated_payment * retention_rate
    final_payment: float = 0.0           # calculated_payment - retention_amount

    # Observations
    observations: list[str] = field(default_factory=list)
    late_count: int = 0
    absent_count: int = 0
    has_biometric: bool = True         # False = no biometric records at all → full pay


@dataclass
class DetailRow:
    """One granular class-slot row for the Detalle sheet."""

    ci: str
    teacher_name: str
    date: date
    day_letter: str
    subject: str
    group_code: str
    semester: str
    scheduled_start: str
    scheduled_end: str
    academic_hours: int
    status: str
    observation: Optional[str]


@dataclass
class PlanillaResult:
    """Result returned by PlanillaGenerator.generate()."""

    file_path: str
    month: int
    year: int
    total_teachers: int
    total_rows: int
    total_hours: int
    total_payment: float
    planilla_output_id: Optional[int]
    warnings: list[str] = field(default_factory=list)
    discount_mode: str = "attendance"


@dataclass(frozen=True)
class ScheduledSlot:
    slot_date: date
    scheduled_start: time
    scheduled_end: time
    academic_hours: int


@dataclass(frozen=True)
class EffectiveDesignationRange:
    requested_start: date
    requested_end: date
    start: date | None
    end: date | None
    is_explicit: bool

    @property
    def is_contract_limited(self) -> bool:
        return self.start != self.requested_start or self.end != self.requested_end


class PayrollDataError(ValueError):
    """Blocking payroll validation error with an actionable API payload."""

    def __init__(
        self,
        message: str,
        *,
        code: str = "invalid_payroll_data",
        sample: list[str] | None = None,
    ):
        self.code = code
        self.sample = sample or []
        super().__init__(message)

    def as_detail(self) -> dict[str, object]:
        return {"message": str(self), "code": self.code, "sample": self.sample}


# ---------------------------------------------------------------------------
# Period-based hour calculation
# ---------------------------------------------------------------------------

# Reuse canonical weekday mapping and day-name normalization from
# attendance_engine to guarantee consistency between attendance processing
# and payment calculation.  No local copies — single source of truth.
from app.services.attendance_engine import WEEKDAY_MAP, _normalize_day
from app.services.exclusion_matching import exclusion_matches_designation


def _is_excluded(
    target_date: date,
    semester: str,
    subject: str,
    group_code: str,
    exclusions: "list[ExcludedDaySchema]",
) -> bool:
    """Return True if ``target_date`` is excluded for the given designation context.

    Union semantics: the first exclusion whose date AND scope criteria match
    immediately returns True. Later entries are not evaluated (no
    double-counting possible).

    Scope rules:
      - ``global``   — matches any designation on that date.
      - ``semester`` — matches when ``exclusion.semester_id == semester``.
      - ``subject``  — matches subject_id + group_id + semester_id. Historical
                       entries without semester_id match by subject+group.

    Args:
        target_date: The calendar date to check.
        semester:    The designation's semester identifier.
        subject:     The designation's subject name.
        group_code:  The designation's group code.
        exclusions:  List of ExcludedDaySchema entries (may be empty).

    Returns:
        True if the date is excluded for this designation context, else False.
    """
    for exc in exclusions:
        if exc.date != target_date:
            continue
        if exclusion_matches_designation(
            exc,
            semester=semester,
            subject=subject,
            group_code=group_code,
        ):
            return True
    return False


def _resolve_payroll_period(
    month: int,
    year: int,
    start_date: date | None = None,
    end_date: date | None = None,
) -> tuple[date, date, bool]:
    """Resolve a full month or a complete custom range of at most 63 days."""
    if (start_date is None) != (end_date is None):
        raise PayrollDataError(
            "start_date y end_date deben enviarse juntos, o ambos deben omitirse",
            code="incomplete_date_range",
        )
    if start_date is not None and end_date is not None:
        if start_date > end_date:
            raise PayrollDataError(
                "start_date no puede ser posterior a end_date",
                code="invalid_date_range",
            )
        days = (end_date - start_date).days + 1
        if days > 63:
            raise PayrollDataError(
                f"El período tiene {days} días; el máximo permitido es 63",
                code="date_range_too_large",
            )
        return start_date, end_date, True

    if not 1 <= month <= 12:
        raise PayrollDataError("El mes debe estar entre 1 y 12", code="invalid_month")
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day), False


def _effective_designation_range(
    designation: Designation,
    month: int,
    year: int,
    start_date: date | None = None,
    end_date: date | None = None,
) -> EffectiveDesignationRange:
    requested_start, requested_end, is_explicit = _resolve_payroll_period(
        month, year, start_date, end_date,
    )
    contract_start = getattr(designation, "contract_start_date", None)
    contract_end = getattr(designation, "contract_end_date", None)
    if not isinstance(contract_start, date):
        contract_start = None
    if not isinstance(contract_end, date):
        contract_end = None
    if contract_start and contract_end and contract_start > contract_end:
        raise PayrollDataError(
            f"La designación {designation.id} tiene fechas contractuales incoherentes: "
            f"{contract_start} > {contract_end}",
            code="invalid_contract_range",
        )

    effective_start = max(requested_start, contract_start) if contract_start else requested_start
    effective_end = min(requested_end, contract_end) if contract_end else requested_end
    if effective_start > effective_end:
        effective_start = None
        effective_end = None

    return EffectiveDesignationRange(
        requested_start=requested_start,
        requested_end=requested_end,
        start=effective_start,
        end=effective_end,
        is_explicit=is_explicit,
    )


def _parse_schedule_time(raw: object) -> time | None:
    if isinstance(raw, time):
        return raw
    if not isinstance(raw, str):
        return None
    try:
        hour, minute = raw.strip().split(":", 1)
        return time(int(hour), int(minute))
    except (TypeError, ValueError):
        return None


def _expand_schedule_to_slots(
    schedule_json: list[dict],
    start_date: date,
    end_date: date,
    *,
    designation_id: int | None = None,
) -> list[ScheduledSlot]:
    """Strictly expand a designation schedule into canonical dated slots."""
    _resolve_payroll_period(start_date.month, start_date.year, start_date, end_date)
    slots_by_weekday: dict[str, list[tuple[time, time, int]]] = {}
    invalid: list[str] = []
    for index, slot in enumerate(schedule_json or []):
        weekday = _normalize_day(slot.get("dia", ""))
        slot_start = _parse_schedule_time(slot.get("hora_inicio"))
        slot_end = _parse_schedule_time(slot.get("hora_fin"))
        try:
            hours = int(slot.get("horas_academicas", 0) or 0)
        except (TypeError, ValueError):
            hours = 0
        if weekday not in WEEKDAY_MAP.values() or slot_start is None or slot_end is None or hours <= 0:
            invalid.append(f"slot {index + 1}: {slot!r}")
            continue
        if slot_start >= slot_end:
            invalid.append(f"slot {index + 1}: hora_inicio debe ser anterior a hora_fin")
            continue
        slots_by_weekday.setdefault(weekday, []).append((slot_start, slot_end, hours))

    if invalid:
        label = f" de la designación {designation_id}" if designation_id is not None else ""
        raise PayrollDataError(
            f"El horario{label} contiene datos inválidos; corregilos antes de generar la planilla",
            code="invalid_schedule",
            sample=invalid[:10],
        )

    result: list[ScheduledSlot] = []
    current = start_date
    while current <= end_date:
        weekday = WEEKDAY_MAP[current.weekday()]
        for slot_start, slot_end, hours in slots_by_weekday.get(weekday, []):
            result.append(ScheduledSlot(current, slot_start, slot_end, hours))
        current += timedelta(days=1)

    keys = [(slot.slot_date, slot.scheduled_start) for slot in result]
    duplicates = [key for key, count in Counter(keys).items() if count > 1]
    if duplicates:
        label = f"La designación {designation_id}" if designation_id is not None else "El horario"
        raise PayrollDataError(
            f"{label} contiene slots duplicados con la misma fecha y hora de inicio",
            code="duplicate_schedule_slot",
            sample=[
                f"{day.isoformat()} {slot_start.strftime('%H:%M')}"
                for day, slot_start in duplicates[:10]
            ],
        )
    return result


def _calculate_designation_base_hours(
    designation: Designation,
    effective_range: EffectiveDesignationRange,
    excluded_days: "list[ExcludedDaySchema]",
) -> int:
    """Calculate payable base before absences using contract and exclusion rules."""
    if effective_range.start is None or effective_range.end is None:
        return 0

    schedule = designation.schedule_json or []
    monthly_hours = int(designation.monthly_hours or 0)
    if monthly_hours > 0 and not schedule:
        raise PayrollDataError(
            f"La designación {designation.id} tiene {monthly_hours} horas mensuales pero no tiene horario",
            code="missing_schedule",
        )

    exact_calendar = effective_range.is_explicit or effective_range.is_contract_limited
    if exact_calendar:
        slots = _expand_schedule_to_slots(
            schedule,
            effective_range.start,
            effective_range.end,
            designation_id=designation.id,
        )
        return sum(
            slot.academic_hours
            for slot in slots
            if not _is_excluded(
                slot.slot_date,
                designation.semester,
                designation.subject,
                designation.group_code,
                excluded_days,
            )
        )

    excluded_scheduled_hours = 0
    if excluded_days:
        slots = _expand_schedule_to_slots(
            schedule,
            effective_range.start,
            effective_range.end,
            designation_id=designation.id,
        )
        excluded_scheduled_hours = sum(
            slot.academic_hours
            for slot in slots
            if _is_excluded(
                slot.slot_date,
                designation.semester,
                designation.subject,
                designation.group_code,
                excluded_days,
            )
        )
    return max(0, monthly_hours - excluded_scheduled_hours)


def _reconcile_daily_hours(
    daily_hours: dict[date, int],
    daily_status: dict[date, str],
    target_hours: int,
) -> None:
    """Cap visible hours chronologically so Excel day cells equal payable hours."""
    remaining = target_hours
    for slot_date in sorted(daily_hours):
        if daily_status.get(slot_date) == "EXCLUDED":
            daily_hours[slot_date] = 0
            continue
        visible = max(0, int(daily_hours[slot_date] or 0))
        daily_hours[slot_date] = min(visible, remaining)
        remaining -= daily_hours[slot_date]
    if remaining:
        raise PayrollDataError(
            f"El horario visible solo representa {target_hours - remaining}h, "
            f"pero el cálculo pagable exige {target_hours}h",
            code="schedule_hours_mismatch",
        )


def _index_schedule_by_weekday(schedule_json: list[dict]) -> dict[str, int]:
    """Index schedule_json slots by normalized weekday name → total academic hours.

    A teacher may have multiple slots on the same weekday (e.g. morning +
    afternoon); their hours are summed.  Invalid ``horas_academicas`` values
    are logged and treated as 0.

    Returns an empty dict when schedule_json is empty/None or no valid
    slots are found.
    """
    if not schedule_json:
        return {}

    hours_by_weekday: dict[str, int] = {}
    for slot in schedule_json:
        dia = _normalize_day(slot.get("dia", ""))
        try:
            hrs = int(slot.get("horas_academicas", 0) or 0)
        except (ValueError, TypeError):
            logger.warning(
                "_index_schedule_by_weekday: invalid horas_academicas=%r in slot %r — treating as 0",
                slot.get("horas_academicas"), slot,
            )
            hrs = 0
        if dia and hrs > 0:
            hours_by_weekday[dia] = hours_by_weekday.get(dia, 0) + hrs

    return hours_by_weekday


def _expand_schedule_to_daily(
    schedule_json: list[dict],
    start_date: date,
    end_date: date,
) -> dict[date, int]:
    """Expand schedule slots into a per-date hour map for a date range.

    Returns ``{date: academic_hours}`` for every day in ``[start_date,
    end_date]`` whose weekday matches a slot in ``schedule_json``.
    Days without a matching slot are omitted (not zeroed).

    This is the **single source of truth** used by both
    ``_calculate_period_hours`` (scalar sum) and the daily-hours
    fill in ``_build_row`` (per-cell display).

    Raises:
        ValueError: if start_date > end_date or the range exceeds 62 days.
    """
    if start_date > end_date:
        raise ValueError(f"start_date ({start_date}) > end_date ({end_date})")
    if (end_date - start_date).days > 62:
        raise ValueError(f"Period too large: {(end_date - start_date).days + 1} days (max 62)")

    hours_by_weekday = _index_schedule_by_weekday(schedule_json)
    if not hours_by_weekday:
        return {}

    result: dict[date, int] = {}
    num_days = (end_date - start_date).days + 1
    for i in range(num_days):
        current = start_date + timedelta(days=i)
        weekday_name = WEEKDAY_MAP.get(current.weekday(), "")
        hrs = hours_by_weekday.get(weekday_name, 0)
        if hrs > 0:
            result[current] = hrs

    return result


def _calculate_period_hours(
    schedule_json: list[dict],
    start_date: date,
    end_date: date,
    semester: str = "",
    subject: str = "",
    group_code: str = "",
    excluded_days: "list[ExcludedDaySchema] | None" = None,
) -> int:
    """Count the total academic hours for a designation in a date range.

    Delegates to ``_expand_schedule_to_daily`` (single source of truth) and
    returns the scalar sum.  When ``excluded_days`` is provided, hours for
    excluded dates are subtracted from the total before returning.

    Args:
        schedule_json: The designation's schedule JSON.
        start_date:    Start of the attendance window.
        end_date:      End of the attendance window.
        semester:      Designation semester — required for exclusion scope matching.
        subject:       Designation subject — required for exclusion scope matching.
        group_code:    Designation group code — required for exclusion scope matching.
        excluded_days: Optional list of day exclusions.  Empty/None = no exclusions.

    Raises:
        ValueError: if start_date > end_date or the range exceeds 62 days.
    """
    daily = _expand_schedule_to_daily(schedule_json, start_date, end_date)
    if not excluded_days:
        return sum(daily.values())
    return sum(
        hrs for d, hrs in daily.items()
        if not _is_excluded(d, semester, subject, group_code, excluded_days)
    )


# ---------------------------------------------------------------------------
# Day-window helper
# ---------------------------------------------------------------------------


def _build_day_window(
    month: int,
    year: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> list[date]:
    """Build the list of dates that the day columns represent.

    When ``start_date`` and ``end_date`` are both provided, returns every date
    in that inclusive range — enabling cross-month windows such as
    ``Mar 21 → Apr 20`` for an April payroll with a custom cutoff.

    Otherwise falls back to days 1..N of the target month (legacy behavior).
    """
    if start_date is not None or end_date is not None:
        resolved_start, resolved_end, _ = _resolve_payroll_period(
            month, year, start_date, end_date,
        )
        num_days = (resolved_end - resolved_start).days + 1
        return [resolved_start + timedelta(days=i) for i in range(num_days)]

    # Fallback: full target month
    _, days_in_month = calendar.monthrange(year, month)
    return [date(year, month, d) for d in range(1, days_in_month + 1)]


def _build_month_blocks(
    month: int,
    year: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> list[MonthBlock]:
    """Build the list of month-blocks that describe the day-columns layout.

    Single-month (no ``start_date``/``end_date``, OR both provided but within
    the same calendar month): returns ONE block covering the target month
    with the full month as the active range.

    Cross-month ranges return one block for every intersected month, including
    middle months in ranges such as Apr 30 through Jun 30.

    A 1-column visual spacer between blocks is NOT part of the block itself;
    it is implied by the gap between ``blocks[0].col_start + blocks[0].days_in_month``
    and ``blocks[1].col_start``.

    Raises:
        ValueError: if ``start_date > end_date`` or the total span exceeds 62 days.
    """
    requested_start, requested_end, is_explicit = _resolve_payroll_period(
        month, year, start_date, end_date,
    )

    # Same-month case: ONE block covering the target calendar month fully.
    if (
        not is_explicit
        or (
            requested_start.month == requested_end.month
            and requested_start.year == requested_end.year
        )
    ):
        # When the cutoff fits in one month but differs from the target
        # (unlikely but possible), honor the provided month/year of the range.
        if is_explicit:
            blk_month = requested_start.month
            blk_year = requested_start.year
            active_start = requested_start.day
            active_end = requested_end.day
        else:
            blk_month = month
            blk_year = year
            active_start = 1
            active_end = calendar.monthrange(blk_year, blk_month)[1]

        _, days_in_month = calendar.monthrange(blk_year, blk_month)
        return [
            MonthBlock(
                month=blk_month,
                year=blk_year,
                month_name=MONTH_NAMES.get(blk_month, str(blk_month)),
                days_in_month=days_in_month,
                col_start=DAY_COL_START,
                active_start=active_start,
                active_end=active_end,
            )
        ]

    blocks: list[MonthBlock] = []
    block_year = requested_start.year
    block_month = requested_start.month
    col_start = DAY_COL_START
    while (block_year, block_month) <= (requested_end.year, requested_end.month):
        days_in_month = calendar.monthrange(block_year, block_month)[1]
        is_first = (block_year, block_month) == (requested_start.year, requested_start.month)
        is_last = (block_year, block_month) == (requested_end.year, requested_end.month)
        blocks.append(
            MonthBlock(
                month=block_month,
                year=block_year,
                month_name=MONTH_NAMES.get(block_month, str(block_month)),
                days_in_month=days_in_month,
                col_start=col_start,
                active_start=requested_start.day if is_first else 1,
                active_end=requested_end.day if is_last else days_in_month,
            )
        )
        col_start += days_in_month + 1
        if block_month == 12:
            block_month = 1
            block_year += 1
        else:
            block_month += 1
    return blocks


# ---------------------------------------------------------------------------
# Main generator class
# ---------------------------------------------------------------------------


class PlanillaGenerator:
    """
    Generates the monthly teacher payroll Excel file.

    Usage::

        gen = PlanillaGenerator(output_dir="backend/data/output")
        result = gen.generate(db, month=3, year=2026)
        print(result.file_path)
    """

    def __init__(self, output_dir: str = "backend/data/output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate(
        self,
        db: Session,
        month: int,
        year: int,
        payment_overrides: Optional[dict[str, float]] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        discount_mode: str = "attendance",
        excluded_days: "list[ExcludedDaySchema] | None" = None,
    ) -> PlanillaResult:
        """
        Generate the planilla Excel for a given month/year.

        Steps:
          1. Build PlanillaRow list from attendance_records + designations + teachers.
          2. Create the Excel workbook (Planilla + Detalle sheets).
          3. Save to file.
          4. Persist/update PlanillaOutput record in DB.
          5. Return PlanillaResult.

        Args:
            db: SQLAlchemy session
            month: Month number (1–12)
            year: Calendar year
            payment_overrides: Optional {"teacher_ci:designation_id": override_amount}
                and/or {teacher_ci: teacher_total_override} for admin adjustments
            start_date: Optional start of attendance window to filter records
            end_date: Optional end of attendance window to filter records
            discount_mode: "attendance" (apply discounts) or "full" (no discounts,
                all teachers receive full assigned hours)
            excluded_days: Optional list of days to exclude from the planilla.
                Empty/None means backward-compatible behavior (no exclusions).

        Returns:
            PlanillaResult with file path and statistics
        """
        if payment_overrides is None:
            payment_overrides = {}
        payment_overrides = normalize_payment_overrides(payment_overrides)
        if excluded_days is None:
            excluded_days = []

        logger.info(
            "PlanillaGenerator.generate: month=%d year=%d start=%s end=%s discount_mode=%s excluded_days=%d",
            month, year, start_date, end_date, discount_mode, len(excluded_days),
        )

        # Step 1: Build data
        rows, detail_rows, warnings = self._build_planilla_data(
            db, month, year, start_date=start_date, end_date=end_date,
            discount_mode=discount_mode, excluded_days=excluded_days,
        )
        validate_payment_override_targets(rows, payment_overrides)
        logger.info("Built %d planilla rows with %d detail slots", len(rows), len(detail_rows))

        # Step 2: Create workbook.
        # Build month-blocks once so headers, data, and totals stay aligned
        # with the attendance window (1 block same-month, 2 blocks cross-month).
        blocks = _build_month_blocks(month, year, start_date, end_date)
        wb = self._create_workbook(
            rows, detail_rows, month, year, payment_overrides, blocks
        )

        # Step 3: Save file
        month_name = MONTH_NAMES.get(month, str(month)).upper()
        filename = f"planilla_{month:02d}_{year}.xlsx"
        file_path = self.output_dir / filename
        wb.save(str(file_path))
        logger.info("Saved planilla to %s", file_path)

        # Step 4: Persist to DB — Model C: total payable hours
        total_hours = sum(r.payable_hours for r in rows)
        row_amounts = []
        for row in rows:
            override = self._get_row_override(row, payment_overrides, rows)
            row_amounts.append(override if override is not None else row.final_payment)
        snapshot = build_calculation_snapshot(
            rows=rows,
            row_amounts=row_amounts,
            month=month,
            year=year,
            start_date=start_date,
            end_date=end_date,
            discount_mode=discount_mode,
            payment_overrides=payment_overrides,
            excluded_days=excluded_days,
        )
        total_payment = Decimal(snapshot["total"])
        unique_teachers = len({r.teacher_ci for r in rows})

        planilla_output = self._persist_planilla_output(
            db=db,
            month=month,
            year=year,
            file_path=str(file_path),
            total_teachers=unique_teachers,
            total_hours=total_hours,
            total_payment=total_payment,
            payment_overrides=payment_overrides,
            start_date=start_date,
            end_date=end_date,
            discount_mode=discount_mode,
            excluded_days=excluded_days,
            calculation_snapshot=snapshot,
        )

        return PlanillaResult(
            file_path=str(file_path),
            month=month,
            year=year,
            total_teachers=unique_teachers,
            total_rows=len(rows),
            total_hours=total_hours,
            total_payment=total_payment,
            planilla_output_id=planilla_output.id if planilla_output else None,
            warnings=warnings,
            discount_mode=discount_mode,
        )

    # ------------------------------------------------------------------
    # Data building
    # ------------------------------------------------------------------

    def _build_planilla_data(
        self,
        db: Session,
        month: int,
        year: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        discount_mode: str = "attendance",
        excluded_days: "list[ExcludedDaySchema] | None" = None,
    ) -> tuple[list[PlanillaRow], list[DetailRow], list[str]]:
        """Build canonical payroll rows shared by every regular-planilla consumer."""
        warnings: list[str] = []
        exclusions = excluded_days or []
        hourly_rate = app_settings_service.get_hourly_rate(db)
        active_period = app_settings_service.get_active_academic_period(db)
        all_designations: list[Designation] = (
            db.query(Designation)
            .filter(
                Designation.academic_period == active_period,
                Designation.designation_type != "practice",
            )
            .all()
        )

        if not all_designations:
            warnings.append("No hay designaciones en la base de datos")
            return [], [], warnings

        period_start, period_end, _ = _resolve_payroll_period(
            month, year, start_date, end_date,
        )
        effective_ranges = {
            desig.id: _effective_designation_range(
                desig, month, year, start_date, end_date,
            )
            for desig in all_designations
        }

        if discount_mode == "full":
            logger.info("discount_mode=full — skipping regular attendance queries")
            att_index: dict[tuple[str, int], list[AttendanceRecord]] = {}
        else:
            designation_ids = [desig.id for desig in all_designations]
            att_query = db.query(AttendanceRecord).filter(
                AttendanceRecord.designation_id.in_(designation_ids),
                AttendanceRecord.date >= period_start,
                AttendanceRecord.date <= period_end,
            )
            att_records: list[AttendanceRecord] = (
                att_query
                .order_by(
                    AttendanceRecord.designation_id,
                    AttendanceRecord.date,
                    AttendanceRecord.scheduled_start,
                )
                .all()
            )
            self._validate_attendance_coverage(
                all_designations,
                att_records,
                effective_ranges,
                exclusions,
            )
            att_index = {}
            for rec in att_records:
                key = (rec.teacher_ci, rec.designation_id)
                att_index.setdefault(key, []).append(rec)

        all_teacher_cis = {d.teacher_ci for d in all_designations}
        teachers: dict[str, Teacher] = {
            t.ci: t
            for t in db.query(Teacher).filter(Teacher.ci.in_(all_teacher_cis)).all()
        }

        planilla_rows: list[PlanillaRow] = []
        detail_rows: list[DetailRow] = []
        for desig in all_designations:
            ci = desig.teacher_ci
            teacher = teachers.get(ci)
            if teacher is None:
                raise PayrollDataError(
                    f"La designación {desig.id} referencia al docente CI {ci}, que no existe",
                    code="missing_teacher",
                )

            key = (ci, desig.id)
            effective_range = effective_ranges[desig.id]
            records = [
                record
                for record in att_index.get(key, [])
                if effective_range.start is not None
                and effective_range.end is not None
                and effective_range.start <= record.date <= effective_range.end
                and not _is_excluded(
                    record.date,
                    desig.semester,
                    desig.subject,
                    desig.group_code,
                    exclusions,
                )
            ]

            row = self._build_row(
                teacher,
                desig,
                records,
                has_biometric=bool(records),
                discount_mode=discount_mode,
                hourly_rate=hourly_rate,
                start_date=start_date,
                end_date=end_date,
                month=month,
                year=year,
                excluded_days=exclusions,
                effective_range=effective_range,
            )
            planilla_rows.append(row)

            for rec in records:
                detail_rows.append(
                    DetailRow(
                        ci=ci,
                        teacher_name=teacher.full_name,
                        date=rec.date,
                        day_letter=WEEKDAY_LETTERS[rec.date.weekday()],
                        subject=desig.subject,
                        group_code=desig.group_code,
                        semester=desig.semester,
                        scheduled_start=rec.scheduled_start.strftime("%H:%M"),
                        scheduled_end=rec.scheduled_end.strftime("%H:%M"),
                        academic_hours=rec.academic_hours,
                        status=rec.status,
                        observation=rec.observation,
                    )
                )

        planilla_rows.sort(key=lambda r: (r.teacher_name, r.subject, r.group_code))
        detail_rows.sort(key=lambda r: (r.teacher_name, r.date, r.scheduled_start))
        logger.info(
            "_build_planilla_data: %d rows, %d detail records",
            len(planilla_rows),
            len(detail_rows),
        )
        return planilla_rows, detail_rows, warnings

    def _validate_attendance_coverage(
        self,
        designations: list[Designation],
        records: list[AttendanceRecord],
        effective_ranges: dict[int, EffectiveDesignationRange],
        excluded_days: "list[ExcludedDaySchema]",
    ) -> None:
        records_by_designation: dict[int, list[AttendanceRecord]] = {}
        for record in records:
            records_by_designation.setdefault(record.designation_id, []).append(record)

        issues: list[str] = []
        missing_count = duplicate_count = unexpected_count = invalid_count = 0
        valid_statuses = {"ATTENDED", "LATE", "ABSENT", "NO_EXIT"}
        for desig in designations:
            effective_range = effective_ranges[desig.id]
            if effective_range.start is None or effective_range.end is None:
                continue
            if int(desig.monthly_hours or 0) > 0 and not (desig.schedule_json or []):
                raise PayrollDataError(
                    f"La designación {desig.id} tiene horas mensuales pero no tiene horario",
                    code="missing_schedule",
                )
            slots = _expand_schedule_to_slots(
                desig.schedule_json or [],
                effective_range.start,
                effective_range.end,
                designation_id=desig.id,
            )
            expected = {
                (desig.teacher_ci, desig.id, slot.slot_date, slot.scheduled_start): slot
                for slot in slots
                if not _is_excluded(
                    slot.slot_date,
                    desig.semester,
                    desig.subject,
                    desig.group_code,
                    excluded_days,
                )
            }
            relevant = [
                record
                for record in records_by_designation.get(desig.id, [])
                if effective_range.start <= record.date <= effective_range.end
                and not _is_excluded(
                    record.date,
                    desig.semester,
                    desig.subject,
                    desig.group_code,
                    excluded_days,
                )
            ]
            actual: dict[tuple[str, int, date, time], list[AttendanceRecord]] = {}
            for record in relevant:
                key = (record.teacher_ci, record.designation_id, record.date, record.scheduled_start)
                actual.setdefault(key, []).append(record)

            for key, slot in expected.items():
                matches = actual.get(key, [])
                label = (
                    f"designación {desig.id}, CI {desig.teacher_ci}, "
                    f"{slot.slot_date.isoformat()} {slot.scheduled_start.strftime('%H:%M')}"
                )
                if not matches:
                    missing_count += 1
                    issues.append(f"Falta AttendanceRecord: {label}")
                    continue
                if len(matches) > 1:
                    duplicate_count += len(matches) - 1
                    issues.append(f"AttendanceRecord duplicado: {label}")
                    continue
                record = matches[0]
                status_value = (record.status or "").upper()
                expected_hours = 0 if status_value == "ABSENT" else slot.academic_hours
                if (
                    status_value not in valid_statuses
                    or record.scheduled_end != slot.scheduled_end
                    or record.academic_hours != expected_hours
                ):
                    invalid_count += 1
                    issues.append(
                        f"AttendanceRecord incoherente: {label}, estado={record.status!r}, "
                        f"fin={record.scheduled_end}, horas={record.academic_hours}; "
                        f"esperado fin={slot.scheduled_end}, horas={expected_hours}"
                    )

            unexpected = set(actual) - set(expected)
            unexpected_count += len(unexpected)
            issues.extend(
                "AttendanceRecord sin slot programado: "
                f"designación {key[1]}, CI {key[0]}, "
                f"{key[2].isoformat()} {key[3].strftime('%H:%M')}"
                for key in sorted(unexpected, key=lambda value: (value[1], value[2], value[3]))
            )

        if issues:
            raise PayrollDataError(
                "La cobertura regular está incompleta o contiene datos incoherentes "
                f"(faltantes={missing_count}, duplicados={duplicate_count}, "
                f"sobrantes={unexpected_count}, inválidos={invalid_count}). "
                "Reprocesá la asistencia antes de generar la planilla",
                code="regular_attendance_coverage",
                sample=issues[:10],
            )

    def _build_row(
        self,
        teacher: Teacher,
        desig: Designation,
        records: list[AttendanceRecord],
        has_biometric: bool = True,
        discount_mode: str = "attendance",
        hourly_rate: float = RATE_PER_HOUR,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        month: int = 0,
        year: int = 0,
        excluded_days: "list[ExcludedDaySchema] | None" = None,
        effective_range: EffectiveDesignationRange | None = None,
    ) -> PlanillaRow:
        """
        Build a single PlanillaRow using Payment Model C.

        Model C rules:
          - Base pay = designation.monthly_hours × hourly_rate (from app_settings)
          - Deduct ONLY hours where status=ABSENT
          - No biometric at all → full pay (has_biometric=False)
          - discount_mode="full" → full pay for ALL teachers (no deductions)
          - attended/late/no_exit hours in daily_hours are for display only
          - excluded_days → set daily_hours[d]=0, daily_status[d]="EXCLUDED",
            reduce base_monthly_hours (but do NOT count as absent_hours)

        ``hourly_rate`` is injected from the caller (defaults to the legacy
        ``RATE_PER_HOUR`` constant so tests that construct rows directly keep
        working).
        """
        if excluded_days is None:
            excluded_days = []
        daily_hours: dict[date, int] = {}
        daily_status: dict[date, str] = {}
        attended_hours = 0   # for informational display only
        absent_hours = 0     # hours to deduct (Model C)
        absent_hours_by_day: dict[date, int] = {}  # track per-day for exclusion undo
        late_count = 0
        absent_count = 0
        absent_count_by_day: dict[date, int] = {}  # track per-day for exclusion undo
        observations: list[str] = []

        if effective_range is None and (month and year or start_date is not None or end_date is not None):
            range_month = month or (start_date.month if start_date is not None else 0)
            range_year = year or (start_date.year if start_date is not None else 0)
            effective_range = _effective_designation_range(
                desig, range_month, range_year, start_date, end_date,
            )

        for rec in records:
            day = rec.date   # full date — supports cross-month day windows
            hours = rec.academic_hours
            status = rec.status.upper()

            # Accumulate hours for the day (could have multiple slots on same day)
            # For absent slots, academic_hours=0 per engine logic, but we count
            # the SCHEDULED hours via designation — here we show 0 for absent days
            daily_hours[day] = daily_hours.get(day, 0) + hours

            # Track worst status for coloring: ABSENT > LATE > NO_EXIT > ATTENDED
            current_status = daily_status.get(day, "")
            if status == "ABSENT":
                daily_status[day] = "ABSENT"
                absent_count += 1
                absent_count_by_day[day] = absent_count_by_day.get(day, 0) + 1
                # Absent slots have academic_hours=0 in the record; we need to
                # count the slot hours from the designation schedule for deduction.
                # The engine sets academic_hours=0 for ABSENT — we must compute
                # absent_hours from the slot's scheduled hours in schedule_json.
                slot_hrs = self._get_slot_hours(desig, rec)
                absent_hours += slot_hrs
                absent_hours_by_day[day] = absent_hours_by_day.get(day, 0) + slot_hrs
            elif status == "LATE" and current_status != "ABSENT":
                daily_status[day] = "LATE"
                late_count += 1
                attended_hours += hours
            elif status == "NO_EXIT" and current_status not in ("ABSENT", "LATE"):
                daily_status[day] = "NO_EXIT"
                attended_hours += hours
            elif status == "ATTENDED" and not current_status:
                daily_status[day] = "ATTENDED"
                attended_hours += hours

            if rec.observation:
                observations.append(f"Día {day.day}/{day.month}: {rec.observation}")

        # ── Fill daily_hours gaps from schedule ─────────────────────────
        # For any scheduled day in the period that has NO attendance record
        # (either because discount_mode="full" skips biometric, the teacher
        # has no biometric, or biometric only covers part of a cross-month
        # window), fill from schedule_json so the Excel shows expected hours.
        #
        # Uses the SAME shared helper as _calculate_period_hours to guarantee
        # consistency — no duplicated logic.
        #
        # Key rules:
        #   - Only fills dates NOT already in daily_hours (won't overwrite real attendance)
        #   - Does NOT set daily_status (no false "ATTENDED" claim without biometric)
        #   - Does NOT bump attended_hours (these are scheduled, not verified)
        #   - Works for both cross-month (start/end provided) and single-month (derived)
        if desig.schedule_json:
            eff_start = effective_range.start if effective_range else None
            eff_end = effective_range.end if effective_range else None

            if eff_start and eff_end:
                scheduled_daily: dict[date, int] = {}
                for slot in _expand_schedule_to_slots(
                    desig.schedule_json,
                    eff_start,
                    eff_end,
                    designation_id=desig.id,
                ):
                    scheduled_daily[slot.slot_date] = (
                        scheduled_daily.get(slot.slot_date, 0) + slot.academic_hours
                    )
                for d, hrs in scheduled_daily.items():
                    if d not in daily_hours:
                        daily_hours[d] = hrs
                        # No daily_status set — cell gets generic "class day" color
                        # via the hours > 0 fallback in _write_data_row.
                        # No attended_hours bump — these are scheduled, not verified.

        # ── Apply day exclusions ─────────────────────────────────────────
        # Exclusions are applied AFTER schedule expansion so that any day
        # that has scheduled hours (whether from attendance or schedule fill)
        # is properly zeroed and flagged. Excluded hours are NOT added to
        # absent_hours — they represent admin decisions, not attendance failures.
        # If a day was already marked ABSENT, we must REMOVE its hours from
        # absent_hours to prevent double-deduction (base_monthly_hours already
        # excludes the day, so deducting absent_hours again would subtract twice).
        if excluded_days:
            for d in list(daily_hours.keys()):
                if _is_excluded(d, desig.semester, desig.subject, desig.group_code, excluded_days):
                    if daily_status.get(d) == "ABSENT" and d in absent_hours_by_day:
                        # Undo the absent deduction for this excluded day
                        absent_hours = max(0, absent_hours - absent_hours_by_day[d])
                        absent_count = max(0, absent_count - absent_count_by_day.get(d, 0))
                    daily_hours[d] = 0
                    daily_status[d] = "EXCLUDED"

        if effective_range is not None:
            base_monthly_hours = _calculate_designation_base_hours(
                desig, effective_range, excluded_days,
            )
        else:
            base_monthly_hours = int(desig.monthly_hours or 0)

        # discount_mode="full" → override: pay full assigned hours, zero deductions
        if discount_mode == "full":
            absent_hours = 0
            absent_count = 0

        if absent_hours > base_monthly_hours and base_monthly_hours > 0:
            logger.warning(
                "Designation %d (CI=%s): absent_hours=%d exceeds base_monthly_hours=%d "
                "— possible duplicate attendance records or schedule mismatch. "
                "Capping absent_hours at base_monthly_hours.",
                desig.id, teacher.ci, absent_hours, base_monthly_hours,
            )
            absent_hours = base_monthly_hours

        payable_hours = max(0, base_monthly_hours - absent_hours)
        if desig.schedule_json and effective_range is not None:
            _reconcile_daily_hours(daily_hours, daily_status, payable_hours)
        calculated_payment = payable_hours * hourly_rate

        # RC-IVA 13% retention
        has_retention = (teacher.invoice_retention or "").strip().upper() == "RETENCION"
        retention_rate = 0.13 if has_retention else 0.0
        retention_amount = round(calculated_payment * retention_rate, 2)
        final_payment = round(calculated_payment - retention_amount, 2)

        # Build observation summary.
        # "full" mode and "sin biométrico" are orthogonal facts about the row: a
        # teacher can lack biometric data AND the payroll be run in "full" mode,
        # so both observations should appear when applicable. Use independent
        # `if`s instead of `if/elif` so one doesn't mask the other.
        # Attendance-based counts only make sense when there IS biometric data
        # AND we are actually discounting — in "full" mode absent_hours/count
        # are zeroed out above, so those checks are naturally no-ops there.
        obs_parts: list[str] = []
        if discount_mode == "full":
            obs_parts.append("Modo sin descuentos — pago completo")
        if has_biometric and discount_mode != "full":
            if late_count > 0:
                obs_parts.append(f"{late_count} tardanza{'s' if late_count > 1 else ''}")
            if absent_count > 0:
                obs_parts.append(f"{absent_count} ausencia{'s' if absent_count > 1 else ''}")
            if absent_hours > 0:
                obs_parts.append(f"{absent_hours}h descontadas")
        if has_retention:
            obs_parts.append("Retención RC-IVA 13%")

        return PlanillaRow(
            teacher_ci=teacher.ci,
            designation_id=desig.id,
            teacher_name=teacher.full_name,
            email=teacher.email,
            phone=teacher.phone,
            subject=desig.subject,
            semester=desig.semester,
            group_code=desig.group_code,
            teacher_type=teacher.external_permanent,
            gender=teacher.gender,
            sap_code=teacher.sap_code,
            invoice_retention=teacher.invoice_retention,
            account_number=teacher.account_number,
            academic_level=teacher.academic_level,
            profession=teacher.profession,
            specialty=teacher.specialty,
            bank=teacher.bank,
            nit=teacher.nit,
            daily_hours=daily_hours,
            daily_status=daily_status,
            # Model C fields
            base_monthly_hours=base_monthly_hours,
            absent_hours=absent_hours,
            payable_hours=payable_hours,
            # Legacy totals mapped to Model C semantics for Excel columns
            total_hours=payable_hours,                        # used for payment sum
            total_theory_hours=base_monthly_hours,            # COL_HRS_TEORIA = assigned
            total_practice_internal_hours=absent_hours,       # COL_HRS_PRACT_INT = deducted
            total_practice_external_hours=attended_hours,     # COL_HRS_PRACT_EXT = attended (info)
            rate_per_hour=hourly_rate,
            calculated_payment=calculated_payment,
            has_retention=has_retention,
            retention_rate=retention_rate,
            retention_amount=retention_amount,
            final_payment=final_payment,
            observations=obs_parts if obs_parts else [],
            late_count=late_count,
            absent_count=absent_count,
            has_biometric=has_biometric,
        )

    def _get_slot_hours(self, desig: Designation, rec: AttendanceRecord) -> int:
        """
        Find the scheduled academic_hours for an ABSENT slot.

        The engine sets academic_hours=0 for ABSENT records; we must recover
        the scheduled hours from the designation's schedule_json by matching
        the slot's weekday + scheduled_start time (and optionally hora_fin).

        Matching priority:
          1. day-of-week + hora_inicio (most specific — avoids cross-day false match)
          2. hora_inicio only (fallback when schedule_json lacks "dia" field)

        Day matching uses _normalize_day() to handle accented variants ("miércoles" /
        "sábado") stored in schedule_json alongside the unaccented WEEKDAY_MAP values.
        """
        schedule: list[dict] = desig.schedule_json or []
        rec_start_str = rec.scheduled_start.strftime("%H:%M")

        # Use the canonical weekday map and day-name normalizer (single source of truth)
        target_weekday = rec.date.weekday()
        target_day_norm = WEEKDAY_MAP.get(target_weekday, "")

        # Pass 1: match by weekday + hora_inicio (most accurate)
        for slot in schedule:
            if slot.get("hora_inicio", "") == rec_start_str:
                slot_dia_norm = _normalize_day(slot.get("dia", ""))
                if slot_dia_norm == target_day_norm:
                    return int(slot.get("horas_academicas", 0))

        # Pass 2: fallback — match by hora_inicio only (slot may lack "dia")
        for slot in schedule:
            if slot.get("hora_inicio", "") == rec_start_str:
                return int(slot.get("horas_academicas", 0))

        logger.debug(
            "Could not find schedule slot for absent record (designation=%d, start=%s, day=%s)",
            desig.id,
            rec_start_str,
            target_day_norm,
        )
        return 0

    # ------------------------------------------------------------------
    # Workbook creation
    # ------------------------------------------------------------------

    def _get_summary_cols(self, blocks: list[MonthBlock]) -> dict[str, int]:
        """Compute summary column positions based on the block layout.

        Summary columns live immediately after the LAST day-block. For a
        single-month layout this is identical to the legacy behavior
        (``DAY_COL_START + 31 = 48``). For cross-month it shifts to the right
        because of block 0 + 1 spacer + block 1.

        Returns a dict with keys matching the module-level ``COL_*`` constants
        (minus the ``COL_`` prefix, lowercased) so call sites can look up the
        current column for each summary field.
        """
        last_block = blocks[-1]
        # First col after the last day-block (last_block.col_start is day 1,
        # so last day sits at col_start + days_in_month - 1).
        base = last_block.col_start + last_block.days_in_month
        return {
            'total_horas': base,
            'grupo_resumen': base + 1,
            'materia_resumen': base + 2,
            'pago_hora': base + 3,
            'hrs_teoria': base + 4,
            'hrs_pract_int': base + 5,
            'hrs_pract_ext': base + 6,
            'total_hrs_check': base + 7,
            'pago_calculado': base + 8,
            'retencion_amt': base + 9,
            'pago_neto': base + 10,
            'pago_ajustado': base + 11,
            'observaciones': base + 12,
        }

    def _create_workbook(
        self,
        rows: list[PlanillaRow],
        detail_rows: list[DetailRow],
        month: int,
        year: int,
        payment_overrides: dict[str, float],
        blocks: list[MonthBlock],
    ) -> Workbook:
        """Create the complete Excel workbook with Planilla + Detalle sheets."""
        wb = Workbook()

        # Sheet 1: Planilla principal
        # wb.active is always non-None on a freshly created Workbook
        ws = wb.create_sheet(title="Planilla", index=0)
        # Remove the default empty sheet that Workbook() creates
        if len(wb.worksheets) > 1:
            default_sheet = wb.worksheets[1]
            del wb[default_sheet.title]

        scols = self._get_summary_cols(blocks)
        # 13 summary columns → last column = observaciones
        total_cols = scols['observaciones']

        self._write_headers(ws, month, year, blocks, scols, total_cols)
        last_data_row = self._write_data_rows(
            ws, rows, month, year, payment_overrides, blocks, scols
        )
        self._write_totals_row(
            ws, rows, last_data_row + 1, payment_overrides,
            blocks, scols, total_cols,
        )
        self._apply_formatting(ws, last_data_row + 1, month, year, total_cols)

        # Sheet 2: Detalle granular
        ws_detail = wb.create_sheet(title="Detalle")
        self._write_detail_sheet(ws_detail, detail_rows, month, year)

        return wb

    # ------------------------------------------------------------------
    # Header writing
    # ------------------------------------------------------------------

    def _write_headers(
        self,
        ws,
        month: int,
        year: int,
        blocks: list[MonthBlock],
        scols: dict[str, int],
        total_cols: int,
    ) -> None:
        """
        Write rows 1–6: title, university, empty, section headers, col headers, weekday row.
        Also sets column widths.
        """
        month_name = MONTH_NAMES.get(month, str(month)).upper()
        last_col_letter = get_column_letter(total_cols)

        # ── Row 1: Title ───────────────────────────────────────────────
        ws.merge_cells(f"A{ROW_TITLE}:{last_col_letter}{ROW_TITLE}")
        title_cell = ws.cell(row=ROW_TITLE, column=1)
        title_cell.value = f"SIPAD — PLANILLA DOCENTE MEDICINA — {month_name} {year}"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color=COLOR_WHITE)
        title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ROW_TITLE].height = 24

        # ── Row 2: University ──────────────────────────────────────────
        ws.merge_cells(f"A{ROW_UNIVERSITY}:{last_col_letter}{ROW_UNIVERSITY}")
        univ_cell = ws.cell(row=ROW_UNIVERSITY, column=1)
        univ_cell.value = "Universidad Privada Domingo Savio — Facultad de Medicina"
        univ_cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE)
        univ_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        univ_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ROW_UNIVERSITY].height = 18

        # ── Row 3: Empty (spacer) ──────────────────────────────────────
        ws.row_dimensions[ROW_EMPTY].height = 6

        # ── Row 4: Section headers ─────────────────────────────────────
        self._write_section_headers(ws, month, month_name, year, blocks, scols)

        # ── Row 5: Column headers (day numbers + identity names) ───────
        self._write_column_headers(ws, month, year, blocks, scols)

        # ── Row 6: Weekday letters under day columns ───────────────────
        self._write_weekday_row(ws, month, year, blocks, scols, total_cols)

        # ── Column widths ──────────────────────────────────────────────
        self._set_column_widths(ws, month, year, blocks, scols)

    def _write_section_headers(
        self,
        ws,
        month: int,
        month_name: str,
        year: int,
        blocks: list[MonthBlock],
        scols: dict[str, int],
    ) -> None:
        """Row 4: Merged section labels.

        Single-month: one merged "ASISTENCIA {MONTH} {YEAR}" label spanning the
        block's day columns — identical to legacy behavior.

        Cross-month: one merged section header PER block. The spacer column
        between blocks gets a neutral styled cell (no value). Each block's
        label includes the active cutoff hint ("desde día X" / "hasta día Y").
        """
        row = ROW_SECTION_HEADERS

        # DATOS DOCENTE (cols A–P) — same in both modes
        ws.merge_cells(
            start_row=row, start_column=COL_SEMESTRE,
            end_row=row, end_column=COL_BANCO
        )
        cell = ws.cell(row=row, column=COL_SEMESTRE)
        cell.value = "DATOS DOCENTE"
        self._style_section_header(cell)

        # ASISTENCIA — one merge per block
        is_cross_month = len(blocks) > 1
        for idx, block in enumerate(blocks):
            start_col = block.col_start
            end_col = block.col_start + block.days_in_month - 1
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col,
            )
            cell = ws.cell(row=row, column=start_col)
            bname = block.month_name.upper()
            if not is_cross_month:
                cell.value = f"ASISTENCIA {bname} {block.year}"
            elif idx == 0:
                # First block: "desde día X"
                cell.value = (
                    f"ASISTENCIA {bname} {block.year} (desde día {block.active_start})"
                )
            else:
                # Last block: "hasta día Y"
                cell.value = (
                    f"ASISTENCIA {bname} {block.year} (hasta día {block.active_end})"
                )
            self._style_section_header(cell)

            # Style the spacer column (only present between blocks)
            if idx < len(blocks) - 1:
                spacer_col = end_col + 1
                spacer_cell = ws.cell(row=row, column=spacer_col)
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        # RESUMEN Y PAGOS — shifts based on block layout via scols
        ws.merge_cells(
            start_row=row, start_column=scols['total_horas'],
            end_row=row, end_column=scols['observaciones']
        )
        cell = ws.cell(row=row, column=scols['total_horas'])
        cell.value = "RESUMEN Y PAGOS"
        self._style_section_header(cell)

        ws.row_dimensions[row].height = 20

    def _write_column_headers(
        self,
        ws,
        month: int,
        year: int,
        blocks: list[MonthBlock],
        scols: dict[str, int],
    ) -> None:
        """Row 5: Actual column headers including day numbers.

        Iterates each block and writes day numbers 1..days_in_month.
        Days outside the block's active cutoff range get a muted style
        (gray bg, lighter font) to signal they're outside the window.
        The spacer column between blocks (cross-month) gets a neutral cell.
        """
        row = ROW_COL_HEADERS

        identity_headers = [
            "Semestre",
            "Apellidos y Nombres",
            "CI",
            "Correo Electrónico",
            "Nro. Celular",
            "Materia",
            "Grupo",
            "Tipo Docente",
            "Género",
            "Código SAP",
            "Factura/Retención",
            "Nro. Cuenta",
            "Nivel Académico",
            "Profesión",
            "Especialidad",
            "Banco",
        ]

        for i, header in enumerate(identity_headers, start=1):
            cell = ws.cell(row=row, column=i)
            cell.value = header
            self._style_col_header(cell)

        # Day number headers — iterate each block. Days outside active cutoff
        # range are rendered with a muted/gray style.
        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                cell = ws.cell(row=row, column=col)
                cell.value = day
                is_active = block.active_start <= day <= block.active_end
                if is_active:
                    self._style_col_header(cell, is_day=True)
                else:
                    # Outside cutoff — muted header style
                    cell.font = Font(
                        name="Calibri", size=9, bold=False, color="A6A6A6"
                    )
                    cell.fill = PatternFill("solid", fgColor=COLOR_OUTSIDE_CUTOFF)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    cell.border = THIN_BORDER

            # Spacer column styling (only between blocks)
            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                spacer_cell = ws.cell(row=row, column=spacer_col)
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        # Summary column headers — Model C semantics, positions are dynamic
        summary_headers = {
            scols['total_horas']: "Hrs\nPagables",        # payable_hours (base - absent)
            scols['grupo_resumen']: "Grupo",
            scols['materia_resumen']: "Materia",
            scols['pago_hora']: "Pago\n/Hora",
            scols['hrs_teoria']: "Hrs\nAsignadas",         # base_monthly_hours
            scols['hrs_pract_int']: "Hrs\nDescontadas",    # absent_hours (deducted)
            scols['hrs_pract_ext']: "Hrs\nAsistidas",      # attended hours (info)
            scols['total_hrs_check']: "Total\nPagable",    # payable_hours (check)
            scols['pago_calculado']: "Total Bruto\n(Bs)",
            scols['retencion_amt']: "Retención\nRC-IVA",   # 13% if has_retention
            scols['pago_neto']: "Pago Neto\n(Bs)",         # final_payment
            scols['pago_ajustado']: "Pago\nAjustado",
            scols['observaciones']: "Observaciones",
        }
        for col, header in summary_headers.items():
            cell = ws.cell(row=row, column=col)
            cell.value = header
            self._style_col_header(cell, wrap=True)

        ws.row_dimensions[row].height = 30

    def _write_weekday_row(
        self,
        ws,
        month: int,
        year: int,
        blocks: list[MonthBlock],
        scols: dict[str, int],
        total_cols: int,
    ) -> None:
        """Row 6: Day-of-week letter under each day column.

        Iterates blocks and writes weekday letters (L/M/M/J/V/S/D) for every
        day of each block. Days outside the active cutoff get a muted gray
        style. Spacer columns between blocks get a neutral fill.
        """
        row = ROW_WEEKDAY

        # Collect the set of columns occupied by day-blocks + spacers so we
        # can skip them when filling "non-day" columns with the weekday bg.
        block_cols: set[int] = set()

        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                block_cols.add(col)
                cell = ws.cell(row=row, column=col)
                target_date = date(block.year, block.month, day)
                cell.value = WEEKDAY_LETTERS[target_date.weekday()]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER

                is_active = block.active_start <= day <= block.active_end
                if is_active:
                    cell.font = Font(
                        name="Calibri", size=8, bold=True, color="595959"
                    )
                    cell.fill = PatternFill("solid", fgColor=COLOR_WEEKDAY_BG)
                else:
                    cell.font = Font(
                        name="Calibri", size=8, bold=False, color="A6A6A6"
                    )
                    cell.fill = PatternFill("solid", fgColor=COLOR_OUTSIDE_CUTOFF)

            # Spacer column (only present between blocks)
            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                block_cols.add(spacer_col)
                spacer_cell = ws.cell(row=row, column=spacer_col)
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        # Fill non-day/non-spacer columns in this row with the weekday bg.
        for col in range(1, total_cols + 1):
            if col in block_cols:
                continue
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=COLOR_WEEKDAY_BG)
            cell.border = THIN_BORDER

        ws.row_dimensions[row].height = 14

    def _set_column_widths(
        self,
        ws,
        month: int,
        year: int,
        blocks: list[MonthBlock],
        scols: dict[str, int],
    ) -> None:
        """Set optimized column widths.

        All day columns (active AND outside-cutoff) use width 3.5 because the
        header shows a plain day number 1..N. The spacer column between blocks
        uses a narrow 1.5 to visually separate months.
        """
        # Identity columns
        widths = {
            COL_SEMESTRE: 10,
            COL_NOMBRE: 28,
            COL_CI: 10,
            COL_EMAIL: 22,
            COL_PHONE: 12,
            COL_MATERIA: 28,
            COL_GRUPO: 8,
            COL_TIPO_DOCENTE: 12,
            COL_GENERO: 8,
            COL_SAP: 12,
            COL_FACTURA: 12,
            COL_CUENTA: 16,
            COL_NIVEL_ACAD: 14,
            COL_PROFESION: 16,
            COL_ESPECIALIDAD: 16,
            COL_BANCO: 14,
        }
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Day + spacer columns
        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                ws.column_dimensions[get_column_letter(col)].width = 3.5
            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                ws.column_dimensions[get_column_letter(spacer_col)].width = 1.5

        # Summary columns — positions are dynamic (scols)
        summary_widths = {
            scols['total_horas']: 8,
            scols['grupo_resumen']: 8,
            scols['materia_resumen']: 22,
            scols['pago_hora']: 8,
            scols['hrs_teoria']: 8,
            scols['hrs_pract_int']: 9,
            scols['hrs_pract_ext']: 9,
            scols['total_hrs_check']: 8,
            scols['pago_calculado']: 12,
            scols['retencion_amt']: 11,
            scols['pago_neto']: 12,
            scols['pago_ajustado']: 12,
            scols['observaciones']: 30,
        }
        for col, width in summary_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

    # ------------------------------------------------------------------
    # Data row writing
    # ------------------------------------------------------------------

    def _write_data_rows(
        self,
        ws,
        rows: list[PlanillaRow],
        month: int,
        year: int,
        payment_overrides: dict[str, float],
        blocks: list[MonthBlock],
        scols: dict[str, int],
    ) -> int:
        """Write all data rows. Returns the row number of the last written row."""
        for i, data in enumerate(rows):
            row_num = DATA_ROW_START + i
            self._write_data_row(
                ws,
                row_num,
                data,
                month,
                year,
                payment_overrides,
                rows,
                blocks,
                scols,
            )

        last_row = DATA_ROW_START + len(rows) - 1
        return last_row if rows else DATA_ROW_START - 1

    def _write_data_row(
        self,
        ws,
        row_num: int,
        data: PlanillaRow,
        month: int,
        year: int,
        payment_overrides: dict[str, float],
        all_rows: list[PlanillaRow],
        blocks: list[MonthBlock],
        scols: dict[str, int],
    ) -> None:
        """Write one teacher×designation data row."""
        override = self._get_row_override(data, payment_overrides, all_rows)

        # Alternate row background for readability
        is_even = (row_num - DATA_ROW_START) % 2 == 0
        row_bg = "FFFFFF" if is_even else "F5F8FA"

        base_font = Font(name="Calibri", size=9)
        base_align = Alignment(vertical="center", wrap_text=False)
        base_fill = PatternFill("solid", fgColor=row_bg)

        def write_identity(col: int, value) -> None:
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.font = base_font
            cell.alignment = base_align
            cell.fill = base_fill
            cell.border = THIN_BORDER

        # Identity columns
        write_identity(COL_SEMESTRE, data.semester)
        write_identity(COL_NOMBRE, data.teacher_name)
        write_identity(COL_CI, data.teacher_ci)
        write_identity(COL_EMAIL, data.email)
        write_identity(COL_PHONE, data.phone)
        write_identity(COL_MATERIA, data.subject)
        write_identity(COL_GRUPO, data.group_code)
        write_identity(COL_TIPO_DOCENTE, data.teacher_type)
        write_identity(COL_GENERO, data.gender)
        write_identity(COL_SAP, data.sap_code)
        write_identity(COL_FACTURA, data.invoice_retention)
        write_identity(COL_CUENTA, data.account_number)
        write_identity(COL_NIVEL_ACAD, data.academic_level)
        write_identity(COL_PROFESION, data.profession)
        write_identity(COL_ESPECIALIDAD, data.specialty)
        write_identity(COL_BANCO, data.bank)

        # Day columns — iterate each block. Days outside the active cutoff
        # are grayed out (no value, no class coloring). Spacer columns between
        # blocks are styled neutrally.
        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                cell = ws.cell(row=row_num, column=col)

                is_active = block.active_start <= day <= block.active_end
                if not is_active:
                    # Outside cutoff — neutral fill, no value
                    cell.value = None
                    cell.fill = PatternFill("solid", fgColor=COLOR_OUTSIDE_CUTOFF)
                    cell.border = THIN_BORDER
                    continue

                # Active day — look up attendance by the full date
                target_date = date(block.year, block.month, day)
                hours = data.daily_hours.get(target_date, 0)
                status = data.daily_status.get(target_date, "")

                # Determine fill color based on status and day type
                if status == "EXCLUDED":
                    fill_color = COLOR_DAY_EXCLUDED
                elif status == "ABSENT":
                    fill_color = COLOR_DAY_ABSENT
                elif status == "LATE":
                    fill_color = COLOR_DAY_LATE
                elif status in ("ATTENDED", "NO_EXIT"):
                    fill_color = COLOR_DAY_CLASS
                elif hours > 0:
                    fill_color = COLOR_DAY_CLASS
                else:
                    # No class on this day — check if it's a weekend
                    if target_date.weekday() >= 5:  # Saturday=5, Sunday=6
                        fill_color = "E8E8E8"   # Slightly darker gray for weekends
                    else:
                        fill_color = COLOR_DAY_WEEKEND

                cell.value = hours if hours > 0 else (0 if status == "EXCLUDED" else None)
                cell.font = Font(name="Calibri", size=9, bold=(hours > 0))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = THIN_BORDER

            # Spacer column styling (only present between blocks)
            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                spacer_cell = ws.cell(row=row_num, column=spacer_col)
                spacer_cell.value = None
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        # Summary columns
        def write_summary(col: int, value, is_currency: bool = False) -> None:
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.font = base_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="EBF3FB")
            cell.border = THIN_BORDER
            if is_currency:
                cell.number_format = '#,##0.00 "Bs"'

        # Model C column semantics (positions are dynamic via scols):
        #   total_horas      = payable_hours (base - absent) — primary hours
        #   hrs_teoria       = base_monthly_hours (assigned load from designation)
        #   hrs_pract_int    = absent_hours (hours deducted)
        #   hrs_pract_ext    = attended hours (informational)
        #   total_hrs_check  = payable_hours (verification = base - deducted)
        write_summary(scols['total_horas'], data.payable_hours)
        write_summary(scols['grupo_resumen'], data.group_code)
        write_summary(scols['materia_resumen'], data.subject)
        write_summary(scols['pago_hora'], data.rate_per_hour, is_currency=True)
        write_summary(scols['hrs_teoria'], data.base_monthly_hours)
        write_summary(scols['hrs_pract_int'], data.absent_hours if data.absent_hours > 0 else None)
        write_summary(scols['hrs_pract_ext'], data.total_practice_external_hours if data.total_practice_external_hours > 0 else None)
        write_summary(scols['total_hrs_check'], data.payable_hours)
        write_summary(scols['pago_calculado'], data.calculated_payment, is_currency=True)

        # Retención RC-IVA
        ret_cell = ws.cell(row=row_num, column=scols['retencion_amt'])
        ret_cell.value = data.retention_amount if data.has_retention else None
        ret_cell.font = Font(name="Calibri", size=9, color="C00000" if data.has_retention else "000000")
        ret_cell.alignment = Alignment(horizontal="center", vertical="center")
        ret_cell.fill = PatternFill("solid", fgColor="FFF0F0" if data.has_retention else "EBF3FB")
        ret_cell.border = THIN_BORDER
        if data.has_retention:
            ret_cell.number_format = '#,##0.00 "Bs"'

        # Pago Neto
        net_cell = ws.cell(row=row_num, column=scols['pago_neto'])
        net_cell.value = data.final_payment
        net_cell.font = Font(name="Calibri", size=9, bold=True, color="1F4E79")
        net_cell.alignment = Alignment(horizontal="center", vertical="center")
        net_cell.fill = PatternFill("solid", fgColor="DEEAF1")
        net_cell.border = THIN_BORDER
        net_cell.number_format = '#,##0.00 "Bs"'

        # Pago Ajustado
        adj_cell = ws.cell(row=row_num, column=scols['pago_ajustado'])
        adj_cell.value = override if override is not None else None
        adj_cell.font = Font(name="Calibri", size=9, bold=(override is not None), color="C00000" if override is not None else "000000")
        adj_cell.alignment = Alignment(horizontal="center", vertical="center")
        adj_cell.fill = PatternFill("solid", fgColor="FFF0F0" if override is not None else "EBF3FB")
        adj_cell.border = THIN_BORDER
        if override is not None:
            adj_cell.number_format = '#,##0.00 "Bs"'

        # Observations
        obs_cell = ws.cell(row=row_num, column=scols['observaciones'])
        obs_text = "; ".join(data.observations) if data.observations else ""
        obs_cell.value = obs_text if obs_text else None
        obs_cell.font = Font(name="Calibri", size=8, color="595959")
        obs_cell.alignment = Alignment(vertical="center", wrap_text=True)
        obs_cell.fill = base_fill
        obs_cell.border = THIN_BORDER

        ws.row_dimensions[row_num].height = 15

    # ------------------------------------------------------------------
    # Totals row
    # ------------------------------------------------------------------

    def _write_totals_row(
        self,
        ws,
        rows: list[PlanillaRow],
        totals_row: int,
        payment_overrides: dict[str, float],
        blocks: list[MonthBlock],
        scols: dict[str, int],
        total_cols: int,
    ) -> None:
        """Write the totals row at the bottom of all data rows."""
        if not rows:
            return

        # Model C totals
        total_hours = sum(r.payable_hours for r in rows)                    # total payable hours
        total_theory = sum(r.base_monthly_hours for r in rows)              # total assigned hours
        total_pract_int = sum(r.absent_hours for r in rows)                 # total deducted hours
        total_pract_ext = sum(r.total_practice_external_hours for r in rows)  # total attended (info)
        total_bruto = sum(r.calculated_payment for r in rows)               # total gross before retention
        total_retention = sum(r.retention_amount for r in rows)             # total RC-IVA deducted
        total_payment = self._calculate_total_payment(rows, payment_overrides)
        total_teachers = len({r.teacher_ci for r in rows})

        totals_fill = PatternFill("solid", fgColor=COLOR_TOTAL_ROW)
        totals_font = Font(name="Calibri", size=9, bold=True)
        totals_align = Alignment(horizontal="center", vertical="center")

        # Label
        ws.merge_cells(
            start_row=totals_row, start_column=1,
            end_row=totals_row, end_column=COL_BANCO
        )
        label_cell = ws.cell(row=totals_row, column=1)
        label_cell.value = f"TOTALES — {total_teachers} docente(s)"
        label_cell.font = totals_font
        label_cell.fill = totals_fill
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.border = THIN_BORDER

        # Day columns — sum per day using full-date keys. Inactive days and
        # spacer columns get styled but empty cells so the row looks coherent.
        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                cell = ws.cell(row=totals_row, column=col)
                cell.font = totals_font
                cell.alignment = totals_align
                cell.border = THIN_BORDER

                is_active = block.active_start <= day <= block.active_end
                if not is_active:
                    cell.value = None
                    cell.fill = PatternFill("solid", fgColor=COLOR_OUTSIDE_CUTOFF)
                    continue

                target_date = date(block.year, block.month, day)
                day_total = sum(r.daily_hours.get(target_date, 0) for r in rows)
                cell.value = day_total if day_total > 0 else None
                cell.fill = totals_fill

            # Spacer column (only between blocks)
            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                spacer_cell = ws.cell(row=totals_row, column=spacer_col)
                spacer_cell.value = None
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        # Summary totals
        def write_total(col: int, value, is_currency: bool = False) -> None:
            cell = ws.cell(row=totals_row, column=col)
            cell.value = value
            cell.font = totals_font
            cell.alignment = totals_align
            cell.fill = totals_fill
            cell.border = THIN_BORDER
            if is_currency:
                cell.number_format = '#,##0.00 "Bs"'

        write_total(scols['total_horas'], total_hours)
        write_total(scols['grupo_resumen'], None)
        write_total(scols['materia_resumen'], None)
        write_total(scols['pago_hora'], None)
        write_total(scols['hrs_teoria'], total_theory)
        write_total(scols['hrs_pract_int'], total_pract_int)
        write_total(scols['hrs_pract_ext'], total_pract_ext)
        write_total(scols['total_hrs_check'], total_hours)
        write_total(scols['pago_calculado'], total_bruto, is_currency=True)
        write_total(scols['retencion_amt'], total_retention if total_retention > 0 else None, is_currency=True)
        write_total(scols['pago_neto'], total_payment, is_currency=True)
        write_total(scols['pago_ajustado'], None)
        write_total(scols['observaciones'], None)

        ws.row_dimensions[totals_row].height = 18

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------

    def _apply_formatting(
        self, ws, last_row: int, month: int, year: int, total_cols: int,
    ) -> None:
        """Apply freeze panes and print settings."""
        # Freeze panes: rows 1-6 (headers) and cols 1-3 (CI + name columns)
        ws.freeze_panes = ws.cell(row=DATA_ROW_START, column=COL_CI + 1)

        # Print settings
        ws.print_title_rows = f"1:{ROW_WEEKDAY}"
        ws.print_title_cols = f"A:{get_column_letter(COL_BANCO)}"
        ws.sheet_view.showGridLines = True

        # Auto-filter on column headers row — use dynamic total_cols so the
        # filter always ends at the real last column (varies with day-window).
        if last_row >= DATA_ROW_START:
            ws.auto_filter.ref = (
                f"A{ROW_COL_HEADERS}:{get_column_letter(total_cols)}{last_row}"
            )

    # ------------------------------------------------------------------
    # Detail sheet
    # ------------------------------------------------------------------

    def _write_detail_sheet(
        self,
        ws,
        detail_rows: list[DetailRow],
        month: int,
        year: int,
    ) -> None:
        """
        Write the Detalle sheet with granular class-slot data.

        Columns: CI | Docente | Fecha | Día | Materia | Grupo | Semestre |
                 Hora Inicio | Hora Fin | Hrs Académicas | Estado | Observación
        """
        month_name = MONTH_NAMES.get(month, str(month)).upper()
        total_detail_cols = 12

        # Title
        ws.merge_cells(f"A1:L1")
        title = ws.cell(row=1, column=1)
        title.value = f"DETALLE DE ASISTENCIA — {month_name} {year}"
        title.font = Font(name="Calibri", size=12, bold=True, color=COLOR_WHITE)
        title.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Column headers
        detail_headers = [
            "CI", "Docente", "Fecha", "Día", "Materia", "Grupo",
            "Semestre", "Hora Inicio", "Hora Fin", "Hrs Académicas", "Estado", "Observación"
        ]
        header_widths = [10, 28, 12, 5, 28, 8, 10, 10, 10, 12, 12, 40]

        for col, (header, width) in enumerate(zip(detail_headers, header_widths), start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(name="Calibri", size=9, bold=True, color=COLOR_WHITE)
            cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[2].height = 18
        ws.freeze_panes = ws.cell(row=3, column=1)

        # Status color mapping
        status_colors = {
            "ATTENDED": COLOR_DAY_CLASS,
            "LATE": COLOR_DAY_LATE,
            "ABSENT": COLOR_DAY_ABSENT,
            "NO_EXIT": "E6F0FF",   # Light blue
        }

        # Data rows
        for i, detail in enumerate(detail_rows):
            row_num = 3 + i
            is_even = i % 2 == 0
            bg = "FFFFFF" if is_even else "F5F8FA"

            status_upper = detail.status.upper()
            status_bg = status_colors.get(status_upper, bg)

            row_data = [
                detail.ci,
                detail.teacher_name,
                detail.date.strftime("%d/%m/%Y"),
                detail.day_letter,
                detail.subject,
                detail.group_code,
                detail.semester,
                detail.scheduled_start,
                detail.scheduled_end,
                detail.academic_hours,
                detail.status,
                detail.observation or "",
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                # Status column and hours get special coloring
                if col == 11:  # Estado
                    cell.fill = PatternFill("solid", fgColor=status_bg)
                    cell.font = Font(name="Calibri", size=9, bold=True)
                elif col == 10:  # Hrs Académicas
                    cell.font = Font(name="Calibri", size=9, bold=True)
                    cell.fill = PatternFill("solid", fgColor=bg)
                else:
                    cell.font = Font(name="Calibri", size=9)
                    cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(
                    horizontal="center" if col in (1, 3, 4, 7, 8, 9, 10, 11) else "left",
                    vertical="center",
                )
                cell.border = THIN_BORDER

            ws.row_dimensions[row_num].height = 14

        # Auto-filter
        if detail_rows:
            last_row = 2 + len(detail_rows)
            ws.auto_filter.ref = f"A2:L{last_row}"

        # Add summary at bottom
        if detail_rows:
            summary_row = 3 + len(detail_rows) + 1
            ws.cell(row=summary_row, column=1).value = "TOTAL REGISTROS:"
            ws.cell(row=summary_row, column=2).value = len(detail_rows)
            total_hrs_cell = ws.cell(row=summary_row, column=10)
            total_hrs_cell.value = sum(d.academic_hours for d in detail_rows)
            total_hrs_cell.font = Font(name="Calibri", size=9, bold=True)
            ws.cell(row=summary_row, column=1).font = Font(name="Calibri", size=9, bold=True)
            ws.cell(row=summary_row, column=2).font = Font(name="Calibri", size=9, bold=True)

    # ------------------------------------------------------------------
    # DB persistence
    # ------------------------------------------------------------------

    def _persist_planilla_output(
        self,
        db: Session,
        month: int,
        year: int,
        file_path: str,
        total_teachers: int,
        total_hours: int,
        total_payment: float,
        payment_overrides: Optional[dict[str, float]] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        discount_mode: str = "attendance",
        excluded_days: "list[ExcludedDaySchema] | None" = None,
        calculation_snapshot: dict | None = None,
    ) -> Optional[PlanillaOutput]:
        """
        Create or update a PlanillaOutput record in the DB.
        Uses upsert logic: if one already exists for month/year, update it.
        Stores payment_overrides and excluded_days as JSON so publish/salary-report
        can reconstruct adjusted amounts and exclusions history.
        """
        try:
            existing = (
                db.query(PlanillaOutput)
                .filter(
                    PlanillaOutput.month == month,
                    PlanillaOutput.year == year,
                )
                .first()
            )

            overrides_data = {key: format(value, ".2f") for key, value in payment_overrides.items()} or None
            exclusions_data = (
                [exc.model_dump(mode="json") for exc in excluded_days]
                if excluded_days
                else None
            )

            if existing:
                existing.file_path = file_path
                existing.total_teachers = total_teachers
                existing.total_hours = total_hours
                existing.total_payment = Decimal(str(total_payment))
                existing.payment_overrides_json = overrides_data
                existing.excluded_days_json = exclusions_data
                existing.calculation_snapshot = calculation_snapshot
                existing.start_date = start_date
                existing.end_date = end_date
                existing.discount_mode = discount_mode
                existing.generated_at = datetime.now()
                existing.status = "generated"
                db.flush()
                logger.info("Updated PlanillaOutput id=%d", existing.id)
                return existing
            else:
                output = PlanillaOutput(
                    month=month,
                    year=year,
                    file_path=file_path,
                    total_teachers=total_teachers,
                    total_hours=total_hours,
                    total_payment=Decimal(str(total_payment)),
                    payment_overrides_json=overrides_data,
                    excluded_days_json=exclusions_data,
                    calculation_snapshot=calculation_snapshot,
                    start_date=start_date,
                    end_date=end_date,
                    discount_mode=discount_mode,
                    status="generated",
                )
                db.add(output)
                db.flush()
                logger.info("Created PlanillaOutput id=%d", output.id)
                return output

        except Exception as exc:
            logger.exception("Failed to persist PlanillaOutput: %s", exc)
            return None

    # ------------------------------------------------------------------
    # Style helpers
    # ------------------------------------------------------------------

    def _style_section_header(self, cell) -> None:
        """Apply section header style (dark blue bg, white bold text, centered)."""
        cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    def _style_col_header(
        self, cell, is_day: bool = False, wrap: bool = False
    ) -> None:
        """Apply column header style (light blue bg, dark bold text)."""
        cell.font = Font(name="Calibri", size=9, bold=True, color="1F3864")
        cell.fill = PatternFill("solid", fgColor=COLOR_COL_HEADER_BG)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=wrap,
        )
        cell.border = THIN_BORDER

    def _get_row_override(
        self,
        row: PlanillaRow,
        payment_overrides: dict[str, float],
        all_rows: list[PlanillaRow],
    ) -> Optional[float]:
        """Resolve the display override for a specific row."""
        teacher_rows = [candidate for candidate in all_rows if candidate.teacher_ci == row.teacher_ci]
        allocations = self._get_teacher_override_allocations(teacher_rows, payment_overrides)
        if allocations is not None:
            return allocations.get(row.designation_id)

        return self._resolve_override(row.teacher_ci, row.designation_id, payment_overrides)

    def _resolve_override(
        self,
        teacher_ci: str,
        designation_id: int,
        overrides: dict[str, float],
    ) -> Optional[float]:
        """Resolve override precedence consistently across row and total calculations."""
        row_key = f"{teacher_ci}:{designation_id}"
        if row_key in overrides:
            return overrides[row_key]
        if teacher_ci in overrides:
            return overrides[teacher_ci]
        return None

    def _distribute_teacher_override(
        self,
        row: PlanillaRow,
        teacher_rows: list[PlanillaRow],
        teacher_override: float,
    ) -> float:
        """Distribute a teacher-level override proportionally by row hours."""
        total_hours = sum(candidate.total_hours for candidate in teacher_rows)
        if total_hours <= 0:
            return teacher_override / len(teacher_rows)
        return teacher_override * (row.total_hours / total_hours)

    def _get_teacher_override_allocations(
        self,
        teacher_rows: list[PlanillaRow],
        payment_overrides: dict[str, float],
    ) -> dict[int, float] | None:
        return get_teacher_override_allocations(teacher_rows, payment_overrides)

    def _calculate_total_payment(
        self,
        rows: list[PlanillaRow],
        payment_overrides: dict[str, float],
    ) -> float:
        """Apply override precedence consistently across the full planilla."""
        return calculate_override_total(rows, payment_overrides)


# ---------------------------------------------------------------------------
# Module-level helper
# ---------------------------------------------------------------------------


def _get_month_from_name(month_name: str) -> int:
    """Reverse lookup: month name → month number. Fallback = 1."""
    name_upper = month_name.upper()
    for num, name in MONTH_NAMES.items():
        if name.upper() == name_upper:
            return num
    return 1
