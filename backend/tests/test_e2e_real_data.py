"""
E2E Integration Test - Planilla Docentes UPDS
Phase 5: Full end-to-end validation using REAL data files.

Run from backend/ directory:
    python -m pytest tests/test_e2e_real_data.py -v -s
  or directly:
    python tests/test_e2e_real_data.py
"""
from __future__ import annotations

import io
import os
import sys
import time
from datetime import datetime
from pathlib import Path

# Force UTF-8 output on Windows to avoid cp1252 issues with Unicode chars.
# Skip when running under pytest (capture=default) to avoid breaking pytest's I/O wrappers.
if sys.platform == "win32" and "pytest" not in sys.modules:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# Make sure we can import the app
BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

# ── Paths ────────────────────────────────────────────────────────────────────
PROJECT_ROOT = BACKEND_DIR.parent
BIOMETRIC_FILE = PROJECT_ROOT / "reporte biometrico marzo_docentes.xls"
DESIGNATIONS_FILE = PROJECT_ROOT / "designacion_new.json"
OUTPUT_DIR = BACKEND_DIR / "data" / "output"
OUTPUT_FILE = OUTPUT_DIR / "planilla_marzo_2026.xlsx"

MONTH = 3
YEAR = 2026

# ── Test results accumulator ─────────────────────────────────────────────────
results: list[dict] = []


def step(name: str):
    """Context manager / decorator to time and record each step."""
    class _Step:
        def __init__(self, label):
            self.label = label
            self.start = None

        def __enter__(self):
            self.start = time.perf_counter()
            print(f"\n{'='*60}")
            print(f"STEP: {self.label}")
            print(f"{'='*60}")
            return self

        def __exit__(self, exc_type, exc_val, exc_tb):
            elapsed = time.perf_counter() - self.start
            status = "PASS" if exc_type is None else "FAIL"
            error = str(exc_val) if exc_val else None
            results.append({
                "step": self.label,
                "status": status,
                "elapsed_s": round(elapsed, 2),
                "error": error,
            })
            if exc_type is not None:
                print(f"  ❌ FAILED: {exc_val}")
                return False   # re-raise
            print(f"  ✅ PASSED ({elapsed:.2f}s)")
            return True

    return _Step(name)


# ─────────────────────────────────────────────────────────────────────────────
# DB setup — use the real PostgreSQL database
# ─────────────────────────────────────────────────────────────────────────────

def get_real_db_session():
    """Return a real PostgreSQL session (not SQLite)."""
    from app.database import SessionLocal
    return SessionLocal()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: Clear existing data
# ─────────────────────────────────────────────────────────────────────────────

def step1_clear_data():
    with step("Step 1: Clear existing data (fresh start)"):
        from app.database import create_tables, engine
        from app.main import _run_column_migrations
        from app.models.attendance import AttendanceRecord
        from app.models.biometric import BiometricRecord, BiometricUpload
        from app.models.designation import Designation
        from app.models.planilla import PlanillaOutput
        from app.models.teacher import Teacher

        # Ensure schema is up-to-date (new columns added since last run)
        create_tables()
        _run_column_migrations()

        db = get_real_db_session()
        try:
            # Delete in FK-safe order
            deleted_att = db.query(AttendanceRecord).delete()
            deleted_bio_rec = db.query(BiometricRecord).delete()
            deleted_bio_up = db.query(BiometricUpload).delete()
            deleted_des = db.query(Designation).delete()
            deleted_plan = db.query(PlanillaOutput).delete()
            deleted_teach = db.query(Teacher).delete()
            db.commit()
            print(f"  Deleted: {deleted_att} attendance_records")
            print(f"  Deleted: {deleted_bio_rec} biometric_records")
            print(f"  Deleted: {deleted_bio_up} biometric_uploads")
            print(f"  Deleted: {deleted_des} designations")
            print(f"  Deleted: {deleted_plan} planilla_outputs")
            print(f"  Deleted: {deleted_teach} teachers")
        finally:
            db.close()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: Load designations
# ─────────────────────────────────────────────────────────────────────────────

def step2_load_designations():
    with step("Step 2: Load designations from JSON"):
        from app.services.designation_loader import DesignationLoader

        assert DESIGNATIONS_FILE.exists(), f"Designations file not found: {DESIGNATIONS_FILE}"

        db = get_real_db_session()
        try:
            loader = DesignationLoader()
            t0 = time.perf_counter()
            result = loader.load_from_json(db=db, json_path=str(DESIGNATIONS_FILE))
            elapsed = time.perf_counter() - t0

            print(f"  Teachers created : {result.teachers_created}")
            print(f"  Teachers reused  : {result.teachers_reused}")
            print(f"  Designations     : {result.designations_loaded}")
            print(f"  Skipped          : {result.total_skipped}")
            print(f"  Warnings         : {len(result.warnings)}")
            if result.warnings:
                for w in result.warnings[:5]:
                    print(f"    ⚠ {w}")
                if len(result.warnings) > 5:
                    print(f"    ... {len(result.warnings)-5} more warnings")
            print(f"  Time             : {elapsed:.2f}s")

            # Verify expectations
            total_teachers = result.teachers_created + result.teachers_reused
            assert result.designations_loaded >= 350, \
                f"Expected ~397 designations, got {result.designations_loaded}"
            assert result.teachers_created >= 100, \
                f"Expected ~130 teachers, got {result.teachers_created}"
            print(f"\n  ✔ {result.teachers_created} teachers created, {result.designations_loaded} designations loaded")
        finally:
            db.close()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: Parse biometric report
# ─────────────────────────────────────────────────────────────────────────────

def step3_parse_biometric():
    """Returns (upload_id, ci_name_map)."""
    with step("Step 3: Parse biometric XLS report"):
        from app.services.biometric_parser import BiometricParser

        assert BIOMETRIC_FILE.exists(), f"Biometric file not found: {BIOMETRIC_FILE}"

        db = get_real_db_session()
        try:
            parser = BiometricParser()
            t0 = time.perf_counter()
            parse_result = parser.parse_file(str(BIOMETRIC_FILE))
            elapsed_parse = time.perf_counter() - t0

            print(f"  Teachers found   : {parse_result.stats['total_teachers']}")
            print(f"  Total records    : {parse_result.stats['total_records']}")
            print(f"  Date range       : {parse_result.stats['date_range']}")
            print(f"  Parse warnings   : {len(parse_result.warnings)}")
            if parse_result.warnings:
                for w in parse_result.warnings[:3]:
                    print(f"    ⚠ {w}")
            print(f"  Parse time       : {elapsed_parse:.2f}s")

            # Save to DB
            t1 = time.perf_counter()
            upload = parser.save_to_db(
                db=db,
                parse_result=parse_result,
                month=MONTH,
                year=YEAR,
                filename=BIOMETRIC_FILE.name,
            )
            db.commit()
            elapsed_save = time.perf_counter() - t1

            print(f"  Upload ID        : {upload.id}")
            print(f"  DB save time     : {elapsed_save:.2f}s")

            # Build ci_name_map for teacher linking
            ci_name_map: dict[str, str] = {}
            for ci, entries in parse_result.records.items():
                if entries:
                    ci_name_map[ci] = entries[0].teacher_name

            assert parse_result.stats["total_teachers"] >= 50, \
                f"Expected ~72+ teachers, got {parse_result.stats['total_teachers']}"
            assert parse_result.stats["total_records"] >= 100, \
                f"Expected many records, got {parse_result.stats['total_records']}"

            # Store for later steps
            step3_parse_biometric.upload_id = upload.id
            step3_parse_biometric.ci_name_map = ci_name_map

            print(f"\n  ✔ Upload id={upload.id}, {len(ci_name_map)} teachers with CI+name pairs")
        finally:
            db.close()


step3_parse_biometric.upload_id = None
step3_parse_biometric.ci_name_map = {}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: Link teachers (TEMP CI → real CI)
# ─────────────────────────────────────────────────────────────────────────────

def step4_link_teachers():
    with step("Step 4: Link teachers (TEMP CI → real CI)"):
        from app.services.designation_loader import DesignationLoader
        from app.models.teacher import Teacher

        ci_name_map = step3_parse_biometric.ci_name_map
        assert ci_name_map, "No ci_name_map — did step 3 run?"

        db = get_real_db_session()
        try:
            loader = DesignationLoader()
            t0 = time.perf_counter()
            linked = loader.link_teachers_by_name(db=db, ci_name_map=ci_name_map)
            elapsed = time.perf_counter() - t0

            # Count remaining TEMP teachers
            temp_count = db.query(Teacher).filter(Teacher.ci.like("TEMP-%")).count()
            total_teachers = db.query(Teacher).count()

            print(f"  Teachers linked  : {linked}")
            print(f"  TEMP remaining   : {temp_count}")
            print(f"  Total teachers   : {total_teachers}")
            print(f"  Time             : {elapsed:.2f}s")

            # Sample some linked teachers
            real_teachers = (
                db.query(Teacher)
                .filter(~Teacher.ci.like("TEMP-%"))
                .limit(5)
                .all()
            )
            print(f"\n  Sample linked teachers:")
            for t in real_teachers:
                print(f"    CI={t.ci}  Name={t.full_name}")

            print(f"\n  ✔ {linked} teachers promoted from TEMP to real CI, {temp_count} remain TEMP")
        finally:
            db.close()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: Process attendance for March 2026
# ─────────────────────────────────────────────────────────────────────────────

def step5_process_attendance():
    with step("Step 5: Process attendance for March 2026"):
        from app.services.attendance_engine import AttendanceEngine
        from app.models.attendance import AttendanceRecord

        upload_id = step3_parse_biometric.upload_id
        assert upload_id is not None, "No upload_id — did step 3 run?"

        db = get_real_db_session()
        try:
            engine = AttendanceEngine()
            t0 = time.perf_counter()
            result = engine.process_month(
                db=db,
                upload_id=upload_id,
                month=MONTH,
                year=YEAR,
            )
            db.commit()
            elapsed = time.perf_counter() - t0

            total = result.total_slots
            rate = round(result.present / total * 100, 1) if total else 0.0

            print(f"  Total slots      : {result.total_slots}")
            print(f"  Attended         : {result.attended}")
            print(f"  Late             : {result.late}")
            print(f"  Absent           : {result.absent}")
            print(f"  No exit          : {result.no_exit}")
            print(f"  Present (att+late+no_exit): {result.present}")
            print(f"  Attendance rate  : {rate}%")
            print(f"  Warnings         : {len(result.warnings)}")
            if result.warnings:
                for w in result.warnings[:3]:
                    print(f"    ⚠ {w}")
            print(f"  Time             : {elapsed:.2f}s")

            # Verify some records were created
            saved_count = db.query(AttendanceRecord).filter(
                AttendanceRecord.month == MONTH,
                AttendanceRecord.year == YEAR,
            ).count()
            print(f"  DB records saved : {saved_count}")

            assert saved_count > 0, "No attendance records were saved"
            assert result.total_slots > 0, "No slots were processed"

            print(f"\n  ✔ {saved_count} attendance records created, {rate}% attendance rate")
        finally:
            db.close()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6: Generate planilla Excel
# ─────────────────────────────────────────────────────────────────────────────

def step6_generate_planilla():
    with step("Step 6: Generate planilla Excel for March 2026"):
        from app.services.planilla_generator import PlanillaGenerator

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        db = get_real_db_session()
        try:
            gen = PlanillaGenerator(output_dir=str(OUTPUT_DIR))
            t0 = time.perf_counter()
            result = gen.generate(
                db=db,
                month=MONTH,
                year=YEAR,
                payment_overrides={},
            )
            db.commit()
            elapsed = time.perf_counter() - t0

            print(f"  File path        : {result.file_path}")
            print(f"  Total teachers   : {result.total_teachers}")
            print(f"  Total hours      : {result.total_hours}")
            print(f"  Total payment    : {result.total_payment} Bs")
            print(f"  Planilla ID      : {result.planilla_output_id}")
            print(f"  Warnings         : {len(result.warnings)}")
            if result.warnings:
                for w in result.warnings[:3]:
                    print(f"    ⚠ {w}")
            print(f"  Time             : {elapsed:.2f}s")

            assert result.file_path, "No file path returned"
            file_p = Path(result.file_path)
            assert file_p.exists(), f"Output file not found: {file_p}"
            file_size = file_p.stat().st_size
            print(f"  File size        : {file_size:,} bytes")

            assert result.total_teachers > 0, "No teachers in planilla"
            assert result.total_hours > 0, "No hours in planilla"
            assert result.total_payment > 0, "No payment in planilla"

            # Store for step 7
            step6_generate_planilla.file_path = result.file_path
            step6_generate_planilla.total_teachers = result.total_teachers
            step6_generate_planilla.total_hours = result.total_hours
            step6_generate_planilla.total_payment = result.total_payment

            print(f"\n  ✔ Planilla generated: {file_p.name} ({file_size:,} bytes)")
        finally:
            db.close()


step6_generate_planilla.file_path = None
step6_generate_planilla.total_teachers = 0
step6_generate_planilla.total_hours = 0
step6_generate_planilla.total_payment = 0.0


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7: Validate the generated Excel
# ─────────────────────────────────────────────────────────────────────────────

def step7_validate_excel():
    with step("Step 7: Validate generated planilla Excel"):
        import openpyxl
        from app.models.attendance import AttendanceRecord
        from app.models.teacher import Teacher

        file_path = step6_generate_planilla.file_path or str(OUTPUT_FILE)
        assert Path(file_path).exists(), f"Planilla file not found: {file_path}"

        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames
        print(f"  Sheets           : {sheet_names}")

        assert "Planilla" in sheet_names, f"Missing 'Planilla' sheet. Got: {sheet_names}"
        assert "Detalle" in sheet_names, f"Missing 'Detalle' sheet. Got: {sheet_names}"

        # ── Validate Planilla sheet ───────────────────────────────────────
        ws_planilla = wb["Planilla"]
        max_row = ws_planilla.max_row
        max_col = ws_planilla.max_column
        print(f"  Planilla sheet   : {max_row} rows × {max_col} cols")

        # Data starts at row 7 (after title + headers)
        data_rows = max_row - 6  # rough count excluding totals
        print(f"  Data rows (est)  : {data_rows}")

        # Check header rows exist
        title_cell = ws_planilla.cell(row=1, column=1).value
        print(f"  Title cell [A1]  : {title_cell!r}")
        col_header_row_val = ws_planilla.cell(row=5, column=2).value
        print(f"  Col header [B5]  : {col_header_row_val!r}")

        assert max_row >= 8, f"Planilla sheet has too few rows: {max_row}"
        assert max_col >= 48, f"Expected ≥48 columns, got {max_col}"

        # Find rows with CI data (column C, starting from row 7)
        ci_values = []
        payment_values = []
        sample_rows = []

        for r in range(7, max_row + 1):
            ci = ws_planilla.cell(row=r, column=3).value  # CI column
            name = ws_planilla.cell(row=r, column=2).value  # Name column
            total_hrs = ws_planilla.cell(row=r, column=48).value  # Total Horas (AV)
            payment = ws_planilla.cell(row=r, column=56).value  # Total Pago Calculado (BD)

            if ci and str(ci).strip() and str(ci).strip() not in ("CI", "C.I.", ""):
                ci_values.append(str(ci).strip())
                if payment is not None:
                    try:
                        payment_values.append(float(payment))
                    except (ValueError, TypeError):
                        pass
                if len(sample_rows) < 5 and name and total_hrs is not None:
                    sample_rows.append({
                        "ci": ci,
                        "name": name,
                        "total_hours": total_hrs,
                        "payment": payment,
                    })

        print(f"  Rows with CI data: {len(ci_values)}")
        print(f"  Rows with payment: {len(payment_values)}")

        if payment_values:
            print(f"  Payment range    : {min(payment_values):.0f} – {max(payment_values):.0f} Bs")
            print(f"  Total payment    : {sum(payment_values):,.0f} Bs")

        # ── Validate Detalle sheet ────────────────────────────────────────
        ws_detalle = wb["Detalle"]
        det_max_row = ws_detalle.max_row
        det_max_col = ws_detalle.max_column
        print(f"\n  Detalle sheet    : {det_max_row} rows × {det_max_col} cols")
        det_header = ws_detalle.cell(row=1, column=1).value
        print(f"  Detalle header   : {det_header!r}")

        # ── Sample output ─────────────────────────────────────────────────
        print(f"\n  {'─'*55}")
        print(f"  {'SAMPLE: 5 TEACHERS FROM PLANILLA':^55}")
        print(f"  {'─'*55}")
        print(f"  {'CI':<15} {'Name':<30} {'Hours':>6} {'Payment':>10}")
        print(f"  {'─'*55}")
        for row in sample_rows[:5]:
            name_short = str(row['name'])[:28]
            hrs = row['total_hours'] if row['total_hours'] is not None else 0
            pay = row['payment'] if row['payment'] is not None else 0
            try:
                pay_fmt = f"{float(pay):>9.0f}"
            except (ValueError, TypeError):
                pay_fmt = f"{str(pay):>10}"
            print(f"  {str(row['ci']):<15} {name_short:<30} {hrs:>6} {pay_fmt} Bs")
        print(f"  {'─'*55}")

        assert len(ci_values) > 0, "No data rows found in Planilla sheet"
        assert len(payment_values) > 0, "No payment values found in Planilla sheet"

        wb.close()
        print(f"\n  ✔ Excel validated — {len(ci_values)} data rows, payments ranging "
              f"{min(payment_values):.0f}–{max(payment_values):.0f} Bs")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 8: Test API endpoints via TestClient
# ─────────────────────────────────────────────────────────────────────────────

def step8_test_api():
    with step("Step 8: Test API endpoints via FastAPI TestClient"):
        from fastapi.testclient import TestClient
        from app.main import app
        from app.models.teacher import Teacher
        from app.models.user import User
        from app.services.auth_service import auth_service

        # Get a real CI for parameterized tests
        db = get_real_db_session()
        try:
            sample_teacher = (
                db.query(Teacher)
                .filter(~Teacher.ci.like("TEMP-%"))
                .first()
            )
            sample_ci = sample_teacher.ci if sample_teacher else None

            # Ensure a test admin user exists and get a token
            admin_user = db.query(User).filter(User.ci == "admin").first()
            if admin_user is None:
                # Create default admin if not present
                auth_service.create_default_admin(db)
                admin_user = db.query(User).filter(User.ci == "admin").first()

            if admin_user is None:
                # Fallback: create a fresh test admin
                admin_user = User(
                    ci="E2E_TEST_ADMIN",
                    full_name="E2E Test Admin",
                    password_hash=auth_service.hash_password("testpass123"),
                    role="admin",
                    is_active=True,
                )
                db.add(admin_user)
                db.commit()
                db.refresh(admin_user)

            admin_token = auth_service.create_access_token(
                data={"sub": str(admin_user.id), "role": "admin"}
            )
        finally:
            db.close()

        auth_headers = {"Authorization": f"Bearer {admin_token}"}
        client = TestClient(app, raise_server_exceptions=False)

        endpoint_results = []

        def check(method, url, expected_status=200, label=None):
            t0 = time.perf_counter()
            resp = getattr(client, method)(url, headers=auth_headers)
            elapsed = time.perf_counter() - t0
            ok = resp.status_code == expected_status
            status_str = "✅" if ok else "❌"
            label = label or url
            result_line = (
                f"  {status_str} {resp.status_code} | {elapsed*1000:.0f}ms | {label}"
            )
            print(result_line)
            endpoint_results.append({
                "url": url,
                "status": resp.status_code,
                "ok": ok,
                "elapsed_ms": round(elapsed * 1000),
            })
            if not ok:
                try:
                    body = resp.json()
                    print(f"     Body: {str(body)[:200]}")
                except Exception:
                    print(f"     Body: {resp.text[:200]}")
            else:
                try:
                    body = resp.json()
                    if isinstance(body, dict):
                        # Print first few keys
                        keys = list(body.keys())[:6]
                        print(f"     Keys: {keys}")
                    elif isinstance(body, list):
                        print(f"     Items: {len(body)}")
                except Exception:
                    pass
            return resp

        print(f"\n  Endpoints:")
        check("get", "/health")
        check("get", "/api/dashboard/summary")
        check("get", "/api/teachers")
        check("get", "/api/teachers?page=1&per_page=10")

        if sample_ci:
            check("get", f"/api/teachers/{sample_ci}", label=f"GET /api/teachers/{{ci}}")
        else:
            print("  ⚠ No real CI available for teacher detail test")

        check("get", f"/api/attendance/{MONTH}/{YEAR}")
        check("get", f"/api/attendance/{MONTH}/{YEAR}/summary")
        check("get", f"/api/observations/{MONTH}/{YEAR}")
        check("get", "/api/planilla/history")

        passed = sum(1 for r in endpoint_results if r["ok"])
        total_ep = len(endpoint_results)
        print(f"\n  API Results: {passed}/{total_ep} endpoints passed")
        assert passed >= total_ep - 1, \
            f"Too many API failures: {total_ep - passed}/{total_ep} failed"

        print(f"\n  ✔ {passed}/{total_ep} API endpoints passed")


# ─────────────────────────────────────────────────────────────────────────────
# FINAL REPORT
# ─────────────────────────────────────────────────────────────────────────────

def print_final_report():
    print(f"\n\n{'#'*65}")
    print(f"  E2E INTEGRATION TEST REPORT — Planilla Docentes UPDS")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'#'*65}\n")

    passed = [r for r in results if r["status"] == "PASS"]
    failed = [r for r in results if r["status"] == "FAIL"]

    print(f"  SUMMARY: {len(passed)}/{len(results)} steps passed\n")
    print(f"  {'Step':<45} {'Status':<8} {'Time':>7}")
    print(f"  {'─'*62}")
    for r in results:
        icon = "✅" if r["status"] == "PASS" else "❌"
        print(f"  {icon} {r['step']:<43} {r['status']:<8} {r['elapsed_s']:>5.1f}s")
        if r["error"]:
            print(f"      Error: {r['error'][:80]}")
    print(f"  {'─'*62}")
    total_time = sum(r["elapsed_s"] for r in results)
    print(f"  {'Total time':<45} {'':<8} {total_time:>5.1f}s")

    print(f"\n  {'─'*62}")
    if failed:
        print(f"\n  ❌ {len(failed)} STEP(S) FAILED:")
        for r in failed:
            print(f"     • {r['step']}: {r['error']}")
    else:
        print(f"\n  ✅ ALL {len(passed)} STEPS PASSED — SYSTEM READY FOR PRODUCTION")

    # Data quality summary
    print(f"\n  DATA QUALITY:")
    print(f"  • Biometric file  : {BIOMETRIC_FILE.name}")
    print(f"  • Designations    : {DESIGNATIONS_FILE.name}")
    if step6_generate_planilla.file_path:
        file_p = Path(step6_generate_planilla.file_path)
        size = file_p.stat().st_size if file_p.exists() else 0
        print(f"  • Output planilla : {file_p.name} ({size:,} bytes)")
        print(f"  • Total teachers  : {step6_generate_planilla.total_teachers}")
        print(f"  • Total hours     : {step6_generate_planilla.total_hours}")
        print(f"  • Total payment   : {step6_generate_planilla.total_payment:,.0f} Bs")

    print(f"\n{'#'*65}\n")


# ─────────────────────────────────────────────────────────────────────────────
# pytest entry point
# ─────────────────────────────────────────────────────────────────────────────

def test_e2e_full_flow():
    """Single pytest test that runs the complete E2E flow."""
    step1_clear_data()
    step2_load_designations()
    step3_parse_biometric()
    step4_link_teachers()
    step5_process_attendance()
    step6_generate_planilla()
    step7_validate_excel()
    step8_test_api()
    print_final_report()

    # Assert overall pass
    failed_steps = [r for r in results if r["status"] == "FAIL"]
    assert not failed_steps, f"E2E test failed at: {[r['step'] for r in failed_steps]}"


# ─────────────────────────────────────────────────────────────────────────────
# Direct execution
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\nPlanilla Docentes UPDS — E2E Integration Test")
    print(f"Working dir: {BACKEND_DIR}")
    print(f"Biometric  : {BIOMETRIC_FILE}")
    print(f"Desig. JSON: {DESIGNATIONS_FILE}")
    print(f"Output dir : {OUTPUT_DIR}")

    try:
        test_e2e_full_flow()
    except Exception as exc:
        print_final_report()
        print(f"\n💥 FATAL: {exc}")
        sys.exit(1)

    sys.exit(0)
