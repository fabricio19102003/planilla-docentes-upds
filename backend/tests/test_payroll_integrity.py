from __future__ import annotations

from datetime import date, time

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, sessionmaker

import app.models  # noqa: F401
from app.database import Base
from app.models.app_setting import AppSetting
from app.models.attendance import AttendanceRecord
from app.models.designation import Designation
from app.models.practice_attendance import PracticeAttendanceLog
from app.models.teacher import Teacher
from app.schemas.attendance import AttendanceProcessRequest
from app.schemas.planilla import ExcludedDaySchema, PlanillaGenerateRequest, SalaryReportRequest
from app.schemas.practice_attendance import PracticeAttendanceBulkCreate
from app.schemas.practice_planilla import PracticePlanillaGenerateRequest
from app.services.planilla_generator import (
    DATA_ROW_START,
    PayrollDataError,
    PlanillaGenerator,
    _build_month_blocks,
    _effective_designation_range,
    _expand_schedule_to_slots,
)
from app.services.practice_planilla_generator import (
    PracticePlanillaCoverageError,
    PracticePlanillaGenerator,
)


@pytest.fixture
def db() -> Session:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()
    session.add_all(
        [
            AppSetting(key="HOURLY_RATE", value="70", description="test"),
            AppSetting(key="PRACTICE_HOURLY_RATE", value="50", description="test"),
            AppSetting(key="ACTIVE_ACADEMIC_PERIOD", value="I/2026", description="test"),
        ]
    )
    session.flush()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(engine)
        engine.dispose()


def add_teacher(db: Session, ci: str) -> Teacher:
    teacher = Teacher(ci=ci, full_name=f"DOCENTE {ci}", invoice_retention=None)
    db.add(teacher)
    db.flush()
    return teacher


def add_designation(
    db: Session,
    teacher_ci: str,
    *,
    designation_type: str = "regular",
    semester: str = "1",
    subject: str = "Materia",
    group: str = "A",
    weekday: str = "lunes",
    start: str = "08:00",
    end: str = "10:00",
    slot_hours: int = 2,
    monthly_hours: int = 8,
    contract_start: date | None = None,
    contract_end: date | None = None,
) -> Designation:
    designation = Designation(
        teacher_ci=teacher_ci,
        subject=subject,
        semester=semester,
        group_code=group,
        academic_period="I/2026",
        designation_type=designation_type,
        schedule_json=[
            {
                "dia": weekday,
                "hora_inicio": start,
                "hora_fin": end,
                "horas_academicas": slot_hours,
            }
        ],
        monthly_hours=monthly_hours,
        weekly_hours=slot_hours,
        contract_start_date=contract_start,
        contract_end_date=contract_end,
    )
    db.add(designation)
    db.flush()
    return designation


def regular_coverage(
    db: Session,
    designation: Designation,
    period_start: date,
    period_end: date,
    *,
    status: str = "ATTENDED",
) -> list[AttendanceRecord]:
    records = []
    for slot in _expand_schedule_to_slots(
        designation.schedule_json,
        period_start,
        period_end,
        designation_id=designation.id,
    ):
        record = AttendanceRecord(
            teacher_ci=designation.teacher_ci,
            designation_id=designation.id,
            date=slot.slot_date,
            scheduled_start=slot.scheduled_start,
            scheduled_end=slot.scheduled_end,
            actual_entry=slot.scheduled_start if status != "ABSENT" else None,
            actual_exit=slot.scheduled_end if status != "ABSENT" else None,
            status=status,
            academic_hours=0 if status == "ABSENT" else slot.academic_hours,
            late_minutes=0,
            month=slot.slot_date.month,
            year=slot.slot_date.year,
        )
        db.add(record)
        records.append(record)
    db.flush()
    return records


def practice_coverage(
    db: Session,
    designation: Designation,
    period_start: date,
    period_end: date,
    *,
    status: str = "attended",
) -> list[PracticeAttendanceLog]:
    records = []
    for slot in _expand_schedule_to_slots(
        designation.schedule_json,
        period_start,
        period_end,
        designation_id=designation.id,
    ):
        record = PracticeAttendanceLog(
            teacher_ci=designation.teacher_ci,
            designation_id=designation.id,
            date=slot.slot_date,
            scheduled_start=slot.scheduled_start,
            scheduled_end=slot.scheduled_end,
            academic_hours=slot.academic_hours,
            status=status,
        )
        db.add(record)
        records.append(record)
    db.flush()
    return records


def test_confirmed_semester_amounts_remain_numeric(db: Session) -> None:
    teacher = add_teacher(db, "SEMESTERS")
    sem1 = add_designation(db, teacher.ci, semester="1", subject="Sem1", group="S1")
    sem7 = add_designation(
        db,
        teacher.ci,
        semester="7",
        subject="Sem7",
        group="S7",
        slot_hours=3,
        monthly_hours=12,
    )
    exclusion = ExcludedDaySchema(date=date(2026, 5, 4), scope="semester", semester_id="7")

    rows, _, _ = PlanillaGenerator()._build_planilla_data(
        db, 5, 2026, discount_mode="full", excluded_days=[exclusion]
    )
    by_semester = {row.semester: row for row in rows}
    assert (by_semester["1"].payable_hours, by_semester["1"].final_payment) == (8, 560)
    assert (by_semester["7"].payable_hours, by_semester["7"].final_payment) == (9, 630)
    assert sum(row.payable_hours for row in rows) == 17
    assert sum(row.final_payment for row in rows) == 1190

    practice_teacher = add_teacher(db, "PRACTICE-SEMESTERS")
    add_designation(
        db,
        practice_teacher.ci,
        designation_type="practice",
        semester="1",
        subject="PSem1",
        group="P1",
    )
    add_designation(
        db,
        practice_teacher.ci,
        designation_type="practice",
        semester="7",
        subject="PSem7",
        group="P7",
        slot_hours=3,
        monthly_hours=12,
    )
    practice_rows, _ = PracticePlanillaGenerator()._build_planilla_data(
        db, 5, 2026, discount_mode="full", excluded_days=[exclusion]
    )
    assert sum(row.payable_hours for row in practice_rows) == 17
    assert sum(row.final_payment for row in practice_rows) == 850


@pytest.mark.parametrize("practice", [False, True])
def test_five_mondays_static_monthly_hours_still_subtracts_exclusion(
    db: Session, practice: bool
) -> None:
    teacher = add_teacher(db, f"FIVE-{practice}")
    add_designation(
        db,
        teacher.ci,
        designation_type="practice" if practice else "regular",
    )
    exclusion = ExcludedDaySchema(date=date(2026, 6, 1), scope="global")
    generator = PracticePlanillaGenerator() if practice else PlanillaGenerator()
    result = generator._build_planilla_data(
        db, 6, 2026, discount_mode="full", excluded_days=[exclusion]
    )
    rows = result[0]
    assert rows[0].payable_hours == 6
    assert sum(rows[0].daily_hours.values()) == 6


@pytest.mark.parametrize("practice", [False, True])
def test_contract_intersection_uses_exact_calendar(db: Session, practice: bool) -> None:
    teacher = add_teacher(db, f"CONTRACT-{practice}")
    add_designation(
        db,
        teacher.ci,
        designation_type="practice" if practice else "regular",
        contract_start=date(2026, 5, 11),
        contract_end=date(2026, 5, 25),
    )
    generator = PracticePlanillaGenerator() if practice else PlanillaGenerator()
    rows = generator._build_planilla_data(db, 5, 2026, discount_mode="full")[0]
    assert rows[0].payable_hours == 6
    assert set(rows[0].daily_hours) == {
        date(2026, 5, 11),
        date(2026, 5, 18),
        date(2026, 5, 25),
    }


def test_missing_coverage_for_one_regular_designation_blocks(db: Session) -> None:
    teacher = add_teacher(db, "MISSING")
    covered = add_designation(db, teacher.ci, subject="Covered", group="C")
    add_designation(db, teacher.ci, subject="Missing", group="M")
    regular_coverage(db, covered, date(2026, 5, 1), date(2026, 5, 31))

    with pytest.raises(PayrollDataError, match="faltantes") as exc_info:
        PlanillaGenerator()._build_planilla_data(db, 5, 2026, discount_mode="attendance")
    assert exc_info.value.code == "regular_attendance_coverage"
    assert any("Missing" not in sample or "Falta" in sample for sample in exc_info.value.sample)


@pytest.mark.parametrize("practice", [False, True])
def test_full_mode_ignores_attendance_but_keeps_contract_schedule_and_exclusion(
    db: Session, practice: bool
) -> None:
    teacher = add_teacher(db, f"FULL-{practice}")
    designation = add_designation(
        db,
        teacher.ci,
        designation_type="practice" if practice else "regular",
        contract_start=date(2026, 5, 11),
        contract_end=date(2026, 5, 25),
    )
    if practice:
        practice_coverage(db, designation, date(2026, 5, 11), date(2026, 5, 25), status="absent")
    else:
        regular_coverage(db, designation, date(2026, 5, 11), date(2026, 5, 25), status="ABSENT")
    exclusion = ExcludedDaySchema(date=date(2026, 5, 18), scope="global")
    generator = PracticePlanillaGenerator() if practice else PlanillaGenerator()
    rows = generator._build_planilla_data(
        db, 5, 2026, discount_mode="full", excluded_days=[exclusion]
    )[0]
    assert rows[0].base_monthly_hours == 4
    assert rows[0].absent_hours == 0
    assert rows[0].payable_hours == 4


def test_practice_duplicate_blocks_legacy_data_and_constraint_protects(db: Session) -> None:
    teacher = add_teacher(db, "DUPLICATE")
    designation = add_designation(db, teacher.ci, designation_type="practice")
    first = PracticeAttendanceLog(
        teacher_ci=teacher.ci,
        designation_id=designation.id,
        date=date(2026, 5, 4),
        scheduled_start=time(8),
        scheduled_end=time(10),
        academic_hours=2,
        status="attended",
    )
    db.add(first)
    db.flush()
    db.add(
        PracticeAttendanceLog(
            teacher_ci=teacher.ci,
            designation_id=designation.id,
            date=date(2026, 5, 4),
            scheduled_start=time(8),
            scheduled_end=time(10),
            academic_hours=2,
            status="attended",
        )
    )
    with pytest.raises(IntegrityError):
        db.flush()
    db.rollback()

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with engine.begin() as connection:
        connection.execute(text("DROP TABLE practice_attendance_logs"))
        connection.execute(
            text(
                """
                CREATE TABLE practice_attendance_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    teacher_ci VARCHAR(20) NOT NULL,
                    designation_id INTEGER NOT NULL,
                    date DATE NOT NULL,
                    scheduled_start TIME NOT NULL,
                    scheduled_end TIME NOT NULL,
                    actual_start TIME,
                    actual_end TIME,
                    academic_hours INTEGER NOT NULL,
                    status VARCHAR(20) NOT NULL,
                    observation TEXT,
                    registered_by VARCHAR(20),
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
    LegacySession = sessionmaker(bind=engine)
    legacy_db = LegacySession()
    try:
        legacy_db.add_all(
            [
                AppSetting(key="PRACTICE_HOURLY_RATE", value="50", description="test"),
                AppSetting(key="ACTIVE_ACADEMIC_PERIOD", value="I/2026", description="test"),
            ]
        )
        legacy_teacher = add_teacher(legacy_db, "LEGACY-DUP")
        legacy_designation = add_designation(
            legacy_db, legacy_teacher.ci, designation_type="practice"
        )
        slots = _expand_schedule_to_slots(
            legacy_designation.schedule_json,
            date(2026, 5, 1),
            date(2026, 5, 31),
            designation_id=legacy_designation.id,
        )
        rows = [
            {
                "teacher_ci": legacy_teacher.ci,
                "designation_id": legacy_designation.id,
                "date": slot.slot_date.isoformat(),
                "scheduled_start": slot.scheduled_start.isoformat(),
                "scheduled_end": slot.scheduled_end.isoformat(),
                "academic_hours": slot.academic_hours,
                "status": "attended",
            }
            for slot in slots
        ]
        rows.append(dict(rows[0]))
        legacy_db.execute(
            text(
                """
                INSERT INTO practice_attendance_logs
                    (teacher_ci, designation_id, date, scheduled_start, scheduled_end, academic_hours, status)
                VALUES
                    (:teacher_ci, :designation_id, :date, :scheduled_start, :scheduled_end, :academic_hours, :status)
                """
            ),
            rows,
        )
        legacy_db.flush()
        with pytest.raises(PracticePlanillaCoverageError, match="duplicados"):
            PracticePlanillaGenerator()._build_planilla_data(
                legacy_db, 5, 2026, discount_mode="attendance"
            )
    finally:
        legacy_db.close()
        engine.dispose()


@pytest.mark.parametrize("practice", [False, True])
def test_unknown_attendance_status_blocks(db: Session, practice: bool) -> None:
    teacher = add_teacher(db, f"UNKNOWN-{practice}")
    designation = add_designation(
        db, teacher.ci, designation_type="practice" if practice else "regular"
    )
    if practice:
        records = practice_coverage(db, designation, date(2026, 5, 1), date(2026, 5, 31))
        records[0].status = "mystery"
        expected = PracticePlanillaCoverageError
        generator = PracticePlanillaGenerator()
    else:
        records = regular_coverage(db, designation, date(2026, 5, 1), date(2026, 5, 31))
        records[0].status = "MYSTERY"
        expected = PayrollDataError
        generator = PlanillaGenerator()
    db.flush()
    with pytest.raises(expected, match="inválidos"):
        generator._build_planilla_data(db, 5, 2026, discount_mode="attendance")


@pytest.mark.parametrize("practice", [False, True])
def test_exclusion_and_absence_same_date_do_not_double_discount(db: Session, practice: bool) -> None:
    teacher = add_teacher(db, f"NO-DOUBLE-{practice}")
    designation = add_designation(
        db, teacher.ci, designation_type="practice" if practice else "regular"
    )
    if practice:
        records = practice_coverage(
            db, designation, date(2026, 5, 1), date(2026, 5, 31), status="attended"
        )
        records[0].status = "absent"
        generator = PracticePlanillaGenerator()
    else:
        records = regular_coverage(
            db, designation, date(2026, 5, 1), date(2026, 5, 31), status="ATTENDED"
        )
        records[0].status = "ABSENT"
        records[0].academic_hours = 0
        generator = PlanillaGenerator()
    excluded_date = records[0].date
    exclusion = ExcludedDaySchema(date=excluded_date, scope="global")
    db.flush()
    rows = generator._build_planilla_data(
        db, 5, 2026, discount_mode="attendance", excluded_days=[exclusion]
    )[0]
    assert rows[0].base_monthly_hours == 6
    assert rows[0].absent_hours == 0
    assert rows[0].payable_hours == 6


def test_three_month_excel_contains_every_month_and_visible_hours(tmp_path, db: Session) -> None:
    teacher = add_teacher(db, "THREE-MONTHS")
    add_designation(
        db,
        teacher.ci,
        weekday="jueves",
        monthly_hours=8,
    )
    generator = PlanillaGenerator(output_dir=str(tmp_path))
    result = generator.generate(
        db,
        6,
        2026,
        start_date=date(2026, 4, 30),
        end_date=date(2026, 6, 30),
        discount_mode="full",
    )
    blocks = _build_month_blocks(6, 2026, date(2026, 4, 30), date(2026, 6, 30))
    assert [block.month for block in blocks] == [4, 5, 6]
    workbook = load_workbook(result.file_path, data_only=True)
    sheet = workbook["Planilla"]
    visible_total = 0
    for block in blocks:
        for day in range(block.active_start, block.active_end + 1):
            visible_total += sheet.cell(
                row=DATA_ROW_START,
                column=block.col_start + day - 1,
            ).value or 0
    assert visible_total == result.total_hours == 18


@pytest.mark.parametrize(
    "schema",
    [
        PlanillaGenerateRequest,
        SalaryReportRequest,
        PracticePlanillaGenerateRequest,
        PracticeAttendanceBulkCreate,
        AttendanceProcessRequest,
    ],
)
@pytest.mark.parametrize("field", ["start_date", "end_date"])
def test_requests_reject_a_single_custom_period_endpoint(schema, field: str) -> None:
    payload = {"month": 5, "year": 2026, field: date(2026, 5, 1)}
    if schema is AttendanceProcessRequest:
        payload["upload_id"] = 1
    with pytest.raises(ValidationError, match="deben enviarse juntos"):
        schema(**payload)


def test_contract_outside_requested_range_pays_zero(db: Session) -> None:
    teacher = add_teacher(db, "OUTSIDE")
    add_designation(
        db,
        teacher.ci,
        contract_start=date(2026, 6, 1),
        contract_end=date(2026, 6, 30),
    )
    rows = PlanillaGenerator()._build_planilla_data(db, 5, 2026, discount_mode="full")[0]
    assert rows[0].payable_hours == 0
    assert rows[0].daily_hours == {}
