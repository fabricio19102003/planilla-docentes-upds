from __future__ import annotations

import logging
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from sqlalchemy import func, or_, text
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app.models.attendance import AttendanceRecord
from app.models.teacher import Teacher
from app.models.user import User
from app.schemas.teacher import (
    PaginatedTeachersResponse,
    TeacherAttendanceSummary,
    TeacherCreate,
    TeacherDetailResponse,
    TeacherResponse,
    TeacherUpdate,
)
from app.services.activity_logger import log_activity
from app.utils.auth import get_current_user, require_admin

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/teachers", tags=["teachers"])


@router.get("", response_model=PaginatedTeachersResponse)
def list_teachers(
    search: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=50, ge=1, le=200),
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> PaginatedTeachersResponse:
    try:
        query = db.query(Teacher)
        if search:
            term = f"%{search.strip()}%"
            query = query.filter(or_(Teacher.full_name.ilike(term), Teacher.ci.ilike(term)))

        total = query.count()
        teachers = (
            query.order_by(Teacher.full_name.asc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        return PaginatedTeachersResponse(
            items=[TeacherResponse.model_validate(teacher) for teacher in teachers],
            total=total,
            page=page,
            per_page=per_page,
        )
    except Exception as exc:
        logger.exception("Failed to load teachers: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo obtener la lista de docentes",
        ) from exc


@router.get("/{ci}", response_model=TeacherDetailResponse)
def get_teacher(
    ci: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> TeacherDetailResponse:
    # Admin can see any teacher; docente can only see their own
    if current_user.role == "docente" and current_user.teacher_ci != ci:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo podés ver tu propio perfil de docente",
        )
    try:
        teacher = (
            db.query(Teacher)
            .options(selectinload(Teacher.designations))
            .filter(Teacher.ci == ci)
            .first()
        )
        if teacher is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Docente no encontrado")

        attendance_rows = db.query(AttendanceRecord).filter(AttendanceRecord.teacher_ci == ci).all()
        summary = TeacherAttendanceSummary(
            total_records=len(attendance_rows),
            attended=sum(1 for row in attendance_rows if row.status == "ATTENDED"),
            late=sum(1 for row in attendance_rows if row.status == "LATE"),
            absent=sum(1 for row in attendance_rows if row.status == "ABSENT"),
            no_exit=sum(1 for row in attendance_rows if row.status == "NO_EXIT"),
            total_academic_hours=sum(row.academic_hours for row in attendance_rows),
        )

        payload = TeacherDetailResponse.model_validate(teacher)
        payload.attendance_summary = summary
        return payload
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to load teacher %s: %s", ci, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo obtener el docente",
        ) from exc


@router.post("", response_model=TeacherResponse, status_code=status.HTTP_201_CREATED)
def create_teacher(
    request: Request,
    payload: TeacherCreate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> TeacherResponse:
    """Create a new teacher manually."""
    try:
        existing = db.query(Teacher).filter(Teacher.ci == payload.ci).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Ya existe un docente con CI {payload.ci}",
            )

        teacher = Teacher(
            ci=payload.ci,
            full_name=payload.full_name,
            email=payload.email,
            phone=payload.phone,
            gender=payload.gender,
            external_permanent=payload.external_permanent,
            academic_level=payload.academic_level,
            profession=payload.profession,
            specialty=payload.specialty,
            bank=payload.bank,
            account_number=payload.account_number,
            sap_code=payload.sap_code,
            invoice_retention=payload.invoice_retention,
        )
        db.add(teacher)

        log_activity(
            db,
            "create_teacher",
            "teachers",
            f"Docente creado: {teacher.full_name} (CI: {teacher.ci})",
            user=current_user,
            details={"ci": teacher.ci, "full_name": teacher.full_name},
            request=request,
        )

        db.commit()
        db.refresh(teacher)

        return TeacherResponse.model_validate(teacher)
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception("Failed to create teacher: %s", exc)
        raise HTTPException(status_code=500, detail="No se pudo crear el docente") from exc


@router.put("/{ci}", response_model=TeacherResponse)
def update_teacher(
    request: Request,
    ci: str,
    payload: TeacherUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> TeacherResponse:
    """Update an existing teacher's information. Supports CI change with cascade."""
    try:
        teacher = db.query(Teacher).filter(Teacher.ci == ci).first()
        if teacher is None:
            raise HTTPException(status_code=404, detail="Docente no encontrado")

        update_data = payload.model_dump(exclude_unset=True)
        new_ci = update_data.pop("ci", None)

        # Handle CI change — must cascade to all FK references
        if new_ci and new_ci != ci:
            # Check new CI doesn't already exist
            existing = db.query(Teacher).filter(Teacher.ci == new_ci).first()
            if existing:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Ya existe un docente con CI {new_ci}",
                )

            from app.models.designation import Designation
            from sqlalchemy import text

            # Update ALL FK references via raw SQL (SQLAlchemy can't cascade PK changes)
            db.execute(text("UPDATE designations SET teacher_ci = :new WHERE teacher_ci = :old"), {"new": new_ci, "old": ci})
            db.execute(text("UPDATE attendance_records SET teacher_ci = :new WHERE teacher_ci = :old"), {"new": new_ci, "old": ci})
            db.execute(text("UPDATE biometric_records SET teacher_ci = :new WHERE teacher_ci = :old"), {"new": new_ci, "old": ci})
            db.execute(text("UPDATE detail_requests SET teacher_ci = :new WHERE teacher_ci = :old"), {"new": new_ci, "old": ci})
            db.execute(text("UPDATE users SET teacher_ci = :new WHERE teacher_ci = :old"), {"new": new_ci, "old": ci})
            # Also update the user's login CI so they can still authenticate after a CI change
            db.execute(text("UPDATE users SET ci = :new WHERE ci = :old AND role = 'docente'"), {"new": new_ci, "old": ci})

            # Update the PK itself
            db.execute(text("UPDATE teachers SET ci = :new WHERE ci = :old"), {"new": new_ci, "old": ci})
            db.flush()

            # Re-fetch with new CI
            teacher = db.query(Teacher).filter(Teacher.ci == new_ci).first()

        # Update remaining fields
        for field, value in update_data.items():
            setattr(teacher, field, value)

        log_activity(
            db,
            "update_teacher",
            "teachers",
            f"Docente actualizado: {teacher.full_name} (CI: {teacher.ci})" + (f" [CI cambiado: {ci} → {new_ci}]" if new_ci and new_ci != ci else ""),
            user=current_user,
            details={"old_ci": ci, "new_ci": new_ci or ci, "fields_updated": list(update_data.keys()) + (["ci"] if new_ci and new_ci != ci else [])},
            request=request,
        )

        db.commit()
        db.refresh(teacher)

        return TeacherResponse.model_validate(teacher)
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception("Failed to update teacher %s: %s", ci, exc)
        raise HTTPException(status_code=500, detail="No se pudo actualizar el docente") from exc


def _normalize_teacher_data(raw: dict) -> dict:
    """Normalize a single teacher record applying all business rules."""
    name = str(raw.get("full_name") or raw.get("nombre") or "").strip().upper()
    ci = str(raw.get("ci") or raw.get("ci_number") or "").strip()
    phone = str(raw.get("phone") or raw.get("telefono") or "").strip() or None
    email = str(raw.get("email") or raw.get("correo") or "").strip().lower() or None
    bank = str(raw.get("bank") or raw.get("banco") or "").strip().title() or None
    account = str(raw.get("account_number") or raw.get("cuenta") or "").strip() or None

    nit_raw = str(raw.get("nit") or "").strip().upper()
    nit = None
    invoice_retention = None
    if nit_raw in ("RETENCION", "RETENCIÓN", "RETENCION "):
        invoice_retention = "RETENCION"
    elif nit_raw and nit_raw not in ("NONE", ""):
        nit = nit_raw

    contract = str(raw.get("contract_type") or raw.get("tipo_contrato") or "").strip() or None

    return {
        "ci": ci,
        "full_name": name,
        "phone": phone,
        "email": email if email and "@" in email else None,
        "bank": bank,
        "account_number": account,
        "nit": nit,
        "invoice_retention": invoice_retention,
        "external_permanent": "SERVICIOS PROFESIONALES" if contract else None,
    }


def _parse_teacher_excel(file: UploadFile) -> list[dict]:
    """Parse teacher data from Excel file using openpyxl."""
    import io
    import openpyxl

    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Find header row (look for "NOMBRE" in first 5 rows)
    header_row = None
    for row_idx in range(1, 6):
        for cell in ws[row_idx]:
            val = str(cell.value or "").strip().upper()
            if "NOMBRE" in val:
                header_row = row_idx
                break
        if header_row:
            break

    if header_row is None:
        raise HTTPException(400, detail="No se encontró la fila de encabezados en el Excel")

    # Map columns by header content
    col_map: dict[str, int] = {}
    for cell in ws[header_row]:
        val = str(cell.value or "").strip().upper()
        col_idx = cell.column - 1  # 0-based
        if "NOMBRE" in val:
            col_map["full_name"] = col_idx
        elif "TEL" in val or "PHONE" in val:
            col_map["phone"] = col_idx
        elif "CORREO" in val or "EMAIL" in val or "MAIL" in val:
            col_map["email"] = col_idx
        elif "C.I" in val or val == "CI" or "CEDULA" in val or "CÉDULA" in val or "C.I." in val:
            col_map["ci"] = col_idx
        elif "CONTRATO" in val or "TIPO" in val:
            col_map["contract_type"] = col_idx
        elif "NIT" in val:
            col_map["nit"] = col_idx
        elif "CUENTA" in val:
            col_map["account_number"] = col_idx
        elif "BANCO" in val or "BANK" in val:
            col_map["bank"] = col_idx

    teachers = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if not row or not any(row):
            continue

        raw: dict = {}
        for field, col_idx in col_map.items():
            if col_idx < len(row):
                raw[field] = row[col_idx]

        # Skip rows without name or CI
        name = str(raw.get("full_name") or "").strip()
        ci = str(raw.get("ci") or "").strip()
        if not name or not ci:
            continue

        teachers.append(_normalize_teacher_data(raw))

    return teachers


def _parse_teacher_json(file: UploadFile) -> list[dict]:
    """Parse teacher data from JSON file."""
    import json

    content = file.file.read()
    data = json.loads(content.decode("utf-8"))

    # Support both array and {teachers: [...]} formats
    if isinstance(data, dict):
        items = data.get("teachers") or data.get("docentes") or data.get("items") or []
    elif isinstance(data, list):
        items = data
    else:
        raise HTTPException(400, detail="Formato JSON no reconocido")

    teachers = []
    for item in items:
        raw = {
            "full_name": item.get("full_name") or item.get("nombre") or item.get("docente") or item.get("name"),
            "ci": item.get("ci") or item.get("ci_number") or item.get("cedula"),
            "phone": item.get("phone") or item.get("telefono") or item.get("tel"),
            "email": item.get("email") or item.get("correo"),
            "bank": item.get("bank") or item.get("banco"),
            "account_number": item.get("account_number") or item.get("cuenta") or item.get("cuenta_bancaria"),
            "nit": item.get("nit"),
            "contract_type": item.get("contract_type") or item.get("tipo_contrato"),
        }

        name = str(raw.get("full_name") or "").strip()
        ci = str(raw.get("ci") or "").strip()
        if not name or not ci:
            continue

        teachers.append(_normalize_teacher_data(raw))

    return teachers


def _upsert_teachers(db: Session, teachers_data: list[dict]) -> tuple[int, int, int, list[str]]:
    """Insert new teachers or update existing ones. Returns (created, updated, skipped, warnings)."""
    created = 0
    updated = 0
    skipped = 0
    warnings: list[str] = []

    for data in teachers_data:
        ci = data.get("ci")
        if not ci:
            skipped += 1
            continue

        existing = db.query(Teacher).filter(Teacher.ci == ci).first()

        if existing:
            # Update existing teacher with new data (only non-None fields)
            changed = False
            for field in ["full_name", "phone", "email", "bank", "account_number", "nit", "invoice_retention", "external_permanent"]:
                new_val = data.get(field)
                if new_val is not None and new_val != getattr(existing, field):
                    setattr(existing, field, new_val)
                    changed = True
            if changed:
                updated += 1
            else:
                skipped += 1
        else:
            # Create new teacher
            teacher = Teacher(
                ci=ci,
                full_name=data["full_name"],
                phone=data.get("phone"),
                email=data.get("email"),
                bank=data.get("bank"),
                account_number=data.get("account_number"),
                nit=data.get("nit"),
                invoice_retention=data.get("invoice_retention"),
                external_permanent=data.get("external_permanent"),
            )
            db.add(teacher)
            created += 1

        db.flush()

    # Try to link TEMP teachers by name to real CIs from the list
    for data in teachers_data:
        ci = data.get("ci")
        name = data.get("full_name")
        if not ci or not name:
            continue

        temp_teacher = db.query(Teacher).filter(
            Teacher.ci.startswith("TEMP-"),
            Teacher.full_name == name,
        ).first()

        if temp_teacher:
            old_ci = temp_teacher.ci
            # Migrate all FK references to the real CI
            for table_col in [
                ("designations", "teacher_ci"),
                ("attendance_records", "teacher_ci"),
                ("biometric_records", "teacher_ci"),
                ("detail_requests", "teacher_ci"),
                ("users", "teacher_ci"),
            ]:
                table, col = table_col
                db.execute(
                    text(f"UPDATE {table} SET {col} = :new WHERE {col} = :old"),
                    {"new": ci, "old": old_ci},
                )
            db.delete(temp_teacher)
            db.flush()
            warnings.append(f"TEMP docente '{name}' vinculado a CI real {ci}")

    return created, updated, skipped, warnings


@router.post("/upload", status_code=status.HTTP_201_CREATED)
def upload_teacher_list(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Upload a teacher list from Excel or JSON. Normalizes and upserts teachers."""
    filename = file.filename or ""
    extension = Path(filename).suffix.lower()

    if extension not in {".json", ".xlsx", ".xls"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Formato no soportado: '{extension}'. Use .xlsx, .xls o .json",
        )

    try:
        if extension in {".xlsx", ".xls"}:
            teachers_data = _parse_teacher_excel(file)
        else:
            teachers_data = _parse_teacher_json(file)

        created, updated, skipped, warnings = _upsert_teachers(db, teachers_data)

        # Try to link unlinked docente users by CI match
        unlinked_users = db.query(User).filter(User.role == "docente", User.teacher_ci.is_(None)).all()
        linked_users = 0
        for user in unlinked_users:
            teacher = db.query(Teacher).filter(Teacher.ci == user.ci).first()
            if teacher:
                user.teacher_ci = teacher.ci
                linked_users += 1
        if linked_users:
            db.flush()
            logger.info("Linked %d docente users after teacher list upload", linked_users)

        log_activity(
            db,
            "upload_teacher_list",
            "upload",
            f"Lista de docentes subida: {created} nuevos, {updated} actualizados, {skipped} omitidos",
            user=current_user,
            details={"filename": filename, "created": created, "updated": updated, "skipped": skipped},
            request=request,
        )

        db.commit()

        return {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "total_processed": created + updated + skipped,
            "warnings": warnings,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception("Teacher list upload failed: %s", exc)
        raise HTTPException(400, detail="No se pudo procesar la lista de docentes") from exc
    finally:
        file.file.close()


@router.delete("/{ci}", status_code=status.HTTP_204_NO_CONTENT)
def delete_teacher(
    request: Request,
    ci: str,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Delete a teacher. This also cascades to their designations."""
    try:
        teacher = db.query(Teacher).filter(Teacher.ci == ci).first()
        if teacher is None:
            raise HTTPException(status_code=404, detail="Docente no encontrado")

        name = teacher.full_name
        log_activity(
            db,
            "delete_teacher",
            "teachers",
            f"Docente eliminado: {name} (CI: {ci})",
            user=current_user,
            details={"ci": ci, "full_name": name},
            request=request,
        )

        db.delete(teacher)
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception("Failed to delete teacher %s: %s", ci, exc)
        raise HTTPException(status_code=500, detail="No se pudo eliminar el docente") from exc
