from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.medicine_schedule_parser import normalize_group, parse_medicine_workbook


def workbook_bytes(*, metadata="PRIMER SEMESTRE | GRUPO M-01 | MAÑANA", title="1RO", seventh=False,
                   teacher="Dra. Pérez", layout="subject_teacher_activity", first_time="08:00 - 09:30"):
    wb = Workbook()
    ws = wb.active
    ws.title = "7MO" if seventh else title
    ws["A1"] = "SEPTIMO SEMESTRE | GRUPO M1" if seventh else metadata
    ws.append([])
    ws.append(["HORARIO", "LUNES", "MARTES", "MIERCOLES"])
    if layout == "teacher_subject":
        rows = [[first_time, teacher], [None, "Anatomía I"], [], ["10:00 - 11:30", teacher], [None, "Anatomía I"]]
    elif layout == "subject_only":
        rows = [[first_time, "Anatomía I"], [], [], ["10:00 - 11:30", "Anatomía I"]]
    else:
        rows = [[first_time, "Anatomía I", "RECESO", None, None, "SIDE TOTAL"], [None, teacher], [None, "TEORÍA"],
                ["10:00 - 11:30", "Anatomía I"], [None, teacher], [None, "LABORATORIO"]]
    for values in rows:
        ws.append(values)
    ws.append([None, "TOTAL HORAS", 99])
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def test_supported_block_preserves_lineage_and_complete_group():
    result = parse_medicine_workbook(workbook_bytes())
    assert not any(issue.severity == "error" for issue in result.issues)
    assert normalize_group("m 001") == "M1"
    assert len(result.offerings) == 1
    offering = result.offerings[0]
    assert (offering.category, offering.semester, offering.group_code, offering.shift) == ("regular", 1, "M1", "morning")
    assert offering.subject_raw == "Anatomía I" and offering.subject_key == "ANATOMIA I"
    assert [meeting.activity for meeting in offering.meetings] == ["theory", "laboratory"]
    assert str(offering.meetings[0].start_time) == "08:00:00"
    assert offering.meetings[0].source_cell == "B4"
    assert offering.meetings[0].raw_payload["teacher"] == "Dra. Pérez"
    assert offering.meetings[0].raw_payload["teacher_cell"] == "B5"
    assert {issue.code for issue in result.issues} == {"break_excluded"}

def test_convalidacion_and_missing_teacher_remain_explicit():
    content = workbook_bytes(metadata="CONVALIDACIÓN | GRUPO ESPECIAL | TARDE", title="Convalidación")
    result = parse_medicine_workbook(content)
    assert result.offerings[0].category == "convalidacion"
    assert (result.offerings[0].semester, result.offerings[0].group_code) == (None, "SPECIAL")
    assert result.offerings[0].meetings[1].teacher_raw == "Dra. Pérez"
    assert result.workbook_sha256 and result.parser_schema_version == "medicine-v1"
    missing = parse_medicine_workbook(workbook_bytes(teacher=None))
    assert all(meeting.teacher_raw is None for meeting in missing.offerings[0].meetings)
    assert "missing_teacher" in {issue.code for issue in missing.issues}

@pytest.mark.parametrize("layout", ["teacher_subject", "subject_only"])
def test_recognized_alternate_layouts_parse_and_noise_is_ignored(layout):
    result = parse_medicine_workbook(workbook_bytes(layout=layout))
    assert len(result.offerings) == 1 and len(result.offerings[0].meetings) == 2
    assert {meeting.activity for meeting in result.offerings[0].meetings} == {"unspecified"}
    assert "alternate_layout" in {issue.code for issue in result.issues}

def test_unambiguous_malformed_separator_is_normalized_with_warning():
    result = parse_medicine_workbook(workbook_bytes(first_time="08:00:09:30"))
    assert str(result.offerings[0].meetings[0].end_time) == "09:30:00"
    assert "normalized_time" in {issue.code for issue in result.issues}

@pytest.mark.parametrize("metadata,code", [
    ("GRUPO M1", "missing_metadata"),
    ("PRIMER SEMESTRE", "missing_metadata"),
    ("PRIMER SEMESTRE | GRUPO M0", "missing_metadata"),
    ("PRIMER SEMESTRE | SEGUNDO SEMESTRE | GRUPO M1", "contradictory_metadata"),
    ("PRIMER SEMESTRE | GRUPO M1 | TARDE", "contradictory_shift"),
])
def test_missing_or_contradictory_metadata_blocks_extraction(metadata, code):
    result = parse_medicine_workbook(workbook_bytes(metadata=metadata, title="Schedule" if metadata == "GRUPO M1" else "1RO"))
    assert result.offerings == []
    assert code in {issue.code for issue in result.issues if issue.severity == "error"}


def test_unsupported_structure_and_seventh_semester_are_not_approximated():
    wb = Workbook()
    stream = BytesIO()
    wb.save(stream)
    unsupported = parse_medicine_workbook(stream.getvalue())
    assert unsupported.offerings == [] and unsupported.issues[0].code == "unsupported_structure"
    seventh = parse_medicine_workbook(workbook_bytes(seventh=True))
    assert seventh.offerings == [] and seventh.unsupported_semesters == [7]
    assert "semester_7_unsupported" in {issue.code for issue in seventh.issues}
