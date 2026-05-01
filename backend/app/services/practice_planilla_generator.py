"""
Service: Practice Planilla Generator

Generates the monthly payroll Excel file for practice (asistencial) teachers.

Key differences from the regular PlanillaGenerator:
  - Queries designations WHERE designation_type = "practice" (not "regular")
  - Uses PRACTICE_HOURLY_RATE from app_settings (default 50 Bs)
  - Attendance source: PracticeAttendanceLog (manual) — NOT AttendanceRecord/biometric
  - No biometric check — attendance is entered manually by admin
  - Same cross-month two-block Excel layout as regular planilla
  - Output file: planilla_practicas_MM_YYYY.xlsx
  - Persists PracticePlanillaOutput (separate table from PlanillaOutput)

Payment model (same as regular, adapted for manual attendance):
  - Base pay = designation.monthly_hours × PRACTICE_HOURLY_RATE
  - When start_date/end_date provided: base = _calculate_period_hours(schedule, start, end)
  - Deduct ONLY hours where PracticeAttendanceLog.status == "absent"
  - Payable hours = max(0, base - absent_hours)
  - No biometric logic — manual logs are the single source of truth
"""
from __future__ import annotations

import calendar
import logging
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.designation import Designation
from app.models.practice_attendance import PracticeAttendanceLog
from app.models.practice_planilla import PracticePlanillaOutput
from app.models.teacher import Teacher
from app.services import app_settings_service
from app.services.planilla_generator import (
    PlanillaRow,
    MonthBlock,
    _calculate_period_hours,
    _expand_schedule_to_daily,
    _build_month_blocks,
    MONTH_NAMES,
    WEEKDAY_LETTERS,
    DAY_COL_START,
    ROW_COL_HEADERS,
    COLOR_DAY_ABSENT,
    COLOR_DAY_LATE,
    COLOR_DAY_CLASS,
    COLOR_DAY_WEEKEND,
    THIN_BORDER,
    COLOR_HEADER_BG,
    COLOR_SECTION_BG,
    COLOR_COL_HEADER_BG,
    COLOR_WEEKDAY_BG,
    COLOR_TOTAL_ROW,
    COLOR_WHITE,
    COLOR_OUTSIDE_CUTOFF,
    COLOR_SPACER,
    MEDIUM_BORDER,
    COL_SEMESTRE,
    COL_NOMBRE,
    COL_CI,
    COL_EMAIL,
    COL_PHONE,
    COL_MATERIA,
    COL_GRUPO,
    COL_TIPO_DOCENTE,
    COL_GENERO,
    COL_SAP,
    COL_FACTURA,
    COL_CUENTA,
    COL_NIVEL_ACAD,
    COL_PROFESION,
    COL_ESPECIALIDAD,
    COL_BANCO,
    DATA_ROW_START,
    ROW_TITLE,
    ROW_UNIVERSITY,
    ROW_EMPTY,
    ROW_SECTION_HEADERS,
    ROW_WEEKDAY,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data Transfer Objects
# ---------------------------------------------------------------------------


@dataclass
class PracticePlanillaResult:
    """Result returned by PracticePlanillaGenerator.generate()."""

    file_path: str
    month: int
    year: int
    total_teachers: int
    total_rows: int
    total_hours: int
    total_payment: float
    planilla_output_id: Optional[int]
    warnings: list[str] = field(default_factory=list)
    discount_mode: str = "attendance"


# ---------------------------------------------------------------------------
# Main generator class
# ---------------------------------------------------------------------------


class PracticePlanillaGenerator:
    """
    Generates the monthly practice-teacher payroll Excel file.

    Usage::

        gen = PracticePlanillaGenerator(output_dir="backend/data/output")
        result = gen.generate(db, month=3, year=2026)
        print(result.file_path)
    """

    def __init__(self, output_dir: str = "backend/data/output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate(
        self,
        db: Session,
        month: int,
        year: int,
        payment_overrides: Optional[dict[str, float]] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        discount_mode: str = "attendance",
    ) -> PracticePlanillaResult:
        """
        Generate the practice planilla Excel for a given month/year.

        Args:
            db: SQLAlchemy session
            month: Month number (1-12)
            year: Calendar year
            payment_overrides: Optional admin adjustments
            start_date: Optional start of attendance window
            end_date: Optional end of attendance window
            discount_mode: "attendance" (apply discounts) or "full" (no discounts)

        Returns:
            PracticePlanillaResult with file path and statistics
        """
        if payment_overrides is None:
            payment_overrides = {}

        logger.info(
            "PracticePlanillaGenerator.generate: month=%d year=%d start=%s end=%s discount_mode=%s",
            month, year, start_date, end_date, discount_mode,
        )

        # Step 1: Build data
        rows, warnings = self._build_planilla_data(
            db, month, year,
            start_date=start_date, end_date=end_date,
            discount_mode=discount_mode,
        )
        logger.info("Built %d practice planilla rows", len(rows))

        # Step 2: Create workbook
        blocks = _build_month_blocks(month, year, start_date, end_date)
        wb = self._create_workbook(rows, month, year, payment_overrides, blocks)

        # Step 3: Save file
        filename = f"planilla_practicas_{month:02d}_{year}.xlsx"
        file_path = self.output_dir / filename
        wb.save(str(file_path))
        logger.info("Saved practice planilla to %s", file_path)

        # Step 4: Persist to DB
        total_hours = sum(r.payable_hours for r in rows)
        total_payment = self._calculate_total_payment(rows, payment_overrides)
        unique_teachers = len({r.teacher_ci for r in rows})

        planilla_output = self._persist_planilla_output(
            db=db,
            month=month,
            year=year,
            file_path=str(file_path),
            total_teachers=unique_teachers,
            total_hours=total_hours,
            total_payment=total_payment,
            payment_overrides=payment_overrides,
            start_date=start_date,
            end_date=end_date,
            discount_mode=discount_mode,
        )

        return PracticePlanillaResult(
            file_path=str(file_path),
            month=month,
            year=year,
            total_teachers=unique_teachers,
            total_rows=len(rows),
            total_hours=total_hours,
            total_payment=total_payment,
            planilla_output_id=planilla_output.id if planilla_output else None,
            warnings=warnings,
            discount_mode=discount_mode,
        )

    # ------------------------------------------------------------------
    # Data building
    # ------------------------------------------------------------------

    def _build_planilla_data(
        self,
        db: Session,
        month: int,
        year: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        discount_mode: str = "attendance",
    ) -> tuple[list[PlanillaRow], list[str]]:
        """
        Build PlanillaRow list from practice designations + manual attendance.

        Practice attendance logic:
          - No biometric — manual PracticeAttendanceLog is the ONLY source
          - If no attendance logs exist, teachers get full pay (same as Model C)
          - Absent hours are summed directly from logs where status == "absent"
          - discount_mode="full" → skip attendance, everyone gets full pay
        """
        warnings: list[str] = []

        # Rate and active period from app_settings
        hourly_rate = app_settings_service.get_practice_hourly_rate(db)
        active_period = app_settings_service.get_active_academic_period(db)

        # Step 1: Load practice designations ONLY
        practice_designations: list[Designation] = (
            db.query(Designation)
            .filter(
                Designation.designation_type == "practice",
                Designation.academic_period == active_period,
            )
            .all()
        )

        if not practice_designations:
            warnings.append("No hay designaciones de práctica en la base de datos")
            return [], warnings

        # Step 2: Load attendance logs (skip in "full" mode)
        att_index: dict[tuple[str, int], list[PracticeAttendanceLog]] = {}

        if discount_mode == "full":
            logger.info(
                "discount_mode=full — skipping practice attendance queries, "
                "all practice teachers receive full pay"
            )
        else:
            if start_date is not None and end_date is not None:
                att_query = db.query(PracticeAttendanceLog).filter(
                    PracticeAttendanceLog.date >= start_date,
                    PracticeAttendanceLog.date <= end_date,
                )
            else:
                # Filter by month/year using date extraction
                att_query = db.query(PracticeAttendanceLog).filter(
                    PracticeAttendanceLog.date >= date(year, month, 1),
                    PracticeAttendanceLog.date <= date(
                        year, month, calendar.monthrange(year, month)[1]
                    ),
                )

            att_records = att_query.all()

            if not att_records:
                logger.info(
                    "No practice attendance logs for %d/%d — all practice docentes get full pay",
                    month, year,
                )
                warnings.append(
                    f"Sin registros de asistencia de prácticas para "
                    f"{MONTH_NAMES.get(month)} {year} — "
                    f"todos los docentes recibirán pago completo"
                )

            for rec in att_records:
                key = (rec.teacher_ci, rec.designation_id)
                att_index.setdefault(key, []).append(rec)

        # Step 3: Bulk-load teachers
        all_teacher_cis = {d.teacher_ci for d in practice_designations}
        teachers: dict[str, Teacher] = {
            t.ci: t
            for t in db.query(Teacher).filter(Teacher.ci.in_(all_teacher_cis)).all()
        }

        planilla_rows: list[PlanillaRow] = []

        # Step 4: One PlanillaRow per practice designation
        for desig in practice_designations:
            ci = desig.teacher_ci
            teacher = teachers.get(ci)

            if teacher is None:
                warnings.append(
                    f"Docente CI {ci} no encontrado en la base (designación {desig.id})"
                )
                continue

            key = (ci, desig.id)
            logs = att_index.get(key, [])

            row = self._build_row(
                teacher=teacher,
                desig=desig,
                logs=logs,
                discount_mode=discount_mode,
                hourly_rate=hourly_rate,
                start_date=start_date,
                end_date=end_date,
                month=month,
                year=year,
            )
            planilla_rows.append(row)

        # Sort: by teacher name, then subject, then group
        planilla_rows.sort(key=lambda r: (r.teacher_name, r.subject, r.group_code))

        logger.info(
            "_build_planilla_data (practice): %d rows, %d teachers",
            len(planilla_rows),
            len({r.teacher_ci for r in planilla_rows}),
        )
        return planilla_rows, warnings

    def _build_row(
        self,
        teacher: Teacher,
        desig: Designation,
        logs: list[PracticeAttendanceLog],
        discount_mode: str = "attendance",
        hourly_rate: float = 50.0,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        month: int = 0,
        year: int = 0,
    ) -> PlanillaRow:
        """
        Build a single PlanillaRow for a practice designation.

        Simpler than regular planilla — no biometric check.
        Absent hours come directly from PracticeAttendanceLog.
        """
        daily_hours: dict[date, int] = {}
        daily_status: dict[date, str] = {}
        attended_hours = 0
        absent_hours = 0
        late_count = 0
        absent_count = 0
        observations: list[str] = []

        # Process manual attendance logs
        for log in logs:
            day = log.date
            hours = log.academic_hours
            status = log.status.lower()

            if status == "absent":
                daily_hours[day] = daily_hours.get(day, 0)  # 0 hours for absent
                daily_status[day] = "ABSENT"
                absent_count += 1
                absent_hours += hours  # Manual logs already have the scheduled hours
            elif status == "late":
                daily_hours[day] = daily_hours.get(day, 0) + hours
                if daily_status.get(day, "") != "ABSENT":
                    daily_status[day] = "LATE"
                late_count += 1
                attended_hours += hours
            elif status in ("attended", "present", "justified"):
                daily_hours[day] = daily_hours.get(day, 0) + hours
                if daily_status.get(day, "") not in ("ABSENT", "LATE"):
                    daily_status[day] = "ATTENDED"
                attended_hours += hours

            if log.observation:
                observations.append(f"Día {day.day}/{day.month}: {log.observation}")

        # Fill daily_hours gaps from schedule (same as regular planilla)
        if desig.schedule_json:
            if start_date and end_date:
                eff_start, eff_end = start_date, end_date
            elif month and year:
                _, last_day = calendar.monthrange(year, month)
                eff_start = date(year, month, 1)
                eff_end = date(year, month, last_day)
            else:
                eff_start, eff_end = None, None

            if eff_start and eff_end:
                scheduled_daily = _expand_schedule_to_daily(
                    desig.schedule_json, eff_start, eff_end
                )
                for d, hrs in scheduled_daily.items():
                    if d not in daily_hours:
                        daily_hours[d] = hrs

        # Base hours calculation (same real-day-count logic as regular)
        if start_date and end_date and desig.schedule_json:
            calendar_hours = _calculate_period_hours(
                desig.schedule_json, start_date, end_date
            )
            if calendar_hours > 0:
                base_monthly_hours = calendar_hours
            else:
                fallback_raw = desig.monthly_hours or 0
                if fallback_raw > 0:
                    num_days = (end_date - start_date).days + 1
                    base_monthly_hours = round(fallback_raw * num_days / 30)
                    observations.append(
                        f"Horario no coincide con período — horas estimadas ({base_monthly_hours}h)"
                    )
                else:
                    base_monthly_hours = 0
        else:
            base_monthly_hours = desig.monthly_hours or 0

        # discount_mode="full" → zero deductions
        if discount_mode == "full":
            absent_hours = 0
            absent_count = 0

        if absent_hours > base_monthly_hours and base_monthly_hours > 0:
            logger.warning(
                "Practice designation %d (CI=%s): absent_hours=%d exceeds "
                "base_monthly_hours=%d — capping",
                desig.id, teacher.ci, absent_hours, base_monthly_hours,
            )
            absent_hours = base_monthly_hours

        payable_hours = max(0, base_monthly_hours - absent_hours)
        calculated_payment = payable_hours * hourly_rate

        # RC-IVA 13% retention
        has_retention = (teacher.invoice_retention or "").strip().upper() == "RETENCION"
        retention_rate = 0.13 if has_retention else 0.0
        retention_amount = round(calculated_payment * retention_rate, 2)
        final_payment = round(calculated_payment - retention_amount, 2)

        # Observations
        obs_parts: list[str] = []
        if discount_mode == "full":
            obs_parts.append("Modo sin descuentos — pago completo")
        if not logs and discount_mode != "full":
            obs_parts.append("Sin registro de asistencia — pago completo")
        if logs and discount_mode != "full":
            if late_count > 0:
                obs_parts.append(f"{late_count} tardanza{'s' if late_count > 1 else ''}")
            if absent_count > 0:
                obs_parts.append(f"{absent_count} ausencia{'s' if absent_count > 1 else ''}")
            if absent_hours > 0:
                obs_parts.append(f"{absent_hours}h descontadas")
        if has_retention:
            obs_parts.append("Retención RC-IVA 13%")

        return PlanillaRow(
            teacher_ci=teacher.ci,
            designation_id=desig.id,
            teacher_name=teacher.full_name,
            email=teacher.email,
            phone=teacher.phone,
            subject=desig.subject,
            semester=desig.semester,
            group_code=desig.group_code,
            teacher_type=teacher.external_permanent,
            gender=teacher.gender,
            sap_code=teacher.sap_code,
            invoice_retention=teacher.invoice_retention,
            account_number=teacher.account_number,
            academic_level=teacher.academic_level,
            profession=teacher.profession,
            specialty=teacher.specialty,
            bank=teacher.bank,
            daily_hours=daily_hours,
            daily_status=daily_status,
            base_monthly_hours=base_monthly_hours,
            absent_hours=absent_hours,
            payable_hours=payable_hours,
            total_hours=payable_hours,
            total_theory_hours=base_monthly_hours,
            total_practice_internal_hours=absent_hours,
            total_practice_external_hours=attended_hours,
            rate_per_hour=hourly_rate,
            calculated_payment=calculated_payment,
            has_retention=has_retention,
            retention_rate=retention_rate,
            retention_amount=retention_amount,
            final_payment=final_payment,
            observations=obs_parts if obs_parts else [],
            late_count=late_count,
            absent_count=absent_count,
            has_biometric=bool(logs),  # True if any manual attendance exists
        )

    # ------------------------------------------------------------------
    # Workbook creation
    # ------------------------------------------------------------------

    def _get_summary_cols(self, blocks: list[MonthBlock]) -> dict[str, int]:
        """Compute summary column positions based on the block layout.

        Same logic as regular planilla — summary cols live after the last
        day-block.
        """
        last_block = blocks[-1]
        base = last_block.col_start + last_block.days_in_month
        return {
            'total_horas': base,
            'grupo_resumen': base + 1,
            'materia_resumen': base + 2,
            'pago_hora': base + 3,
            'hrs_asignadas': base + 4,
            'hrs_descontadas': base + 5,
            'hrs_asistidas': base + 6,
            'total_pagable': base + 7,
            'pago_calculado': base + 8,
            'retencion_amt': base + 9,
            'pago_neto': base + 10,
            'pago_ajustado': base + 11,
            'observaciones': base + 12,
        }

    def _create_workbook(
        self,
        rows: list[PlanillaRow],
        month: int,
        year: int,
        payment_overrides: dict[str, float],
        blocks: list[MonthBlock],
    ) -> Workbook:
        """Create the Excel workbook with a single Planilla sheet."""
        wb = Workbook()

        ws = wb.create_sheet(title="Planilla Prácticas", index=0)
        if len(wb.worksheets) > 1:
            default_sheet = wb.worksheets[1]
            del wb[default_sheet.title]

        scols = self._get_summary_cols(blocks)
        total_cols = scols['observaciones']

        self._write_headers(ws, month, year, blocks, scols, total_cols)
        last_data_row = self._write_data_rows(
            ws, rows, month, year, payment_overrides, blocks, scols
        )
        self._write_totals_row(
            ws, rows, last_data_row + 1, payment_overrides,
            blocks, scols, total_cols,
        )
        self._apply_formatting(ws, last_data_row + 1, total_cols)

        return wb

    # ------------------------------------------------------------------
    # Header writing
    # ------------------------------------------------------------------

    def _write_headers(
        self,
        ws,
        month: int,
        year: int,
        blocks: list[MonthBlock],
        scols: dict[str, int],
        total_cols: int,
    ) -> None:
        """Write rows 1-6: title, university, empty, section headers, col headers, weekday row."""
        month_name = MONTH_NAMES.get(month, str(month)).upper()
        last_col_letter = get_column_letter(total_cols)

        # Row 1: Title
        ws.merge_cells(f"A{ROW_TITLE}:{last_col_letter}{ROW_TITLE}")
        title_cell = ws.cell(row=ROW_TITLE, column=1)
        title_cell.value = f"SIPAD — PLANILLA DOCENTES ASISTENCIALES — {month_name} {year}"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color=COLOR_WHITE)
        title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ROW_TITLE].height = 24

        # Row 2: University
        ws.merge_cells(f"A{ROW_UNIVERSITY}:{last_col_letter}{ROW_UNIVERSITY}")
        univ_cell = ws.cell(row=ROW_UNIVERSITY, column=1)
        univ_cell.value = "Universidad Privada Domingo Savio — Facultad de Medicina"
        univ_cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE)
        univ_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        univ_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ROW_UNIVERSITY].height = 18

        # Row 3: Spacer
        ws.row_dimensions[ROW_EMPTY].height = 6

        # Row 4: Section headers
        self._write_section_headers(ws, month, month_name, year, blocks, scols)

        # Row 5: Column headers
        self._write_column_headers(ws, blocks, scols)

        # Row 6: Weekday letters
        self._write_weekday_row(ws, blocks, scols, total_cols)

        # Column widths
        self._set_column_widths(ws, blocks, scols)

    def _write_section_headers(
        self, ws, month: int, month_name: str, year: int,
        blocks: list[MonthBlock], scols: dict[str, int],
    ) -> None:
        """Row 4: Merged section labels."""
        row = ROW_SECTION_HEADERS

        # DATOS DOCENTE
        ws.merge_cells(
            start_row=row, start_column=COL_SEMESTRE,
            end_row=row, end_column=COL_BANCO,
        )
        cell = ws.cell(row=row, column=COL_SEMESTRE)
        cell.value = "DATOS DOCENTE"
        self._style_section_header(cell)

        # ASISTENCIA — one merge per block
        is_cross_month = len(blocks) > 1
        for idx, block in enumerate(blocks):
            start_col = block.col_start
            end_col = block.col_start + block.days_in_month - 1
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col,
            )
            cell = ws.cell(row=row, column=start_col)
            bname = block.month_name.upper()
            if not is_cross_month:
                cell.value = f"ASISTENCIA PRÁCTICAS {bname} {block.year}"
            elif idx == 0:
                cell.value = f"ASISTENCIA {bname} {block.year} (desde día {block.active_start})"
            else:
                cell.value = f"ASISTENCIA {bname} {block.year} (hasta día {block.active_end})"
            self._style_section_header(cell)

            if idx < len(blocks) - 1:
                spacer_col = end_col + 1
                spacer_cell = ws.cell(row=row, column=spacer_col)
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        # RESUMEN Y PAGOS
        ws.merge_cells(
            start_row=row, start_column=scols['total_horas'],
            end_row=row, end_column=scols['observaciones'],
        )
        cell = ws.cell(row=row, column=scols['total_horas'])
        cell.value = "RESUMEN Y PAGOS"
        self._style_section_header(cell)

        ws.row_dimensions[row].height = 20

    def _write_column_headers(
        self, ws, blocks: list[MonthBlock], scols: dict[str, int],
    ) -> None:
        """Row 5: Actual column headers including day numbers."""
        row = ROW_COL_HEADERS

        identity_headers = [
            "Semestre", "Apellidos y Nombres", "CI",
            "Correo Electrónico", "Nro. Celular", "Materia", "Grupo",
            "Tipo Docente", "Género", "Código SAP", "Factura/Retención",
            "Nro. Cuenta", "Nivel Académico", "Profesión", "Especialidad", "Banco",
        ]
        for i, header in enumerate(identity_headers, start=1):
            cell = ws.cell(row=row, column=i)
            cell.value = header
            self._style_col_header(cell)

        # Day number headers
        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                cell = ws.cell(row=row, column=col)
                cell.value = day
                is_active = block.active_start <= day <= block.active_end
                if is_active:
                    self._style_col_header(cell, is_day=True)
                else:
                    cell.font = Font(name="Calibri", size=9, bold=False, color="A6A6A6")
                    cell.fill = PatternFill("solid", fgColor=COLOR_OUTSIDE_CUTOFF)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = THIN_BORDER

            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                spacer_cell = ws.cell(row=row, column=spacer_col)
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        # Summary column headers
        summary_headers = {
            scols['total_horas']: "Hrs\nPagables",
            scols['grupo_resumen']: "Grupo",
            scols['materia_resumen']: "Materia",
            scols['pago_hora']: "Pago\n/Hora",
            scols['hrs_asignadas']: "Hrs\nAsignadas",
            scols['hrs_descontadas']: "Hrs\nDescontadas",
            scols['hrs_asistidas']: "Hrs\nAsistidas",
            scols['total_pagable']: "Total\nPagable",
            scols['pago_calculado']: "Total Bruto\n(Bs)",
            scols['retencion_amt']: "Retención\nRC-IVA",
            scols['pago_neto']: "Pago Neto\n(Bs)",
            scols['pago_ajustado']: "Pago\nAjustado",
            scols['observaciones']: "Observaciones",
        }
        for col, header in summary_headers.items():
            cell = ws.cell(row=row, column=col)
            cell.value = header
            self._style_col_header(cell, wrap=True)

        ws.row_dimensions[row].height = 30

    def _write_weekday_row(
        self, ws, blocks: list[MonthBlock],
        scols: dict[str, int], total_cols: int,
    ) -> None:
        """Row 6: Day-of-week letters under each day column."""
        row = ROW_WEEKDAY
        block_cols: set[int] = set()

        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                block_cols.add(col)
                cell = ws.cell(row=row, column=col)
                target_date = date(block.year, block.month, day)
                cell.value = WEEKDAY_LETTERS[target_date.weekday()]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER

                is_active = block.active_start <= day <= block.active_end
                if is_active:
                    cell.font = Font(name="Calibri", size=8, bold=True, color="595959")
                    cell.fill = PatternFill("solid", fgColor=COLOR_WEEKDAY_BG)
                else:
                    cell.font = Font(name="Calibri", size=8, bold=False, color="A6A6A6")
                    cell.fill = PatternFill("solid", fgColor=COLOR_OUTSIDE_CUTOFF)

            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                block_cols.add(spacer_col)
                spacer_cell = ws.cell(row=row, column=spacer_col)
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        for col in range(1, total_cols + 1):
            if col in block_cols:
                continue
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=COLOR_WEEKDAY_BG)
            cell.border = THIN_BORDER

        ws.row_dimensions[row].height = 14

    def _set_column_widths(
        self, ws, blocks: list[MonthBlock], scols: dict[str, int],
    ) -> None:
        """Set column widths."""
        widths = {
            COL_SEMESTRE: 10, COL_NOMBRE: 28, COL_CI: 10,
            COL_EMAIL: 22, COL_PHONE: 12, COL_MATERIA: 28,
            COL_GRUPO: 8, COL_TIPO_DOCENTE: 12, COL_GENERO: 8,
            COL_SAP: 12, COL_FACTURA: 12, COL_CUENTA: 16,
            COL_NIVEL_ACAD: 14, COL_PROFESION: 16, COL_ESPECIALIDAD: 16,
            COL_BANCO: 14,
        }
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                ws.column_dimensions[get_column_letter(col)].width = 3.5
            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                ws.column_dimensions[get_column_letter(spacer_col)].width = 1.5

        summary_widths = {
            scols['total_horas']: 8,
            scols['grupo_resumen']: 8,
            scols['materia_resumen']: 22,
            scols['pago_hora']: 8,
            scols['hrs_asignadas']: 8,
            scols['hrs_descontadas']: 9,
            scols['hrs_asistidas']: 9,
            scols['total_pagable']: 8,
            scols['pago_calculado']: 12,
            scols['retencion_amt']: 11,
            scols['pago_neto']: 12,
            scols['pago_ajustado']: 12,
            scols['observaciones']: 30,
        }
        for col, width in summary_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

    # ------------------------------------------------------------------
    # Data row writing
    # ------------------------------------------------------------------

    def _write_data_rows(
        self, ws, rows: list[PlanillaRow], month: int, year: int,
        payment_overrides: dict[str, float],
        blocks: list[MonthBlock], scols: dict[str, int],
    ) -> int:
        """Write all data rows. Returns the row number of the last written row."""
        for i, data in enumerate(rows):
            row_num = DATA_ROW_START + i
            self._write_data_row(
                ws, row_num, data, payment_overrides, rows, blocks, scols,
            )

        last_row = DATA_ROW_START + len(rows) - 1
        return last_row if rows else DATA_ROW_START - 1

    def _write_data_row(
        self, ws, row_num: int, data: PlanillaRow,
        payment_overrides: dict[str, float],
        all_rows: list[PlanillaRow],
        blocks: list[MonthBlock], scols: dict[str, int],
    ) -> None:
        """Write one practice teacher data row."""
        override = self._resolve_override(
            data.teacher_ci, data.designation_id, payment_overrides
        )

        is_even = (row_num - DATA_ROW_START) % 2 == 0
        row_bg = "FFFFFF" if is_even else "F5F8FA"

        base_font = Font(name="Calibri", size=9)
        base_align = Alignment(vertical="center", wrap_text=False)
        base_fill = PatternFill("solid", fgColor=row_bg)

        def write_identity(col: int, value) -> None:
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.font = base_font
            cell.alignment = base_align
            cell.fill = base_fill
            cell.border = THIN_BORDER

        # Identity columns
        write_identity(COL_SEMESTRE, data.semester)
        write_identity(COL_NOMBRE, data.teacher_name)
        write_identity(COL_CI, data.teacher_ci)
        write_identity(COL_EMAIL, data.email)
        write_identity(COL_PHONE, data.phone)
        write_identity(COL_MATERIA, data.subject)
        write_identity(COL_GRUPO, data.group_code)
        write_identity(COL_TIPO_DOCENTE, data.teacher_type)
        write_identity(COL_GENERO, data.gender)
        write_identity(COL_SAP, data.sap_code)
        write_identity(COL_FACTURA, data.invoice_retention)
        write_identity(COL_CUENTA, data.account_number)
        write_identity(COL_NIVEL_ACAD, data.academic_level)
        write_identity(COL_PROFESION, data.profession)
        write_identity(COL_ESPECIALIDAD, data.specialty)
        write_identity(COL_BANCO, data.bank)

        # Day columns
        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                cell = ws.cell(row=row_num, column=col)

                is_active = block.active_start <= day <= block.active_end
                if not is_active:
                    cell.value = None
                    cell.fill = PatternFill("solid", fgColor=COLOR_OUTSIDE_CUTOFF)
                    cell.border = THIN_BORDER
                    continue

                target_date = date(block.year, block.month, day)
                hours = data.daily_hours.get(target_date, 0)
                status = data.daily_status.get(target_date, "")

                if status == "ABSENT":
                    fill_color = COLOR_DAY_ABSENT
                elif status == "LATE":
                    fill_color = COLOR_DAY_LATE
                elif status in ("ATTENDED", "NO_EXIT"):
                    fill_color = COLOR_DAY_CLASS
                elif hours > 0:
                    fill_color = COLOR_DAY_CLASS
                else:
                    if target_date.weekday() >= 5:
                        fill_color = "E8E8E8"
                    else:
                        fill_color = COLOR_DAY_WEEKEND

                cell.value = hours if hours > 0 else None
                cell.font = Font(name="Calibri", size=9, bold=(hours > 0))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = THIN_BORDER

            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                spacer_cell = ws.cell(row=row_num, column=spacer_col)
                spacer_cell.value = None
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        # Summary columns
        def write_summary(col: int, value, is_currency: bool = False) -> None:
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.font = base_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="EBF3FB")
            cell.border = THIN_BORDER
            if is_currency:
                cell.number_format = '#,##0.00 "Bs"'

        write_summary(scols['total_horas'], data.payable_hours)
        write_summary(scols['grupo_resumen'], data.group_code)
        write_summary(scols['materia_resumen'], data.subject)
        write_summary(scols['pago_hora'], data.rate_per_hour, is_currency=True)
        write_summary(scols['hrs_asignadas'], data.base_monthly_hours)
        write_summary(
            scols['hrs_descontadas'],
            data.absent_hours if data.absent_hours > 0 else None,
        )
        write_summary(
            scols['hrs_asistidas'],
            data.total_practice_external_hours if data.total_practice_external_hours > 0 else None,
        )
        write_summary(scols['total_pagable'], data.payable_hours)
        write_summary(scols['pago_calculado'], data.calculated_payment, is_currency=True)

        # Retention
        ret_cell = ws.cell(row=row_num, column=scols['retencion_amt'])
        ret_cell.value = data.retention_amount if data.has_retention else None
        ret_cell.font = Font(
            name="Calibri", size=9,
            color="C00000" if data.has_retention else "000000",
        )
        ret_cell.alignment = Alignment(horizontal="center", vertical="center")
        ret_cell.fill = PatternFill(
            "solid", fgColor="FFF0F0" if data.has_retention else "EBF3FB"
        )
        ret_cell.border = THIN_BORDER
        if data.has_retention:
            ret_cell.number_format = '#,##0.00 "Bs"'

        # Pago Neto
        net_cell = ws.cell(row=row_num, column=scols['pago_neto'])
        net_cell.value = data.final_payment
        net_cell.font = Font(name="Calibri", size=9, bold=True, color="1F4E79")
        net_cell.alignment = Alignment(horizontal="center", vertical="center")
        net_cell.fill = PatternFill("solid", fgColor="DEEAF1")
        net_cell.border = THIN_BORDER
        net_cell.number_format = '#,##0.00 "Bs"'

        # Pago Ajustado
        adj_cell = ws.cell(row=row_num, column=scols['pago_ajustado'])
        adj_cell.value = override if override is not None else None
        adj_cell.font = Font(
            name="Calibri", size=9, bold=(override is not None),
            color="C00000" if override is not None else "000000",
        )
        adj_cell.alignment = Alignment(horizontal="center", vertical="center")
        adj_cell.fill = PatternFill(
            "solid", fgColor="FFF0F0" if override is not None else "EBF3FB"
        )
        adj_cell.border = THIN_BORDER
        if override is not None:
            adj_cell.number_format = '#,##0.00 "Bs"'

        # Observations
        obs_cell = ws.cell(row=row_num, column=scols['observaciones'])
        obs_text = "; ".join(data.observations) if data.observations else ""
        obs_cell.value = obs_text if obs_text else None
        obs_cell.font = Font(name="Calibri", size=8, color="595959")
        obs_cell.alignment = Alignment(vertical="center", wrap_text=True)
        obs_cell.fill = base_fill
        obs_cell.border = THIN_BORDER

        ws.row_dimensions[row_num].height = 15

    # ------------------------------------------------------------------
    # Totals row
    # ------------------------------------------------------------------

    def _write_totals_row(
        self, ws, rows: list[PlanillaRow], totals_row: int,
        payment_overrides: dict[str, float],
        blocks: list[MonthBlock], scols: dict[str, int],
        total_cols: int,
    ) -> None:
        """Write the totals row at the bottom."""
        if not rows:
            return

        total_hours = sum(r.payable_hours for r in rows)
        total_assigned = sum(r.base_monthly_hours for r in rows)
        total_deducted = sum(r.absent_hours for r in rows)
        total_attended = sum(r.total_practice_external_hours for r in rows)
        total_bruto = sum(r.calculated_payment for r in rows)
        total_retention = sum(r.retention_amount for r in rows)
        total_payment = self._calculate_total_payment(rows, payment_overrides)
        total_teachers = len({r.teacher_ci for r in rows})

        totals_fill = PatternFill("solid", fgColor=COLOR_TOTAL_ROW)
        totals_font = Font(name="Calibri", size=9, bold=True)
        totals_align = Alignment(horizontal="center", vertical="center")

        # Label
        ws.merge_cells(
            start_row=totals_row, start_column=1,
            end_row=totals_row, end_column=COL_BANCO,
        )
        label_cell = ws.cell(row=totals_row, column=1)
        label_cell.value = f"TOTALES — {total_teachers} docente(s) asistencial(es)"
        label_cell.font = totals_font
        label_cell.fill = totals_fill
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.border = THIN_BORDER

        # Day columns totals
        for idx, block in enumerate(blocks):
            for day in range(1, block.days_in_month + 1):
                col = block.col_start + (day - 1)
                cell = ws.cell(row=totals_row, column=col)
                cell.font = totals_font
                cell.alignment = totals_align
                cell.border = THIN_BORDER

                is_active = block.active_start <= day <= block.active_end
                if not is_active:
                    cell.value = None
                    cell.fill = PatternFill("solid", fgColor=COLOR_OUTSIDE_CUTOFF)
                    continue

                target_date = date(block.year, block.month, day)
                day_total = sum(r.daily_hours.get(target_date, 0) for r in rows)
                cell.value = day_total if day_total > 0 else None
                cell.fill = totals_fill

            if idx < len(blocks) - 1:
                spacer_col = block.col_start + block.days_in_month
                spacer_cell = ws.cell(row=totals_row, column=spacer_col)
                spacer_cell.value = None
                spacer_cell.fill = PatternFill("solid", fgColor=COLOR_SPACER)
                spacer_cell.border = THIN_BORDER

        # Summary totals
        def write_total(col: int, value, is_currency: bool = False) -> None:
            cell = ws.cell(row=totals_row, column=col)
            cell.value = value
            cell.font = totals_font
            cell.alignment = totals_align
            cell.fill = totals_fill
            cell.border = THIN_BORDER
            if is_currency:
                cell.number_format = '#,##0.00 "Bs"'

        write_total(scols['total_horas'], total_hours)
        write_total(scols['grupo_resumen'], None)
        write_total(scols['materia_resumen'], None)
        write_total(scols['pago_hora'], None)
        write_total(scols['hrs_asignadas'], total_assigned)
        write_total(scols['hrs_descontadas'], total_deducted)
        write_total(scols['hrs_asistidas'], total_attended)
        write_total(scols['total_pagable'], total_hours)
        write_total(scols['pago_calculado'], total_bruto, is_currency=True)
        write_total(
            scols['retencion_amt'],
            total_retention if total_retention > 0 else None,
            is_currency=True,
        )
        write_total(scols['pago_neto'], total_payment, is_currency=True)
        write_total(scols['pago_ajustado'], None)
        write_total(scols['observaciones'], None)

        ws.row_dimensions[totals_row].height = 18

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------

    def _apply_formatting(self, ws, last_row: int, total_cols: int) -> None:
        """Apply freeze panes and print settings."""
        ws.freeze_panes = ws.cell(row=DATA_ROW_START, column=COL_CI + 1)
        ws.print_title_rows = f"1:{ROW_WEEKDAY}"
        ws.print_title_cols = f"A:{get_column_letter(COL_BANCO)}"
        ws.sheet_view.showGridLines = True

        if last_row >= DATA_ROW_START:
            ws.auto_filter.ref = (
                f"A{ROW_COL_HEADERS}:{get_column_letter(total_cols)}{last_row}"
            )

    # ------------------------------------------------------------------
    # DB persistence
    # ------------------------------------------------------------------

    def _persist_planilla_output(
        self,
        db: Session,
        month: int,
        year: int,
        file_path: str,
        total_teachers: int,
        total_hours: int,
        total_payment: float,
        payment_overrides: Optional[dict[str, float]] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        discount_mode: str = "attendance",
    ) -> Optional[PracticePlanillaOutput]:
        """Create or update a PracticePlanillaOutput record (upsert by month/year)."""
        try:
            existing = (
                db.query(PracticePlanillaOutput)
                .filter(
                    PracticePlanillaOutput.month == month,
                    PracticePlanillaOutput.year == year,
                )
                .first()
            )

            overrides_data = payment_overrides if payment_overrides else None

            if existing:
                existing.file_path = file_path
                existing.total_teachers = total_teachers
                existing.total_hours = total_hours
                existing.total_payment = Decimal(str(total_payment))
                existing.payment_overrides_json = overrides_data
                existing.start_date = start_date
                existing.end_date = end_date
                existing.discount_mode = discount_mode
                existing.generated_at = datetime.now()
                existing.status = "generated"
                db.flush()
                logger.info("Updated PracticePlanillaOutput id=%d", existing.id)
                return existing
            else:
                output = PracticePlanillaOutput(
                    month=month,
                    year=year,
                    file_path=file_path,
                    total_teachers=total_teachers,
                    total_hours=total_hours,
                    total_payment=Decimal(str(total_payment)),
                    payment_overrides_json=overrides_data,
                    start_date=start_date,
                    end_date=end_date,
                    discount_mode=discount_mode,
                    status="generated",
                )
                db.add(output)
                db.flush()
                logger.info("Created PracticePlanillaOutput id=%d", output.id)
                return output

        except Exception as exc:
            logger.exception("Failed to persist PracticePlanillaOutput: %s", exc)
            return None

    # ------------------------------------------------------------------
    # Style helpers
    # ------------------------------------------------------------------

    def _style_section_header(self, cell) -> None:
        cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    def _style_col_header(
        self, cell, is_day: bool = False, wrap: bool = False,
    ) -> None:
        cell.font = Font(name="Calibri", size=9, bold=True, color="1F3864")
        cell.fill = PatternFill("solid", fgColor=COLOR_COL_HEADER_BG)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=wrap,
        )
        cell.border = THIN_BORDER

    def _resolve_override(
        self, teacher_ci: str, designation_id: int,
        overrides: dict[str, float],
    ) -> Optional[float]:
        """Resolve override precedence."""
        row_key = f"{teacher_ci}:{designation_id}"
        if row_key in overrides:
            return overrides[row_key]
        if teacher_ci in overrides:
            return overrides[teacher_ci]
        return None

    def _calculate_total_payment(
        self, rows: list[PlanillaRow], payment_overrides: dict[str, float],
    ) -> float:
        """Calculate total payment with override precedence."""
        total = 0.0
        for row in rows:
            override = self._resolve_override(
                row.teacher_ci, row.designation_id, payment_overrides
            )
            if override is not None:
                total += override
            else:
                total += row.final_payment
        return total
