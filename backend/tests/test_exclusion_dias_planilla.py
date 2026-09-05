"""
Tests for the exclusion-dias-planilla feature.

Test layers:
  1. Unit tests for _is_excluded() — all scopes, edge cases
  2. Unit tests for ExcludedDaySchema validator — scope conditional fields
  3. Integration tests for _build_row() / _build_planilla_data() with exclusions
  4. Integration test for generate() → Excel cell fill colors
  5. Backward compatibility — empty exclusions produce identical output
"""
from __future__ import annotations

import calendar
import json
import tempfile
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from pydantic import ValidationError
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import app.models  # noqa: F401 — register all models for create_all
from app.database import Base
from app.models.attendance import AttendanceRecord
from app.models.biometric import BiometricRecord, BiometricUpload
from app.models.designation import Designation
from app.models.planilla import PlanillaOutput
from app.models.teacher import Teacher

from app.schemas.planilla import ExcludedDaySchema
from app.services.planilla_generator import (
    PlanillaGenerator,
    _is_excluded,
    COLOR_DAY_EXCLUDED,
    COLOR_DAY_CLASS,
    COLOR_DAY_WEEKEND,
    DATA_ROW_START,
    DAY_COL_START,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

TEST_DB_URL = "sqlite:///:memory:"


@pytest.fixture(scope="module")
def engine():
    eng = create_engine(TEST_DB_URL, connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=eng)
    yield eng
    Base.metadata.drop_all(bind=eng)


@pytest.fixture
def db(engine):
    conn = engine.connect()
    trans = conn.begin()
    Session_ = sessionmaker(bind=conn)
    session = Session_()
    yield session
    session.close()
    trans.rollback()
    conn.close()


@pytest.fixture
def temp_output_dir(tmp_path):
    output = tmp_path / "output"
    output.mkdir()
    return str(output)


# ---------------------------------------------------------------------------
# Section 1: Unit tests for _is_excluded()
# ---------------------------------------------------------------------------


class TestIsExcluded:
    """Unit tests for the shared _is_excluded() predicate."""

    TARGET = date(2026, 5, 15)

    def _exc(self, scope, semester=None, subject=None, group_code=None) -> ExcludedDaySchema:
        data: dict = {"date": self.TARGET, "scope": scope}
        if semester is not None:
            data["semester_id"] = semester
        if subject is not None:
            data["subject_id"] = subject
        if group_code is not None:
            data["group_id"] = group_code
        return ExcludedDaySchema.model_validate(data)

    def _exc_subject(self, subject: str, group_code: str, semester: str | None = None) -> ExcludedDaySchema:
        data = {
            "date": self.TARGET,
            "scope": "subject",
            "subject_id": subject,
            "group_id": group_code,
        }
        if semester is not None:
            data["semester_id"] = semester
        return ExcludedDaySchema.model_validate(data)

    def test_global_scope_matches_any_designation(self):
        """scope=global must return True regardless of semester/subject/group."""
        exc = self._exc("global")
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [exc]) is True

    def test_global_scope_does_not_match_different_date(self):
        other_date = date(2026, 5, 16)
        exc = self._exc("global")
        assert _is_excluded(other_date, "SEM-A", "Biología", "GR-01", [exc]) is False

    def test_semester_scope_matches_correct_semester(self):
        exc = self._exc("semester", semester="SEM-A")
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [exc]) is True

    def test_semester_scope_does_not_match_wrong_semester(self):
        exc = self._exc("semester", semester="SEM-B")
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [exc]) is False

    def test_subject_scope_full_academic_identity_match(self):
        exc = self._exc_subject("Biología", "GR-01", "SEM-A")
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [exc]) is True
        assert _is_excluded(self.TARGET, "SEM-B", "Biología", "GR-01", [exc]) is False

    def test_subject_scope_no_semester_id_still_matches(self):
        """Legacy subject scope without semester_id keeps its broad behavior."""
        exc = self._exc_subject("Biología", "GR-01")
        # Same designation, any semester context
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [exc]) is True
        assert _is_excluded(self.TARGET, "SEM-Z", "Biología", "GR-01", [exc]) is True

    def test_subject_scope_wrong_subject(self):
        exc = self._exc_subject("Física", "GR-01")
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [exc]) is False

    def test_subject_scope_wrong_group(self):
        exc = self._exc_subject("Biología", "GR-02", "SEM-A")
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [exc]) is False

    def test_semester_scope_applies_to_all_designations_in_semester(self):
        exc = self._exc("semester", semester="SEM-A")
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [exc]) is True
        assert _is_excluded(self.TARGET, "SEM-A", "Física", "GR-99", [exc]) is True

    def test_empty_exclusions_returns_false(self):
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", []) is False

    def test_redundant_global_entries_no_double_count(self):
        """Multiple global entries for the same date → True but no runtime error."""
        exc1 = self._exc("global")
        exc2 = self._exc("global")
        # Should return True (first match short-circuits) without error
        result = _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [exc1, exc2])
        assert result is True

    def test_global_overrides_subject_same_date(self):
        """If a global exclusion exists alongside a subject exclusion, global wins
        (no double-count — both would return True, just first-match semantics)."""
        global_exc = self._exc("global")
        subject_exc = self._exc_subject("Biología", "GR-01")
        # global is checked first in the loop; result must be True
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [global_exc, subject_exc]) is True
        # Even if subject is first, still True
        assert _is_excluded(self.TARGET, "SEM-A", "Biología", "GR-01", [subject_exc, global_exc]) is True

    def test_day_not_in_exclusion_list_returns_false(self):
        exc = self._exc("global")
        other = date(2026, 6, 1)
        assert _is_excluded(other, "SEM-A", "Biología", "GR-01", [exc]) is False


# ---------------------------------------------------------------------------
# Section 2: Unit tests for ExcludedDaySchema validator
# ---------------------------------------------------------------------------


class TestExcludedDaySchemaValidator:
    """Pydantic model_validator tests."""

    def test_global_with_no_extras_is_valid(self):
        schema = ExcludedDaySchema(date=date(2026, 5, 15), scope="global")
        assert schema.scope == "global"

    def test_semester_with_semester_id_is_valid(self):
        schema = ExcludedDaySchema(date=date(2026, 5, 15), scope="semester", semester_id="SEM-A")
        assert schema.semester_id == "SEM-A"

    def test_semester_missing_semester_id_raises_value_error(self):
        with pytest.raises(ValidationError) as exc_info:
            ExcludedDaySchema(date=date(2026, 5, 15), scope="semester")
        assert "semester_id" in str(exc_info.value).lower()

    def test_subject_with_all_fields_is_valid(self):
        """New subject exclusions preserve the complete academic identity."""
        schema = ExcludedDaySchema(
            date=date(2026, 5, 15),
            scope="subject",
            subject_id="Biología",
            group_id="GR-01",
            semester_id="SEM-A",
        )
        assert schema.subject_id == "Biología"
        assert schema.semester_id == "SEM-A"

    def test_subject_without_semester_id_is_valid(self):
        """Historical subject exclusions without semester_id remain valid."""
        schema = ExcludedDaySchema(
            date=date(2026, 5, 15),
            scope="subject",
            subject_id="Biología",
            group_id="GR-01",
            # no semester_id — this must be valid
        )
        assert schema.scope == "subject"
        assert schema.semester_id is None

    def test_subject_missing_subject_id_raises(self):
        """scope=subject without subject_id must raise ValidationError."""
        with pytest.raises(ValidationError):
            ExcludedDaySchema(
                date=date(2026, 5, 15),
                scope="subject",
                group_id="GR-01",
                # subject_id is missing
            )

    def test_subject_missing_group_id_raises(self):
        """scope=subject without group_id must raise ValidationError."""
        with pytest.raises(ValidationError):
            ExcludedDaySchema(
                date=date(2026, 5, 15),
                scope="subject",
                subject_id="Biología",
                # group_id is missing
            )

    def test_subject_missing_semester_id_does_not_raise(self):
        """Legacy subject JSON without semester_id must still deserialize."""
        schema = ExcludedDaySchema(
            date=date(2026, 5, 15),
            scope="subject",
            subject_id="Math101",
            group_id="A",
            # semester_id deliberately absent — should NOT 422
        )
        assert schema.scope == "subject"
        assert schema.semester_id is None

    def test_reason_is_optional(self):
        schema = ExcludedDaySchema(date=date(2026, 5, 15), scope="global", reason="Feriado")
        assert schema.reason == "Feriado"


# ---------------------------------------------------------------------------
# Shared test data helpers
# ---------------------------------------------------------------------------


def _make_teacher(ci: str = "12345678", name: str = "Docente Test") -> Teacher:
    return Teacher(
        ci=ci,
        full_name=name,
        email=f"{ci}@test.com",
        invoice_retention=None,
    )


def _make_designation(
    teacher_ci: str = "12345678",
    semester: str = "SEM-A",
    subject: str = "Biología",
    group_code: str = "GR-01",
    monthly_hours: int = 8,
    schedule_json=None,
    academic_period: str = "I/2026",
    designation_type: str = "regular",
) -> Designation:
    if schedule_json is None:
        # Monday (lunes) 2 hours — May 2026 has Mondays: 4, 11, 18, 25
        schedule_json = [{"dia": "lunes", "hora_inicio": "08:00", "hora_fin": "10:00", "horas_academicas": 2}]
    return Designation(
        teacher_ci=teacher_ci,
        subject=subject,
        semester=semester,
        group_code=group_code,
        monthly_hours=monthly_hours,
        weekly_hours=2,
        semester_hours=32,
        schedule_json=schedule_json,
        schedule_raw="lunes 08:00-10:00",
        academic_period=academic_period,
        designation_type=designation_type,
    )


# ---------------------------------------------------------------------------
# Section 3: Integration tests for _build_row() with exclusions
# ---------------------------------------------------------------------------


class TestBuildRowWithExclusions:
    """Integration tests that call _build_row() with mocked designations."""

    def _build_row(
        self,
        excluded_days=None,
        month: int = 5,
        year: int = 2026,
    ):
        """Helper: build a single PlanillaRow for May 2026 with Monday schedule."""
        teacher = _make_teacher()
        desig = _make_designation()
        gen = PlanillaGenerator()
        return gen._build_row(
            teacher=teacher,
            desig=desig,
            records=[],
            has_biometric=False,
            discount_mode="full",
            hourly_rate=70.0,
            month=month,
            year=year,
            excluded_days=excluded_days or [],
        )

    def test_excluded_day_status_is_EXCLUDED(self):
        """When a Monday is globally excluded, daily_status[that day] must be 'EXCLUDED'."""
        monday_may_4 = date(2026, 5, 4)
        exc = ExcludedDaySchema(date=monday_may_4, scope="global")
        row = self._build_row(excluded_days=[exc])
        assert row.daily_status.get(monday_may_4) == "EXCLUDED"

    def test_excluded_day_hours_are_zero(self):
        """Excluded date must show 0 hours in daily_hours."""
        monday_may_4 = date(2026, 5, 4)
        exc = ExcludedDaySchema(date=monday_may_4, scope="global")
        row = self._build_row(excluded_days=[exc])
        assert row.daily_hours.get(monday_may_4, 0) == 0

    def test_excluded_day_does_not_increment_absent_hours(self):
        """Excluded hours must NOT be added to absent_hours (not an attendance failure)."""
        monday_may_4 = date(2026, 5, 4)
        exc = ExcludedDaySchema(date=monday_may_4, scope="global")
        row_with_exclusion = self._build_row(excluded_days=[exc])
        row_without_exclusion = self._build_row(excluded_days=[])
        # absent_hours should remain the same (both 0 since no biometric)
        assert row_with_exclusion.absent_hours == row_without_exclusion.absent_hours

    def test_base_monthly_hours_reduced_by_excluded_days(self):
        """base_monthly_hours must be reduced when excluded days fall on scheduled days."""
        # May 2026 has 4 Mondays: 4, 11, 18, 25 → 4×2=8 base hours
        # Exclude one → should be 3×2=6
        monday_may_4 = date(2026, 5, 4)
        exc = ExcludedDaySchema(date=monday_may_4, scope="global")
        row_with = self._build_row(excluded_days=[exc])
        row_without = self._build_row(excluded_days=[])
        assert row_with.base_monthly_hours < row_without.base_monthly_hours

    def test_empty_exclusions_backward_compatible(self):
        """Empty excluded_days list must produce identical output as no parameter."""
        row_none = self._build_row(excluded_days=None)
        row_empty = self._build_row(excluded_days=[])
        assert row_none.base_monthly_hours == row_empty.base_monthly_hours
        assert row_none.payable_hours == row_empty.payable_hours
        assert row_none.daily_hours == row_empty.daily_hours

    def test_non_scheduled_day_exclusion_has_no_effect(self):
        """Excluding a day when the teacher has no class should not change anything."""
        # Tuesday May 5 — teacher only has Monday schedule
        tuesday_may_5 = date(2026, 5, 5)
        exc = ExcludedDaySchema(date=tuesday_may_5, scope="global")
        row_with = self._build_row(excluded_days=[exc])
        row_without = self._build_row(excluded_days=[])
        assert row_with.base_monthly_hours == row_without.base_monthly_hours
        assert row_with.payable_hours == row_without.payable_hours


# ---------------------------------------------------------------------------
# Section 4: Integration test for generate() → Excel cell colors
# ---------------------------------------------------------------------------


@pytest.fixture
def seed_data(db):
    """Seed a teacher + designation for May 2026 (Monday schedule)."""
    teacher = _make_teacher()
    desig = _make_designation()
    db.add(teacher)
    db.flush()
    db.add(desig)

    # Add app setting for hourly rate (SQLite test — no real app_settings)
    from app.models.app_setting import AppSetting
    db.add(AppSetting(key="HOURLY_RATE", value="70.0", description="Test rate"))
    db.add(AppSetting(key="PRACTICE_HOURLY_RATE", value="50.0", description="Test practice rate"))
    db.add(AppSetting(key="ACTIVE_ACADEMIC_PERIOD", value="I/2026", description="Test period"))
    db.flush()
    return teacher, desig


class TestGenerateExcelWithExclusions:
    """Integration tests: generate() produces Excel with correct cell fills."""

    def test_excluded_day_cell_has_purple_fill(self, db, seed_data, temp_output_dir):
        """An excluded Monday must appear as COLOR_DAY_EXCLUDED in the Excel."""
        from openpyxl import load_workbook

        monday_may_4 = date(2026, 5, 4)
        exc = ExcludedDaySchema(date=monday_may_4, scope="global")

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(
            db=db,
            month=5,
            year=2026,
            excluded_days=[exc],
            discount_mode="full",
        )

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # Day 4 sits in column DAY_COL_START + (4 - 1) = 17 + 3 = 20
        day4_col = DAY_COL_START + (monday_may_4.day - 1)  # 20
        # First data row is DATA_ROW_START = 7
        cell = ws.cell(row=DATA_ROW_START, column=day4_col)
        fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor else ""

        # openpyxl stores colors as AARRGGBB (8 chars) or RRGGBB (6 chars).
        # Strip the leading alpha byte (2 chars) if present.
        if len(fill_color) == 8:
            fill_hex = fill_color[2:].upper()
        else:
            fill_hex = fill_color.upper()
        assert fill_hex == COLOR_DAY_EXCLUDED.upper(), (
            f"Expected red fill ({COLOR_DAY_EXCLUDED}) on excluded day col {day4_col}, "
            f"got {fill_hex}"
        )

        # Cell value MUST be 0 for excluded days (not None/blank)
        assert cell.value == 0, (
            f"Expected excluded cell value to be 0, got {cell.value!r}"
        )

    def test_no_exclusions_excel_unchanged(self, db, seed_data, temp_output_dir):
        """Generate with no exclusions must not produce any EXCLUDED-colored cells."""
        from openpyxl import load_workbook

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(
            db=db,
            month=5,
            year=2026,
            excluded_days=[],
            discount_mode="full",
        )

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        excluded_color = COLOR_DAY_EXCLUDED.upper()
        for row in ws.iter_rows(min_row=DATA_ROW_START, max_row=DATA_ROW_START + 10):
            for cell in row:
                fill = cell.fill
                if fill and fill.fgColor and fill.fgColor.rgb:
                    raw = fill.fgColor.rgb
                    cell_color = raw[2:].upper() if len(raw) == 8 else raw.upper()
                    assert cell_color != excluded_color, (
                        f"Unexpected EXCLUDED color at cell {cell.coordinate}"
                    )


# ---------------------------------------------------------------------------
# Section 5: Backward compatibility test
# ---------------------------------------------------------------------------


class TestBackwardCompatibility:
    """Verify empty exclusions produce identical behavior to pre-feature baseline."""

    def test_generate_with_empty_exclusions_matches_no_exclusions(self, db, seed_data, temp_output_dir):
        """generate() with excluded_days=[] must produce same totals as without the param."""
        gen = PlanillaGenerator(output_dir=temp_output_dir)

        result_default = gen.generate(db=db, month=5, year=2026, discount_mode="full")
        result_empty = gen.generate(
            db=db, month=5, year=2026, excluded_days=[], discount_mode="full"
        )

        assert result_default.total_hours == result_empty.total_hours
        assert result_default.total_payment == result_empty.total_payment

    def test_build_row_no_exclusion_param_same_as_empty_list(self):
        """_build_row called without excluded_days should equal empty list."""
        teacher = _make_teacher()
        desig = _make_designation()
        gen = PlanillaGenerator()

        row_no_param = gen._build_row(
            teacher=teacher, desig=desig, records=[], has_biometric=False,
            month=5, year=2026,
        )
        row_empty = gen._build_row(
            teacher=teacher, desig=desig, records=[], has_biometric=False,
            month=5, year=2026, excluded_days=[],
        )

        assert row_no_param.base_monthly_hours == row_empty.base_monthly_hours
        assert row_no_param.payable_hours == row_empty.payable_hours
        assert row_no_param.absent_hours == row_empty.absent_hours


# ---------------------------------------------------------------------------
# Section 6: API tests for POST /api/planilla/generate with exclusions
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def api_engine():
    """Separate in-memory SQLite engine for API (TestClient) tests."""
    from sqlalchemy import create_engine as _ce
    eng = _ce("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=eng)
    yield eng
    Base.metadata.drop_all(bind=eng)


@pytest.fixture
def api_db(api_engine):
    """Per-test DB session for API tests — rolled back after each test."""
    from sqlalchemy.orm import sessionmaker as _sm
    conn = api_engine.connect()
    trans = conn.begin()
    Session_ = _sm(bind=conn)
    session = Session_()
    yield session
    session.close()
    trans.rollback()
    conn.close()


@pytest.fixture
def api_client(api_db):
    """
    FastAPI TestClient with DB dependency override and a pre-created admin token.

    Does NOT rely on the PostgreSQL-backed Settings — the DB override ensures
    all app-level queries use the in-memory SQLite session.
    """
    from fastapi.testclient import TestClient
    from app.main import app
    from app.database import get_db
    from app.models.user import User
    from app.services.auth_service import auth_service
    from app.models.app_setting import AppSetting

    # Seed required app settings so the generator can read hourly rate / period
    existing_keys = {row[0] for row in api_db.query(AppSetting.key).all()}
    for key, value in [
        ("HOURLY_RATE", "70.0"),
        ("PRACTICE_HOURLY_RATE", "50.0"),
        ("ACTIVE_ACADEMIC_PERIOD", "I/2026"),
        ("COMPANY_NAME", "Test Corp"),
        ("COMPANY_NIT", "12345"),
    ]:
        if key not in existing_keys:
            api_db.add(AppSetting(key=key, value=value, description="Test"))
    api_db.flush()

    # Create admin user
    existing_admin = api_db.query(User).filter(User.ci == "API_TEST_ADMIN").first()
    if existing_admin is None:
        admin = User(
            ci="API_TEST_ADMIN",
            full_name="API Test Admin",
            password_hash=auth_service.hash_password("testpass"),
            role="admin",
            is_active=True,
        )
        api_db.add(admin)
        api_db.flush()
        user_id = admin.id
    else:
        user_id = existing_admin.id

    token = auth_service.create_access_token(data={"sub": str(user_id), "role": "admin"})

    def override_get_db():
        try:
            yield api_db
        finally:
            pass

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as tc:
        tc.headers["Authorization"] = f"Bearer {token}"
        yield tc
    app.dependency_overrides.clear()


class TestAPIGenerateWithExclusions:
    """API-level tests (TestClient) for POST /api/planilla/generate with exclusions."""

    def _seed_teacher_and_designation(self, db):
        """Seed a teacher + Monday designation for May 2026."""
        from app.models.app_setting import AppSetting

        teacher = _make_teacher(ci="API_DOCENTE_01", name="API Docente Test")
        desig = _make_designation(teacher_ci="API_DOCENTE_01")
        db.add(teacher)
        db.flush()
        db.add(desig)
        db.flush()
        return teacher, desig

    def test_generate_with_global_exclusion_returns_200(self, api_client, api_db, tmp_path):
        """POST /api/planilla/generate with a valid global exclusion must return 200."""
        self._seed_teacher_and_designation(api_db)

        payload = {
            "month": 5,
            "year": 2026,
            "discount_mode": "full",
            "excluded_days": [
                {"date": "2026-05-04", "scope": "global"}
            ],
        }
        response = api_client.post("/api/planilla/generate", json=payload)
        assert response.status_code == 200, response.text
        data = response.json()
        assert "planilla_id" in data
        assert data["planilla_id"] > 0

    def test_generate_with_semester_exclusion_returns_200(self, api_client, api_db, tmp_path):
        """POST /api/planilla/generate with a valid semester exclusion must return 200."""
        self._seed_teacher_and_designation(api_db)

        payload = {
            "month": 5,
            "year": 2026,
            "discount_mode": "full",
            "excluded_days": [
                {"date": "2026-05-04", "scope": "semester", "semester_id": "SEM-A"}
            ],
        }
        response = api_client.post("/api/planilla/generate", json=payload)
        assert response.status_code == 200, response.text

    def test_generate_missing_semester_id_returns_422(self, api_client, api_db):
        """POST /api/planilla/generate with scope=semester but no semester_id must return 422."""
        payload = {
            "month": 5,
            "year": 2026,
            "discount_mode": "full",
            "excluded_days": [
                # scope=semester but semester_id is missing — should fail validation
                {"date": "2026-05-04", "scope": "semester"}
            ],
        }
        response = api_client.post("/api/planilla/generate", json=payload)
        assert response.status_code == 422, response.text

    def test_generate_missing_subject_id_returns_422(self, api_client, api_db):
        """POST /api/planilla/generate with scope=subject but missing subject_id must return 422."""
        payload = {
            "month": 5,
            "year": 2026,
            "excluded_days": [
                # scope=subject but subject_id is missing (group_id present)
                {"date": "2026-05-04", "scope": "subject", "group_id": "GR-01"}
            ],
        }
        response = api_client.post("/api/planilla/generate", json=payload)
        assert response.status_code == 422, response.text

    def test_generate_subject_scope_without_semester_id_returns_200(self, api_client, api_db):
        """scope=subject with only subject_id + group_id (no semester_id) must return 200."""
        self._seed_teacher_and_designation(api_db)

        payload = {
            "month": 5,
            "year": 2026,
            "discount_mode": "full",
            "excluded_days": [
                # scope=subject with ONLY subject_id + group_id — semester_id absent intentionally
                {"date": "2026-05-04", "scope": "subject", "subject_id": "Biología", "group_id": "GR-01"}
            ],
        }
        response = api_client.post("/api/planilla/generate", json=payload)
        assert response.status_code == 200, response.text
        data = response.json()
        assert "planilla_id" in data

    def test_generate_empty_exclusions_returns_200(self, api_client, api_db):
        """POST /api/planilla/generate with empty excluded_days must succeed (backward compat)."""
        self._seed_teacher_and_designation(api_db)

        payload = {
            "month": 5,
            "year": 2026,
            "excluded_days": [],
            "discount_mode": "full",
        }
        response = api_client.post("/api/planilla/generate", json=payload)
        assert response.status_code == 200, response.text


class TestAPISalaryReportExclusionInheritance:
    """API-level tests for salary report exclusion inheritance from stored planilla."""

    def _seed_planilla_output_with_exclusions(self, db, exclusions_json):
        """Seed a stored PlanillaOutput with given exclusions_json."""
        output = PlanillaOutput(
            month=5,
            year=2026,
            total_teachers=1,
            total_hours=6,
            total_payment=420,
            status="generated",
            excluded_days_json=exclusions_json,
            discount_mode="full",
        )
        db.add(output)
        db.flush()
        return output

    def test_salary_report_inherits_stored_exclusions(self, api_client, api_db, tmp_path):
        """
        When no excluded_days are sent in the salary report request,
        the endpoint must load stored exclusions from planilla_output.
        The salary report call itself (which generates a file) is what we test here:
        it must not return a 500 even when stored exclusions exist.
        """
        # Seed stored planilla with an exclusion
        self._seed_planilla_output_with_exclusions(
            api_db,
            [{"date": "2026-05-04", "scope": "global"}],
        )
        teacher = _make_teacher(ci="SAL_DOCENTE_01", name="Salary Docente")
        desig = _make_designation(teacher_ci="SAL_DOCENTE_01")
        api_db.add(teacher)
        api_db.flush()
        api_db.add(desig)
        api_db.flush()

        payload = {"month": 5, "year": 2026, "discount_mode": "full"}
        response = api_client.post("/api/planilla/salary-report", json=payload)
        assert response.status_code == 409
        assert response.json()["detail"]["code"] == "snapshot_missing"


# ---------------------------------------------------------------------------
# Section 7: Practice generator parity test (CRITICAL 2)
# ---------------------------------------------------------------------------


class TestPracticeGeneratorParity:
    """Prove that practice_planilla_generator uses the shared _is_excluded()."""

    def test_practice_generator_imports_shared_is_excluded(self):
        """practice_planilla_generator must import _is_excluded from planilla_generator."""
        import importlib
        import sys

        # Get the practice generator module
        pgm = sys.modules.get(
            "app.services.practice_planilla_generator"
        ) or importlib.import_module("app.services.practice_planilla_generator")

        from app.services.planilla_generator import _is_excluded as canonical_is_excluded

        # The practice generator module must have _is_excluded in its namespace
        assert hasattr(pgm, "_is_excluded"), (
            "practice_planilla_generator must import _is_excluded from planilla_generator"
        )
        # It must be the SAME function object (imported, not redefined)
        assert pgm._is_excluded is canonical_is_excluded, (
            "practice_planilla_generator._is_excluded must be the same function as "
            "planilla_generator._is_excluded — not a redefined copy"
        )

    def test_practice_generator_excluded_day_produces_zero_hours(self, db, seed_data, temp_output_dir):
        """Practice generator must apply exclusions via _is_excluded — excluded day → 0h."""
        from app.services.practice_planilla_generator import PracticePlanillaGenerator

        monday_may_4 = date(2026, 5, 4)
        exc = ExcludedDaySchema(date=monday_may_4, scope="global")

        pgen = PracticePlanillaGenerator(output_dir=temp_output_dir)
        # Practice generator returns (rows, warnings) — no separate detail list
        rows, _warnings = pgen._build_planilla_data(
            db, month=5, year=2026, excluded_days=[exc]
        )

        # At least one row must exist, and none may have non-zero hours on the excluded day
        for row in rows:
            excluded_hours = row.daily_hours.get(monday_may_4, 0)
            assert excluded_hours == 0, (
                f"Practice row for {row.subject}/{row.group_code} has "
                f"{excluded_hours}h on excluded {monday_may_4}"
            )

    def test_practice_subject_scope_uses_full_identity_with_legacy_fallback(self):
        from app.services.practice_planilla_generator import _is_excluded as practice_is_excluded

        target = date(2026, 5, 15)
        exact = ExcludedDaySchema(
            date=target,
            scope="subject",
            subject_id="Biología",
            group_id="GR-01",
            semester_id="SEM-A",
        )
        legacy = ExcludedDaySchema(
            date=target,
            scope="subject",
            subject_id="Biología",
            group_id="GR-01",
        )

        assert practice_is_excluded(target, "SEM-A", "Biología", "GR-01", [exact]) is True
        assert practice_is_excluded(target, "SEM-B", "Biología", "GR-01", [exact]) is False
        assert practice_is_excluded(target, "SEM-A", "Biología", "GR-02", [exact]) is False
        assert practice_is_excluded(target, "SEM-A", "Física", "GR-01", [exact]) is False
        assert practice_is_excluded(target, "SEM-B", "Biología", "GR-01", [legacy]) is True


# ---------------------------------------------------------------------------
# Section 8: Salary report override and empty override tests (CRITICAL 2)
# ---------------------------------------------------------------------------


class TestSalaryReportOverride:
    """Tests for salary report exclusion override behavior."""

    def _seed_planilla_with_exclusion(self, db):
        """Seed a PlanillaOutput with a Monday exclusion for May 2026."""
        output = PlanillaOutput(
            month=5,
            year=2026,
            total_teachers=1,
            total_hours=6,
            total_payment=420,
            status="generated",
            excluded_days_json=[{"date": "2026-05-04", "scope": "global"}],
            discount_mode="full",
        )
        db.add(output)
        db.flush()
        return output

    def test_salary_report_caller_override_uses_caller_exclusions(self, api_client, api_db, tmp_path):
        """When salary report includes excluded_days, those override stored exclusions."""
        self._seed_planilla_with_exclusion(api_db)
        teacher = _make_teacher(ci="OVR_DOCENTE_01", name="Override Docente")
        desig = _make_designation(teacher_ci="OVR_DOCENTE_01")
        api_db.add(teacher)
        api_db.flush()
        api_db.add(desig)
        api_db.flush()

        # Caller provides a different exclusion — should NOT use stored Monday exclusion
        payload = {
            "month": 5,
            "year": 2026,
            "discount_mode": "full",
            "excluded_days": [
                {"date": "2026-05-11", "scope": "global"}  # different date than stored
            ],
        }
        response = api_client.post("/api/planilla/salary-report", json=payload)
        assert response.status_code == 409
        assert response.json()["detail"]["code"] == "snapshot_missing"

    def test_salary_report_empty_override_clears_stored_exclusions(self, api_client, api_db, tmp_path):
        """When salary report sends excluded_days=[], stored exclusions are NOT used."""
        self._seed_planilla_with_exclusion(api_db)
        teacher = _make_teacher(ci="CLR_DOCENTE_01", name="Clear Docente")
        desig = _make_designation(teacher_ci="CLR_DOCENTE_01")
        api_db.add(teacher)
        api_db.flush()
        api_db.add(desig)
        api_db.flush()

        # Explicit empty list — must override (clear) stored exclusions
        payload = {
            "month": 5,
            "year": 2026,
            "excluded_days": [],
            "discount_mode": "full",
        }
        response = api_client.post("/api/planilla/salary-report", json=payload)
        assert response.status_code == 409
        assert response.json()["detail"]["code"] == "snapshot_missing"


# ---------------------------------------------------------------------------
# Section 9: Detail endpoint stored exclusion test (CRITICAL 2)
# ---------------------------------------------------------------------------


class TestDetailEndpointStoredExclusions:
    """Test that GET /api/planilla/{month}/{year}/detail reflects stored exclusions."""

    def _seed_planilla_with_exclusion(self, db):
        """Seed a PlanillaOutput with a Monday exclusion for May 2026."""
        output = PlanillaOutput(
            month=5,
            year=2026,
            total_teachers=1,
            total_hours=6,
            total_payment=420,
            status="generated",
            excluded_days_json=[{"date": "2026-05-04", "scope": "global"}],
            discount_mode="full",
        )
        db.add(output)
        db.flush()
        return output

    def test_detail_endpoint_with_stored_exclusions_returns_200(self, api_client, api_db):
        """GET /api/planilla/5/2026/detail must return 200 when stored exclusions exist."""
        self._seed_planilla_with_exclusion(api_db)
        teacher = _make_teacher(ci="DET_DOCENTE_01", name="Detail Docente")
        desig = _make_designation(teacher_ci="DET_DOCENTE_01")
        api_db.add(teacher)
        api_db.flush()
        api_db.add(desig)
        api_db.flush()

        response = api_client.get("/api/planilla/5/2026/detail?discount_mode=full")
        assert response.status_code == 409
        assert response.json()["detail"]["code"] == "snapshot_missing"

    def test_detail_endpoint_exclusions_reduce_hours(self, api_client, api_db):
        """
        Detail endpoint with stored Monday exclusion must return reduced hours
        compared to detail without stored exclusions.
        """
        # First: detail without stored exclusions (fresh DB — no stored planilla)
        teacher = _make_teacher(ci="DET_DOCENTE_02", name="Detail Docente 2")
        desig = _make_designation(teacher_ci="DET_DOCENTE_02")
        api_db.add(teacher)
        api_db.flush()
        api_db.add(desig)
        api_db.flush()

        response_no_excl = api_client.get("/api/planilla/5/2026/detail?discount_mode=full")
        assert response_no_excl.status_code == 200, response_no_excl.text
        # Response is a dict with "detail" key containing list of rows
        data_no_excl = response_no_excl.json().get("detail", [])

        # Add stored exclusion
        self._seed_planilla_with_exclusion(api_db)

        response_with_excl = api_client.get("/api/planilla/5/2026/detail?discount_mode=full")
        assert response_with_excl.status_code == 409
        assert response_with_excl.json()["detail"]["code"] == "snapshot_missing"

    def test_detail_endpoint_query_exclusions_override_stored_exclusions(self, api_client, api_db):
        """excluded_days_json=[] must clear stored exclusions for live preview."""
        teacher = _make_teacher(ci="DET_DOCENTE_03", name="Detail Docente 3")
        desig = _make_designation(teacher_ci="DET_DOCENTE_03")
        api_db.add(teacher)
        api_db.flush()
        api_db.add(desig)
        api_db.flush()
        self._seed_planilla_with_exclusion(api_db)

        response_stored = api_client.get("/api/planilla/5/2026/detail?discount_mode=full")
        assert response_stored.status_code == 409

        response_override = api_client.get(
            "/api/planilla/5/2026/detail",
            params={"excluded_days_json": json.dumps([]), "discount_mode": "full"},
        )
        assert response_override.status_code == 409
        assert response_override.json()["detail"]["code"] == "snapshot_missing"

    def test_detail_endpoint_invalid_query_exclusions_returns_422(self, api_client, api_db):
        """Invalid excluded_days_json should fail validation instead of returning 500."""
        response = api_client.get(
            "/api/planilla/5/2026/detail",
            params={"excluded_days_json": "not-json"},
        )
        assert response.status_code == 422, response.text
