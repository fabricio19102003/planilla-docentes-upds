"""Router: Practice Planilla

Endpoints for generating, viewing, and downloading the practice-teacher
(docentes asistenciales) payroll.  Separate from the regular planilla
router because practice teachers have their own rate, attendance source,
and output table.
"""
from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.designation import Designation
from app.models.practice_planilla import PracticePlanillaOutput
from app.models.teacher import Teacher
from app.models.user import User
from app.schemas.practice_planilla import (
    PracticePlanillaGenerateRequest,
    PracticePlanillaGenerateResponse,
    PracticePlanillaOutputResponse,
)
from app.services import app_settings_service
from app.services.practice_planilla_generator import PracticePlanillaGenerator
from app.services.activity_logger import log_activity
from app.utils.auth import require_admin

MONTH_NAMES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/practice-planilla", tags=["practice-planilla"])


def _output_dir() -> Path:
    path = Path(__file__).resolve().parents[2] / "data" / "output"
    path.mkdir(parents=True, exist_ok=True)
    return path


# ------------------------------------------------------------------
# POST /api/practice-planilla/generate
# ------------------------------------------------------------------

@router.post("/generate", response_model=PracticePlanillaGenerateResponse)
def generate_practice_planilla(
    payload: PracticePlanillaGenerateRequest,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> PracticePlanillaGenerateResponse:
    """Generate the practice-teacher planilla for a given month/year."""
    try:
        logger.info(
            "Generating practice planilla for %02d/%d", payload.month, payload.year
        )
        generator = PracticePlanillaGenerator(output_dir=str(_output_dir()))
        result = generator.generate(
            db=db,
            month=payload.month,
            year=payload.year,
            payment_overrides=payload.payment_overrides,
            start_date=payload.start_date,
            end_date=payload.end_date,
            discount_mode=payload.discount_mode,
        )

        log_activity(
            db,
            "generate_practice_planilla",
            "practice-planilla",
            f"Planilla prácticas generada: {MONTH_NAMES.get(payload.month, str(payload.month))} {payload.year}",
            user=current_user,
            details={
                "month": payload.month,
                "year": payload.year,
                "total_teachers": result.total_teachers,
                "total_hours": result.total_hours,
                "total_payment": float(result.total_payment),
                "discount_mode": payload.discount_mode,
            },
            request=request,
        )

        db.commit()

        if result.planilla_output_id is None:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="No se pudo registrar la planilla de prácticas generada",
            )

        return PracticePlanillaGenerateResponse(
            planilla_id=result.planilla_output_id,
            month=result.month,
            year=result.year,
            file_path=result.file_path,
            total_teachers=result.total_teachers,
            total_hours=result.total_hours,
            total_payment=result.total_payment,
            warnings=result.warnings,
            discount_mode=result.discount_mode,
        )
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception("Practice planilla generation failed: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo generar la planilla de prácticas",
        ) from exc


# ------------------------------------------------------------------
# GET /api/practice-planilla/history
# ------------------------------------------------------------------

@router.get("/history", response_model=list[PracticePlanillaOutputResponse])
def practice_planilla_history(
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[PracticePlanillaOutputResponse]:
    """List all generated practice planillas, ordered by most recent first."""
    try:
        rows = (
            db.query(PracticePlanillaOutput)
            .order_by(PracticePlanillaOutput.generated_at.desc())
            .all()
        )
        return [PracticePlanillaOutputResponse.model_validate(row) for row in rows]
    except Exception as exc:
        logger.exception("Failed to load practice planilla history: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo obtener el historial de planillas de prácticas",
        ) from exc


# ------------------------------------------------------------------
# GET /api/practice-planilla/{month}/{year}/detail
# ------------------------------------------------------------------

@router.get("/{month}/{year}/detail")
def get_practice_planilla_detail(
    month: int,
    year: int,
    start_date: date | None = None,
    end_date: date | None = None,
    discount_mode: Literal["attendance", "full"] = Query("attendance"),
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Return detailed practice planilla breakdown per teacher/designation."""
    try:
        generator = PracticePlanillaGenerator()
        rows, warnings = generator._build_planilla_data(
            db, month=month, year=year,
            start_date=start_date, end_date=end_date,
            discount_mode=discount_mode,
        )

        detail = []
        for row in rows:
            detail.append({
                "teacher_ci": row.teacher_ci,
                "teacher_name": row.teacher_name,
                "subject": row.subject,
                "semester": row.semester,
                "group_code": row.group_code,
                "base_monthly_hours": row.base_monthly_hours,
                "absent_hours": row.absent_hours,
                "payable_hours": row.payable_hours,
                "rate_per_hour": row.rate_per_hour,
                "calculated_payment": row.calculated_payment,
                "has_retention": row.has_retention,
                "retention_amount": row.retention_amount,
                "final_payment": row.final_payment,
                "has_biometric": row.has_biometric,
                "late_count": row.late_count,
                "absent_count": row.absent_count,
                "observations": row.observations,
            })

        # Group by teacher for totals
        teacher_totals: dict = {}
        for row in rows:
            ci = row.teacher_ci
            if ci not in teacher_totals:
                teacher_totals[ci] = {
                    "teacher_ci": ci,
                    "teacher_name": row.teacher_name,
                    "total_base_hours": 0,
                    "total_absent_hours": 0,
                    "total_payable_hours": 0,
                    "total_payment": 0.0,
                    "designation_count": 0,
                    "has_retention": row.has_retention,
                    "retention_amount": 0.0,
                    "final_payment": 0.0,
                }
            teacher_totals[ci]["total_base_hours"] += row.base_monthly_hours
            teacher_totals[ci]["total_absent_hours"] += row.absent_hours
            teacher_totals[ci]["total_payable_hours"] += row.payable_hours
            teacher_totals[ci]["total_payment"] += row.calculated_payment
            teacher_totals[ci]["retention_amount"] += row.retention_amount
            teacher_totals[ci]["final_payment"] += row.final_payment
            teacher_totals[ci]["designation_count"] += 1

        computed_total = sum(r.final_payment for r in rows)

        return {
            "month": month,
            "year": year,
            "total_teachers": len(teacher_totals),
            "total_designations": len(detail),
            "total_payment": computed_total,
            "detail": detail,
            "teacher_totals": list(teacher_totals.values()),
            "warnings": warnings,
        }
    except Exception as exc:
        logger.exception("Failed to load practice planilla detail: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo obtener el detalle de la planilla de prácticas",
        ) from exc


# ------------------------------------------------------------------
# GET /api/practice-planilla/{planilla_id}/download
# ------------------------------------------------------------------

@router.get("/{planilla_id}/download")
def download_practice_planilla(
    planilla_id: int,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> FileResponse:
    """Download the Excel file for a generated practice planilla."""
    try:
        planilla = (
            db.query(PracticePlanillaOutput)
            .filter(PracticePlanillaOutput.id == planilla_id)
            .first()
        )
        if planilla is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Planilla de prácticas no encontrada",
            )

        file_path = Path(planilla.file_path or "")
        if not planilla.file_path or not file_path.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Archivo de planilla de prácticas no encontrado",
            )

        return FileResponse(
            path=file_path,
            filename=file_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Practice planilla download failed: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo descargar la planilla de prácticas",
        ) from exc


# ------------------------------------------------------------------
# POST /api/practice-planilla/salary-report
# ------------------------------------------------------------------

@router.post("/salary-report")
def generate_practice_salary_report(
    payload: PracticePlanillaGenerateRequest,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> FileResponse:
    """Generate and return the salary report Excel for practice teachers.

    Reuses PracticePlanillaGenerator to build rows, then formats them
    using the same salary-report template style.
    """
    try:
        from app.services.practice_planilla_generator import PracticePlanillaGenerator as PPG

        # Resolve discount_mode from stored planilla if available
        stored = (
            db.query(PracticePlanillaOutput)
            .filter(
                PracticePlanillaOutput.month == payload.month,
                PracticePlanillaOutput.year == payload.year,
            )
            .order_by(PracticePlanillaOutput.generated_at.desc())
            .first()
        )
        effective_mode = payload.discount_mode
        effective_sd = payload.start_date
        effective_ed = payload.end_date
        if stored is not None:
            if stored.discount_mode in ("attendance", "full"):
                if payload.discount_mode == "attendance" and stored.discount_mode == "full":
                    effective_mode = "full"
            if effective_sd is None and stored.start_date is not None:
                effective_sd = stored.start_date
            if effective_ed is None and stored.end_date is not None:
                effective_ed = stored.end_date

        # Build rows using practice generator
        gen = PPG()
        rows, _warnings = gen._build_planilla_data(
            db,
            month=payload.month,
            year=payload.year,
            start_date=effective_sd,
            end_date=effective_ed,
            discount_mode=effective_mode,
        )
        rows.sort(key=lambda r: (r.teacher_name, r.subject, r.group_code))

        # Build salary report using the shared SalaryReportGenerator template
        from app.services.salary_report_generator import SalaryReportGenerator

        output_dir = _output_dir()
        salary_gen = SalaryReportGenerator(output_dir=str(output_dir))

        # Bulk-load teachers
        cis = {r.teacher_ci for r in rows}
        teachers: dict[str, Teacher] = {
            t.ci: t
            for t in db.query(Teacher).filter(Teacher.ci.in_(cis)).all()
        } if cis else {}

        from openpyxl import Workbook

        wb = Workbook()
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        month_name = MONTH_NAMES.get(payload.month, str(payload.month)).upper()
        ws = wb.create_sheet(title=f"{month_name} {payload.year}")

        company_name = app_settings_service.get_company_name(db)
        company_nit = app_settings_service.get_company_nit(db)

        salary_gen._apply_column_widths(ws)
        # Override title to indicate practice teachers
        salary_gen._write_title_block(ws, company_name, company_nit, month_name, payload.year)
        # Update the title cell to mention practice
        ws["A4"].value = (
            f"PLANILLA HONORARIOS DOCENTES ASISTENCIALES - MES DE {month_name} {payload.year}"
        )
        salary_gen._write_header_row(ws)
        last_data_row = salary_gen._write_data_rows(ws, rows, teachers)
        total_row = last_data_row + 1 if rows else 7
        salary_gen._write_totals_row(ws, total_row, last_data_row if rows else 6)
        salary_gen._apply_print_setup(ws, total_row)

        filename = f"planilla_salario_practicas_{payload.month:02d}_{payload.year}.xlsx"
        file_path = output_dir / filename
        wb.save(str(file_path))

        download_name = f"Planilla_Salario_Practicas_{MONTH_NAMES.get(payload.month, str(payload.month))}_{payload.year}.xlsx"

        log_activity(
            db,
            "generate_practice_salary_report",
            "practice-planilla",
            f"Planilla salarios prácticas generada: {MONTH_NAMES.get(payload.month)} {payload.year}",
            user=current_user,
            details={
                "month": payload.month,
                "year": payload.year,
                "discount_mode": effective_mode,
                "file": str(file_path),
            },
            request=request,
        )
        db.commit()

        return FileResponse(
            path=file_path,
            filename=download_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception("Practice salary report generation failed: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo generar la planilla de salarios de prácticas",
        ) from exc
