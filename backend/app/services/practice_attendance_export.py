from __future__ import annotations

import logging
from collections import defaultdict
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from sqlalchemy.orm import Session

from app.models.practice_attendance import PracticeAttendanceLog
from app.models.teacher import Teacher
from app.models.designation import Designation
from app.services import app_settings_service

logger = logging.getLogger(__name__)

MONTH_NAMES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# ── UPDS Brand Colors ────────────────────────────────────────────────────────
NAVY = colors.HexColor("#003366")
BLUE = colors.HexColor("#0066CC")
LIGHT_BLUE = colors.HexColor("#E8F4FD")
GREEN_BG = colors.HexColor("#F0FFF4")
YELLOW_BG = colors.HexColor("#FFFBEB")
RED_BG = colors.HexColor("#FEF2F2")
BLUE_BG = colors.HexColor("#EFF6FF")
LIGHT_GRAY = colors.HexColor("#F5F5F5")
DARK_GRAY = colors.HexColor("#666666")

OUTPUT_DIR = Path("data/output")
ASSETS_DIR = Path(__file__).resolve().parents[2] / "data" / "assets"
ISOLOGO_PATH = ASSETS_DIR / "isologo_upds.png"

STATUS_LABELS = {
    "attended": "ASISTIO",
    "absent": "AUSENTE",
    "late": "TARDANZA",
    "justified": "JUSTIFICADO",
}
STATUS_COLORS = {
    "attended": GREEN_BG,
    "absent": RED_BG,
    "late": YELLOW_BG,
    "justified": BLUE_BG,
}


def _query_logs(
    db: Session,
    month: int,
    year: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    teacher_ci: Optional[str] = None,
) -> list[PracticeAttendanceLog]:
    """Query attendance logs with shared filter logic."""
    query = db.query(PracticeAttendanceLog).join(
        Teacher, PracticeAttendanceLog.teacher_ci == Teacher.ci
    ).join(
        Designation, PracticeAttendanceLog.designation_id == Designation.id
    )

    if start_date and end_date:
        query = query.filter(
            PracticeAttendanceLog.date >= start_date,
            PracticeAttendanceLog.date <= end_date,
        )
    else:
        _, last_day = monthrange(year, month)
        query = query.filter(
            PracticeAttendanceLog.date >= date(year, month, 1),
            PracticeAttendanceLog.date <= date(year, month, last_day),
        )

    if teacher_ci:
        query = query.filter(PracticeAttendanceLog.teacher_ci == teacher_ci)

    return query.order_by(
        Teacher.full_name,
        PracticeAttendanceLog.date,
        PracticeAttendanceLog.scheduled_start,
    ).all()


def _date_range_str(
    month: int,
    year: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> str:
    if start_date and end_date:
        return f"{start_date.strftime('%d/%m/%Y')} — {end_date.strftime('%d/%m/%Y')}"
    return f"{MONTH_NAMES.get(month, '')} {year}"


# ── PDF ──────────────────────────────────────────────────────────────────────


def generate_practice_attendance_pdf(
    db: Session,
    month: int,
    year: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    teacher_ci: Optional[str] = None,
    generated_by: str = "Sistema",
    generated_by_ci: str = "",
    client_ip: str = "unknown",
) -> Path:
    """Generate a branded UPDS PDF of the practice attendance list.

    The PDF replicates the physical attendance list that practice teachers
    used to sign on paper. Grouped by teacher, then by date. Includes
    professional audit footer with user/IP/timestamp.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    period_name = app_settings_service.get_active_academic_period(db)

    logs = _query_logs(db, month, year, start_date, end_date, teacher_ci)

    # Group by teacher
    teacher_groups: dict[str, list] = defaultdict(list)
    teacher_names: dict[str, str] = {}
    for log in logs:
        teacher_groups[log.teacher_ci].append(log)
        if log.teacher_ci not in teacher_names:
            teacher = db.query(Teacher).filter(Teacher.ci == log.teacher_ci).first()
            teacher_names[log.teacher_ci] = teacher.full_name if teacher else log.teacher_ci

    date_range = _date_range_str(month, year, start_date, end_date)
    generation_ts = datetime.now()
    filename = f"asistencia_practicas_{month:02d}_{year}.pdf"
    filepath = OUTPUT_DIR / filename

    doc = SimpleDocTemplate(
        str(filepath),
        pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()

    # ── Branded UPDS Header ──────────────────────────────────────────────
    elements: list = []

    # Logo + institution name side by side
    if ISOLOGO_PATH.exists():
        logo = Image(str(ISOLOGO_PATH), width=2 * cm, height=2 * cm)
        logo.hAlign = "LEFT"
        elements.append(logo)
        elements.append(Spacer(1, 2 * mm))

    # Title bar (navy background)
    title_bar_style = ParagraphStyle(
        "TitleBar", parent=styles["Normal"],
        fontSize=13, textColor=colors.white,
        fontName="Helvetica-Bold", leading=17, alignment=TA_LEFT,
    )
    title_table = Table(
        [[Paragraph("Universidad Privada Domingo Savio — UPDS", title_bar_style)]],
        colWidths=["100%"],
    )
    title_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAVY),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 2 * mm))

    # Subtitle line
    sub_style = ParagraphStyle(
        "SubLine", parent=styles["Normal"],
        fontSize=9, textColor=DARK_GRAY, alignment=TA_LEFT,
    )
    elements.append(Paragraph(
        f"Facultad de Medicina · Periodo {period_name}",
        sub_style,
    ))
    elements.append(Spacer(1, 4 * mm))

    # Report title
    report_title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Normal"],
        fontSize=14, textColor=NAVY, fontName="Helvetica-Bold",
        alignment=TA_CENTER, spaceAfter=2 * mm,
    )
    elements.append(Paragraph(
        f"Control de Asistencia — Practicas Internas",
        report_title_style,
    ))

    report_period_style = ParagraphStyle(
        "ReportPeriod", parent=styles["Normal"],
        fontSize=11, textColor=BLUE, alignment=TA_CENTER, spaceAfter=4 * mm,
    )
    elements.append(Paragraph(f"Periodo: {date_range}", report_period_style))
    elements.append(Spacer(1, 3 * mm))

    # ── Teacher sections ─────────────────────────────────────────────────
    teacher_header_style = ParagraphStyle(
        "TeacherHeader", parent=styles["Heading2"],
        fontSize=12, textColor=NAVY, spaceBefore=6 * mm, spaceAfter=2 * mm,
    )

    for ci in sorted(teacher_groups.keys(), key=lambda c: teacher_names.get(c, "")):
        group_logs = teacher_groups[ci]
        teacher_name = teacher_names.get(ci, ci)

        elements.append(Paragraph(f"{teacher_name} (CI: {ci})", teacher_header_style))

        # Summary line
        total = len(group_logs)
        attended = sum(1 for l in group_logs if l.status in ("attended", "late", "justified"))
        absent = sum(1 for l in group_logs if l.status == "absent")
        rate = round(attended / total * 100, 1) if total > 0 else 0

        summary_style = ParagraphStyle("Summary", parent=styles["Normal"], fontSize=9, textColor=DARK_GRAY)
        elements.append(Paragraph(
            f"Clases programadas: {total} | Asistidas: {attended} | Ausentes: {absent} | Tasa: {rate}%",
            summary_style,
        ))
        elements.append(Spacer(1, 2 * mm))

        # Table
        headers = [
            "Fecha", "Materia", "Grupo", "Horario Program.",
            "Hora Llegada", "Hora Salida", "Hrs Acad.", "Estado", "Observacion",
        ]

        table_data = [headers]
        for log in group_logs:
            desig = db.query(Designation).filter(Designation.id == log.designation_id).first()
            subject = desig.subject if desig else "—"
            group_code = desig.group_code if desig else "—"

            table_data.append([
                log.date.strftime("%d/%m/%Y"),
                Paragraph(subject, ParagraphStyle("Cell", fontSize=7, leading=8)),
                group_code,
                f"{log.scheduled_start.strftime('%H:%M')} - {log.scheduled_end.strftime('%H:%M')}",
                log.actual_start.strftime("%H:%M") if log.actual_start else "—",
                log.actual_end.strftime("%H:%M") if log.actual_end else "—",
                str(log.academic_hours),
                STATUS_LABELS.get(log.status, log.status.upper()),
                Paragraph(log.observation or "", ParagraphStyle("CellObs", fontSize=7, leading=8)),
            ])

        col_widths = [55, 120, 40, 75, 50, 50, 35, 60, 120]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.gray),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ]

        for i, log in enumerate(group_logs, start=1):
            bg = STATUS_COLORS.get(log.status, colors.white)
            style_commands.append(("BACKGROUND", (7, i), (7, i), bg))

        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        elements.append(Spacer(1, 6 * mm))

    if not teacher_groups:
        elements.append(Paragraph(
            "No se encontraron registros de asistencia para el periodo seleccionado.",
            styles["Normal"],
        ))

    # ── Professional Audit Footer ────────────────────────────────────────
    elements.append(Spacer(1, 10 * mm))

    # Separator line
    separator = Table([[""]], colWidths=["100%"])
    separator.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), 1, NAVY),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(separator)

    # Audit info table
    footer_left_style = ParagraphStyle(
        "FooterLeft", parent=styles["Normal"],
        fontSize=7, textColor=DARK_GRAY, alignment=TA_LEFT,
    )
    footer_right_style = ParagraphStyle(
        "FooterRight", parent=styles["Normal"],
        fontSize=7, textColor=DARK_GRAY, alignment=TA_RIGHT,
    )

    footer_data = [
        [
            Paragraph(
                f"<b>Generado por:</b> {generated_by} (CI: {generated_by_ci})",
                footer_left_style,
            ),
            Paragraph(
                f"<b>Fecha/Hora:</b> {generation_ts.strftime('%d/%m/%Y %H:%M:%S')}",
                footer_right_style,
            ),
        ],
        [
            Paragraph(
                f"<b>Direccion IP:</b> {client_ip}",
                footer_left_style,
            ),
            Paragraph(
                f"<b>Sistema:</b> SIPAD v1.0 — Universidad Privada Domingo Savio",
                footer_right_style,
            ),
        ],
    ]

    footer_table = Table(footer_data, colWidths=["50%", "50%"])
    footer_table.setStyle(TableStyle([
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(footer_table)

    # Disclaimer
    disclaimer_style = ParagraphStyle(
        "Disclaimer", parent=styles["Normal"],
        fontSize=6, textColor=colors.HexColor("#999999"), alignment=TA_CENTER,
        spaceBefore=3 * mm,
    )
    elements.append(Paragraph(
        "Este documento fue generado automaticamente por SIPAD. "
        "Cualquier alteracion manual invalida su contenido. "
        "Para verificar su autenticidad, contacte a la administracion de la Facultad de Medicina.",
        disclaimer_style,
    ))

    doc.build(elements)
    logger.info("Generated practice attendance PDF: %s", filepath)
    return filepath


# ── Excel ────────────────────────────────────────────────────────────────────


def generate_practice_attendance_excel(
    db: Session,
    month: int,
    year: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    teacher_ci: Optional[str] = None,
) -> Path:
    """Generate an Excel spreadsheet of the practice attendance data."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    company_name = app_settings_service.get_company_name(db)

    logs = _query_logs(db, month, year, start_date, end_date, teacher_ci)

    wb = openpyxl.Workbook()
    ws = wb.active

    date_range = _date_range_str(month, year, start_date, end_date)
    ws.title = "Asistencia Practicas"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_font = Font(name="Calibri", size=14, bold=True, color="003366")
    subtitle_font = Font(name="Calibri", size=11, color="0066CC")

    ws.merge_cells("A1:L1")
    ws["A1"] = company_name
    ws["A1"].font = header_font

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Control de Asistencia — Practicas Internas — {date_range}"
    ws["A2"].font = subtitle_font

    # Headers row 4
    headers = [
        "Docente", "CI", "Fecha", "Materia", "Grupo", "Semestre",
        "Horario Programado", "Hora Llegada", "Hora Salida",
        "Hrs Academicas", "Estado", "Observacion",
    ]
    col_widths = [30, 12, 12, 35, 8, 12, 18, 12, 12, 10, 14, 30]

    header_fill = PatternFill("solid", fgColor="003366")
    header_font_white = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Status colors for Excel
    STATUS_FILLS = {
        "attended": PatternFill("solid", fgColor="C6EFCE"),
        "absent": PatternFill("solid", fgColor="FFC7CE"),
        "late": PatternFill("solid", fgColor="FFEB9C"),
        "justified": PatternFill("solid", fgColor="BDD7EE"),
    }

    data_font = Font(name="Calibri", size=9)

    row_num = 5
    for log in logs:
        teacher = db.query(Teacher).filter(Teacher.ci == log.teacher_ci).first()
        desig = db.query(Designation).filter(Designation.id == log.designation_id).first()

        values = [
            teacher.full_name if teacher else log.teacher_ci,
            log.teacher_ci,
            log.date.strftime("%d/%m/%Y"),
            desig.subject if desig else "",
            desig.group_code if desig else "",
            desig.semester if desig else "",
            f"{log.scheduled_start.strftime('%H:%M')} - {log.scheduled_end.strftime('%H:%M')}",
            log.actual_start.strftime("%H:%M") if log.actual_start else "",
            log.actual_end.strftime("%H:%M") if log.actual_end else "",
            log.academic_hours,
            STATUS_LABELS.get(log.status, log.status.upper()),
            log.observation or "",
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx not in (1, 4, 12) else left_align

        # Status cell coloring
        status_cell = ws.cell(row=row_num, column=11)
        if log.status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[log.status]

        # Alternate row background
        if row_num % 2 == 0:
            alt_fill = PatternFill("solid", fgColor="F9FAFB")
            for col_idx in range(1, len(values) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                if col_idx != 11:  # Don't override status color
                    cell.fill = alt_fill

        row_num += 1

    # Summary row
    row_num += 1
    ws.cell(row=row_num, column=1, value="RESUMEN").font = Font(
        name="Calibri", size=10, bold=True, color="003366",
    )
    row_num += 1
    total_logs = len(logs)
    total_attended = sum(1 for l in logs if l.status in ("attended", "late", "justified"))
    total_absent = sum(1 for l in logs if l.status == "absent")
    rate = round(total_attended / total_logs * 100, 1) if total_logs > 0 else 0

    ws.cell(row=row_num, column=1, value=f"Total registros: {total_logs}")
    ws.cell(row=row_num, column=3, value=f"Asistidas: {total_attended}")
    ws.cell(row=row_num, column=5, value=f"Ausentes: {total_absent}")
    ws.cell(row=row_num, column=7, value=f"Tasa: {rate}%")

    # Auto-filter
    last_data_row = row_num - 2 if total_logs > 0 else 4
    ws.auto_filter.ref = f"A4:L{last_data_row}"

    # Freeze header
    ws.freeze_panes = "A5"

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    filename = f"asistencia_practicas_{month:02d}_{year}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(str(filepath))
    logger.info("Generated practice attendance Excel: %s", filepath)
    return filepath
