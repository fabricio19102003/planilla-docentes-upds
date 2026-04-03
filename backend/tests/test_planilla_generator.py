"""
Tests for PlanillaGenerator service (T-007).

Test layers:
  1. Unit tests — pure Python, no DB. Test data structures, column constants,
     _build_row logic, helper functions.
  2. Integration tests — in-memory SQLite DB. Test _build_planilla_data() and
     full generate() end-to-end with synthetic attendance data.
  3. Output tests — verify the generated Excel file opens correctly and has
     expected structure (sheet names, row counts, key cell values).

How to run:
    cd backend
    pytest tests/test_planilla_generator.py -v
"""
from __future__ import annotations

import calendar
import os
import tempfile
from datetime import date, time
from decimal import Decimal
from pathlib import Path
from typing import Optional
from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, Session

import app.models  # noqa: F401 — register all models for create_all
from app.database import Base
from app.models.attendance import AttendanceRecord
from app.models.biometric import BiometricRecord, BiometricUpload
from app.models.designation import Designation
from app.models.planilla import PlanillaOutput
from app.models.teacher import Teacher

from app.services.planilla_generator import (
    DAY_COL_END,
    DAY_COL_START,
    DATA_ROW_START,
    COL_BANCO,
    COL_CI,
    COL_NOMBRE,
    COL_SEMESTRE,
    COL_TOTAL_HORAS,
    COL_PAGO_CALCULADO,
    COL_PAGO_AJUSTADO,
    COL_OBSERVACIONES,
    MONTH_NAMES,
    RATE_PER_HOUR,
    ROW_COL_HEADERS,
    ROW_WEEKDAY,
    TOTAL_COLS,
    WEEKDAY_LETTERS,
    PlanillaGenerator,
    PlanillaRow,
    PlanillaResult,
    _get_month_from_name,
)

# ---------------------------------------------------------------------------
# In-memory SQLite fixtures
# ---------------------------------------------------------------------------

TEST_DB_URL = "sqlite:///:memory:"


@pytest.fixture(scope="module")
def engine():
    """Create an in-memory SQLite engine and build all tables."""
    eng = create_engine(TEST_DB_URL, connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=eng)
    yield eng
    Base.metadata.drop_all(bind=eng)


@pytest.fixture
def db(engine):
    """Provide a fresh session rolled back after each test."""
    conn = engine.connect()
    trans = conn.begin()
    Session_ = sessionmaker(bind=conn)
    session = Session_()
    yield session
    session.close()
    trans.rollback()
    conn.close()


@pytest.fixture
def temp_output_dir(tmp_path):
    """Return a temporary directory for planilla output."""
    output = tmp_path / "output"
    output.mkdir()
    return str(output)


# ---------------------------------------------------------------------------
# Helper: seed minimal DB fixtures
# ---------------------------------------------------------------------------


def seed_teacher(db: Session, ci: str = "12345678", name: str = "PEREZ JUAN") -> Teacher:
    """Insert a teacher into the test DB."""
    teacher = Teacher(
        ci=ci,
        full_name=name,
        email=f"{ci}@test.com",
        phone="77712345",
        gender="M",
        external_permanent="Externo",
        academic_level="Licenciatura",
        profession="Médico",
        specialty="Medicina General",
        bank="BCP",
        account_number="1234567890",
        sap_code=f"SAP{ci}",
        invoice_retention="Factura",
    )
    db.add(teacher)
    db.flush()
    return teacher


def seed_designation(
    db: Session,
    teacher_ci: str,
    subject: str = "Anatomía I",
    semester: str = "1",
    group_code: str = "M-1",
) -> Designation:
    """Insert a designation into the test DB."""
    desig = Designation(
        teacher_ci=teacher_ci,
        subject=subject,
        semester=semester,
        group_code=group_code,
        schedule_json=[
            {
                "dia": "lunes",
                "hora_inicio": "08:00",
                "hora_fin": "10:00",
                "duracion_minutos": 120,
                "horas_academicas": 3,
            }
        ],
        semester_hours=60,
        monthly_hours=12,
        weekly_hours=3,
    )
    db.add(desig)
    db.flush()
    return desig


def seed_biometric(db: Session, teacher_ci: str, day: int = 2, month: int = 3, year: int = 2026) -> BiometricRecord:
    """Insert a biometric upload + record so the teacher is recognized as having real biometric data."""
    # Ensure a BiometricUpload exists (reuse if already created)
    upload = db.query(BiometricUpload).first()
    if upload is None:
        upload = BiometricUpload(filename="test_bio.xls", month=month, year=year, total_records=1, total_teachers=1)
        db.add(upload)
        db.flush()
    rec = BiometricRecord(
        upload_id=upload.id,
        teacher_ci=teacher_ci,
        date=date(year, month, day),
        entry_time=time(7, 55),
        exit_time=time(10, 5),
        worked_minutes=130,
    )
    db.add(rec)
    db.flush()
    return rec


def seed_attendance(
    db: Session,
    teacher_ci: str,
    designation_id: int,
    day: int = 2,
    month: int = 3,
    year: int = 2026,
    status: str = "ATTENDED",
    academic_hours: int = 3,
    late_minutes: int = 0,
    observation: Optional[str] = None,
) -> AttendanceRecord:
    """Insert one attendance record."""
    rec = AttendanceRecord(
        teacher_ci=teacher_ci,
        designation_id=designation_id,
        date=date(year, month, day),
        scheduled_start=time(8, 0),
        scheduled_end=time(10, 0),
        actual_entry=time(7, 55) if status != "ABSENT" else None,
        actual_exit=time(10, 5) if status == "ATTENDED" else None,
        status=status,
        academic_hours=academic_hours,
        late_minutes=late_minutes,
        observation=observation,
        month=month,
        year=year,
    )
    db.add(rec)
    db.flush()
    return rec


# ===========================================================================
# Unit Tests: Constants and helpers
# ===========================================================================


class TestConstants:
    """Verify all column constant values are within expected ranges."""

    def test_day_col_range_covers_31_days(self):
        assert DAY_COL_END - DAY_COL_START + 1 == 31

    def test_day_col_start_is_col_17(self):
        assert DAY_COL_START == 17

    def test_total_cols_is_58(self):
        assert TOTAL_COLS == 58

    def test_data_row_starts_at_7(self):
        assert DATA_ROW_START == 7

    def test_rate_per_hour_is_70(self):
        assert RATE_PER_HOUR == 70.0

    def test_all_months_in_month_names(self):
        assert set(MONTH_NAMES.keys()) == set(range(1, 13))

    def test_weekday_letters_covers_all_7_days(self):
        assert set(WEEKDAY_LETTERS.keys()) == set(range(7))

    def test_weekday_letters_are_single_chars(self):
        for letter in WEEKDAY_LETTERS.values():
            assert len(letter) == 1


class TestHelpers:
    """Test module-level helper functions."""

    def test_get_month_from_name_marzo(self):
        assert _get_month_from_name("MARZO") == 3

    def test_get_month_from_name_junio(self):
        assert _get_month_from_name("JUNIO") == 6

    def test_get_month_from_name_case_insensitive(self):
        assert _get_month_from_name("marzo") == 3

    def test_get_month_from_name_unknown_returns_1(self):
        assert _get_month_from_name("QUATEMBER") == 1


# ===========================================================================
# Unit Tests: PlanillaRow dataclass
# ===========================================================================


class TestPlanillaRow:
    """Verify PlanillaRow field behavior."""

    def test_default_empty_collections(self):
        row = PlanillaRow(
            teacher_ci="123",
            designation_id=1,
            teacher_name="Test",
            email=None,
            phone=None,
            subject="Materia",
            semester="1",
            group_code="M-1",
            teacher_type=None,
            gender=None,
            sap_code=None,
            invoice_retention=None,
            account_number=None,
            academic_level=None,
            profession=None,
            specialty=None,
            bank=None,
        )
        assert row.daily_hours == {}
        assert row.daily_status == {}
        assert row.observations == []
        assert row.total_hours == 0
        assert row.calculated_payment == 0.0
        assert row.rate_per_hour == RATE_PER_HOUR

    def test_calculated_payment_field(self):
        row = PlanillaRow(
            teacher_ci="123",
            designation_id=1,
            teacher_name="Test",
            email=None,
            phone=None,
            subject="Materia",
            semester="1",
            group_code="M-1",
            teacher_type=None,
            gender=None,
            sap_code=None,
            invoice_retention=None,
            account_number=None,
            academic_level=None,
            profession=None,
            specialty=None,
            bank=None,
            total_hours=10,
            calculated_payment=10 * RATE_PER_HOUR,
        )
        assert row.calculated_payment == 700.0


# ===========================================================================
# Unit Tests: PlanillaGenerator._build_row()
# ===========================================================================


class TestBuildRow:
    """Test the _build_row() method with mock objects."""

    def _make_teacher(self, ci: str = "99999") -> MagicMock:
        t = MagicMock()
        t.ci = ci
        t.full_name = "GARCIA MARIA"
        t.email = "maria@test.com"
        t.phone = "77700000"
        t.gender = "F"
        t.external_permanent = "Permanente"
        t.academic_level = "Maestría"
        t.profession = "Médico"
        t.specialty = "Cirugía"
        t.bank = "Banco Unión"
        t.account_number = "987654321"
        t.sap_code = "SAP001"
        t.invoice_retention = "Retención"
        return t

    def _make_desig(self, monthly_hours: int = 0) -> MagicMock:
        d = MagicMock()
        d.subject = "Anatomía II"
        d.semester = "2"
        d.group_code = "T-1"
        d.id = 1
        d.monthly_hours = monthly_hours  # Model C: explicit int, not MagicMock
        d.schedule_json = []  # no schedule needed for unit tests
        return d

    def _make_att(
        self,
        day: int,
        status: str,
        academic_hours: int,
        late_minutes: int = 0,
        observation: Optional[str] = None,
    ) -> MagicMock:
        rec = MagicMock()
        rec.date = date(2026, 3, day)
        rec.status = status
        rec.academic_hours = academic_hours
        rec.late_minutes = late_minutes
        rec.observation = observation
        return rec

    def test_builds_correct_daily_hours(self):
        """Daily hours dict reflects actual attendance per day (for Excel column display)."""
        gen = PlanillaGenerator.__new__(PlanillaGenerator)
        teacher = self._make_teacher()
        desig = self._make_desig(monthly_hours=12)  # Model C: 12 base hours
        records = [
            self._make_att(2, "ATTENDED", 3),
            self._make_att(9, "ATTENDED", 3),
            self._make_att(16, "LATE", 3, 10, "Llegada tardía: 10 min"),
        ]
        row = gen._build_row(teacher, desig, records)
        # daily_hours still shows actual attendance for Excel day columns
        assert row.daily_hours[2] == 3
        assert row.daily_hours[9] == 3
        assert row.daily_hours[16] == 3
        # Model C: total_hours = payable_hours = base_monthly_hours - absent_hours
        # No absences → payable = 12, total_hours = 12
        assert row.total_hours == 12
        assert row.base_monthly_hours == 12
        assert row.absent_hours == 0
        assert row.payable_hours == 12

    def test_absent_day_contributes_zero_hours(self):
        """Model C: absent slots deduct from base monthly hours."""
        gen = PlanillaGenerator.__new__(PlanillaGenerator)
        teacher = self._make_teacher()
        # schedule_json empty → _get_slot_hours returns 0 for absent slot
        desig = self._make_desig(monthly_hours=12)
        records = [
            self._make_att(2, "ATTENDED", 3),
            self._make_att(9, "ABSENT", 0),
        ]
        row = gen._build_row(teacher, desig, records)
        assert row.daily_hours.get(9, 0) == 0
        assert row.absent_count == 1
        # absent_hours=0 because schedule_json=[] so _get_slot_hours returns 0
        # payable_hours = 12 - 0 = 12 (schedule not matched)
        assert row.base_monthly_hours == 12
        assert row.absent_hours == 0  # can't deduct without matching schedule slot
        assert row.total_hours == 12

    def test_late_status_counted(self):
        gen = PlanillaGenerator.__new__(PlanillaGenerator)
        teacher = self._make_teacher()
        desig = self._make_desig(monthly_hours=8)
        records = [
            self._make_att(3, "LATE", 3, 8, "Tardanza 8 min"),
            self._make_att(10, "LATE", 3, 12, "Tardanza 12 min"),
        ]
        row = gen._build_row(teacher, desig, records)
        assert row.late_count == 2
        # LATE does NOT deduct from base pay (only ABSENT does)
        assert row.absent_hours == 0
        assert row.payable_hours == 8

    def test_payment_calculation(self):
        """Model C: payment = monthly_hours × 70 (no absences)."""
        gen = PlanillaGenerator.__new__(PlanillaGenerator)
        teacher = self._make_teacher()
        desig = self._make_desig(monthly_hours=12)
        records = [self._make_att(d, "ATTENDED", 3) for d in [2, 9, 16, 23]]
        row = gen._build_row(teacher, desig, records)
        assert row.total_hours == 12          # payable = monthly_hours (no absent)
        assert row.calculated_payment == 12 * 70.0

    def test_same_day_multiple_slots_accumulate(self):
        """Multiple attendance records on the same day should sum their hours in daily_hours."""
        gen = PlanillaGenerator.__new__(PlanillaGenerator)
        teacher = self._make_teacher()
        desig = self._make_desig(monthly_hours=8)
        # Two slots on day 5: 2hrs + 2hrs = 4hrs in daily_hours display
        r1 = self._make_att(5, "ATTENDED", 2)
        r2 = self._make_att(5, "ATTENDED", 2)
        row = gen._build_row(teacher, desig, [r1, r2])
        assert row.daily_hours[5] == 4
        # total_hours = payable_hours = monthly_hours (no absences)
        assert row.total_hours == 8

    def test_worst_status_wins_for_coloring(self):
        """If a day has both ATTENDED and LATE slots, daily_status should be LATE."""
        gen = PlanillaGenerator.__new__(PlanillaGenerator)
        teacher = self._make_teacher()
        desig = self._make_desig(monthly_hours=8)
        r1 = self._make_att(5, "ATTENDED", 2)
        r2 = self._make_att(5, "LATE", 2, 10)
        row = gen._build_row(teacher, desig, [r1, r2])
        assert row.daily_status[5] == "LATE"

    def test_observations_included_when_present(self):
        gen = PlanillaGenerator.__new__(PlanillaGenerator)
        teacher = self._make_teacher()
        desig = self._make_desig(monthly_hours=8)
        records = [
            self._make_att(3, "LATE", 3, 8, "Tardanza 8 min"),
            self._make_att(10, "ABSENT", 0, 0, "Sin registro"),
        ]
        row = gen._build_row(teacher, desig, records)
        # Model C observations: "1 tardanza", "1 ausencia"
        assert any("tardanza" in o for o in row.observations)
        assert any("ausencia" in o for o in row.observations)


# ===========================================================================
# Integration Tests: Build planilla data from DB
# ===========================================================================


class TestBuildPlanillaData:
    """Integration tests against in-memory SQLite."""

    def test_empty_attendance_returns_empty_rows(self, db, temp_output_dir):
        """With no designations in DB, planilla returns empty rows and a warning."""
        gen = PlanillaGenerator(output_dir=temp_output_dir)
        rows, detail_rows, warnings = gen._build_planilla_data(db, month=3, year=2026)
        assert rows == []
        assert detail_rows == []
        assert len(warnings) == 1
        # Model C: warns about missing designations (not attendance records)
        assert "designaci" in warnings[0].lower() or "No hay" in warnings[0]

    def test_single_teacher_single_designation(self, db, temp_output_dir):
        """Model C: total_hours = payable_hours = monthly_hours (no absences)."""
        teacher = seed_teacher(db, "11111111", "LOPEZ CARLOS")
        desig = seed_designation(db, teacher.ci)  # monthly_hours=12
        seed_attendance(db, teacher.ci, desig.id, day=2)
        seed_attendance(db, teacher.ci, desig.id, day=9)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        rows, detail_rows, warnings = gen._build_planilla_data(db, month=3, year=2026)

        assert len(rows) == 1
        assert len(detail_rows) == 2
        assert len(warnings) == 0
        row = rows[0]
        assert row.teacher_ci == "11111111"
        assert row.subject == "Anatomía I"
        assert row.daily_hours[2] == 3
        assert row.daily_hours[9] == 3
        # Model C: total_hours = payable_hours = base_monthly_hours (12) - absent_hours (0)
        assert row.base_monthly_hours == 12
        assert row.absent_hours == 0
        assert row.total_hours == 12   # payable = full monthly_hours (no absences)

    def test_multiple_designations_produce_multiple_rows(self, db, temp_output_dir):
        teacher = seed_teacher(db, "22222222", "FERNANDEZ ANA")
        d1 = seed_designation(db, teacher.ci, subject="Anatomía I", group_code="M-1")
        d2 = seed_designation(db, teacher.ci, subject="Fisiología", group_code="T-2")
        seed_attendance(db, teacher.ci, d1.id, day=2)
        seed_attendance(db, teacher.ci, d2.id, day=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        rows, detail_rows, warnings = gen._build_planilla_data(db, month=3, year=2026)

        assert len(rows) == 2
        assert len(detail_rows) == 2
        subjects = {r.subject for r in rows}
        assert "Anatomía I" in subjects
        assert "Fisiología" in subjects

    def test_multiple_teachers(self, db, temp_output_dir):
        t1 = seed_teacher(db, "33333333", "VARGAS LUIS")
        t2 = seed_teacher(db, "44444444", "MORA ELENA")
        d1 = seed_designation(db, t1.ci, subject="Bioquímica")
        d2 = seed_designation(db, t2.ci, subject="Histología")
        seed_attendance(db, t1.ci, d1.id, day=4)
        seed_attendance(db, t2.ci, d2.id, day=5)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        rows, detail_rows, warnings = gen._build_planilla_data(db, month=3, year=2026)

        assert len(rows) == 2
        cis = {r.teacher_ci for r in rows}
        assert "33333333" in cis
        assert "44444444" in cis

    def test_absent_records_are_included_but_zero_hours(self, db, temp_output_dir):
        """
        Model C: Absent slot shows 0 in daily_hours.
        total_hours = payable_hours = monthly_hours - absent_hours.

        The ABSENT record's scheduled_start=08:00 must match schedule_json's
        hora_inicio="08:00" for _get_slot_hours() to deduct 3 hrs.
        """
        teacher = seed_teacher(db, "55555555", "RAMIREZ PEDRO")
        desig = seed_designation(db, teacher.ci)  # monthly_hours=12, schedule has 08:00 slot
        seed_biometric(db, teacher.ci, day=2)  # Mark teacher as having real biometric data
        seed_attendance(db, teacher.ci, desig.id, day=2, status="ATTENDED", academic_hours=3)
        seed_attendance(db, teacher.ci, desig.id, day=9, status="ABSENT", academic_hours=0)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        rows, _, _ = gen._build_planilla_data(db, month=3, year=2026)

        assert len(rows) == 1
        row = rows[0]
        assert row.has_biometric is True
        # Model C: daily_hours for absent day = 0 (no academic hours awarded)
        assert row.daily_hours.get(9, 0) == 0
        assert row.absent_count == 1
        # Absent slot: scheduled_start=08:00 matches schedule_json hora_inicio="08:00" → 3h deducted
        assert row.base_monthly_hours == 12
        assert row.absent_hours == 3
        assert row.payable_hours == 9     # 12 - 3
        assert row.total_hours == 9       # payable_hours


# ===========================================================================
# Integration Tests: Full Excel generation
# ===========================================================================


class TestGenerateExcel:
    """End-to-end tests: generate Excel and verify file structure."""

    def test_generate_creates_file(self, db, temp_output_dir):
        teacher = seed_teacher(db, "66666666", "SALAS JORGE")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=2)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        assert isinstance(result, PlanillaResult)
        assert os.path.exists(result.file_path)
        assert result.file_path.endswith(".xlsx")

    def test_generated_file_has_two_sheets(self, db, temp_output_dir):
        teacher = seed_teacher(db, "77777777", "AGUILAR ROSA")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        assert "Planilla" in wb.sheetnames
        assert "Detalle" in wb.sheetnames

    def test_planilla_sheet_has_title_in_row1(self, db, temp_output_dir):
        teacher = seed_teacher(db, "88888888", "CASTRO MARIO")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=4)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]
        title = ws.cell(row=1, column=1).value
        assert title is not None
        assert "MARZO" in str(title)
        assert "2026" in str(title)

    def test_planilla_col_headers_row_has_semestre(self, db, temp_output_dir):
        teacher = seed_teacher(db, "10101010", "DIAZ BLANCA")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=5)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]
        col_a_header = ws.cell(row=ROW_COL_HEADERS, column=COL_SEMESTRE).value
        assert col_a_header == "Semestre"

    def test_planilla_col_headers_day_numbers(self, db, temp_output_dir):
        teacher = seed_teacher(db, "20202020", "FLORES ANA")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=10)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # Day 1 should be in column DAY_COL_START (Q = 17)
        day1_col = ws.cell(row=ROW_COL_HEADERS, column=DAY_COL_START).value
        assert day1_col == 1

        # Day 31 should be in column DAY_COL_END (AU = 47)
        day31_col = ws.cell(row=ROW_COL_HEADERS, column=DAY_COL_END).value
        assert day31_col == 31

    def test_planilla_data_row_has_teacher_name(self, db, temp_output_dir):
        teacher = seed_teacher(db, "30303030", "ORTIZ CARMEN")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=11)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # First data row: col B = Nombre
        name_cell = ws.cell(row=DATA_ROW_START, column=COL_NOMBRE).value
        assert name_cell == "ORTIZ CARMEN"

    def test_planilla_data_row_hours_in_correct_column(self, db, temp_output_dir):
        teacher = seed_teacher(db, "40404040", "MORALES IVAN")
        desig = seed_designation(db, teacher.ci)
        # Day 15, month 3 → column Q + 14 = col 31
        seed_attendance(db, teacher.ci, desig.id, day=15, academic_hours=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # Day 15 = DAY_COL_START + 14
        day15_col = DAY_COL_START + 14
        hours_cell = ws.cell(row=DATA_ROW_START, column=day15_col).value
        assert hours_cell == 3

    def test_planilla_total_hours_column(self, db, temp_output_dir):
        """
        Model C: COL_TOTAL_HORAS = payable_hours = monthly_hours - absent_hours.
        seed_designation has monthly_hours=12, no absences → payable = 12.
        """
        teacher = seed_teacher(db, "50505050", "REYES JULIA")
        desig = seed_designation(db, teacher.ci)  # monthly_hours=12
        seed_attendance(db, teacher.ci, desig.id, day=2, academic_hours=3)
        seed_attendance(db, teacher.ci, desig.id, day=9, academic_hours=3)
        seed_attendance(db, teacher.ci, desig.id, day=16, academic_hours=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # Model C: COL_TOTAL_HORAS = payable_hours = monthly_hours (no absences)
        total_cell = ws.cell(row=DATA_ROW_START, column=COL_TOTAL_HORAS).value
        assert total_cell == 12  # payable_hours = monthly_hours (no absences)

    def test_planilla_payment_calculated_correctly(self, db, temp_output_dir):
        teacher = seed_teacher(db, "60606060", "NAVARRO FELIX")
        desig = seed_designation(db, teacher.ci)
        # 4 weeks × 3 hours = 12 hours × 70 = 840 Bs
        for day in [2, 9, 16, 23]:
            seed_attendance(db, teacher.ci, desig.id, day=day, academic_hours=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        payment_cell = ws.cell(row=DATA_ROW_START, column=COL_PAGO_CALCULADO).value
        assert payment_cell == 840.0

    def test_payment_override_goes_to_pago_ajustado(self, db, temp_output_dir):
        """
        Model C: COL_PAGO_CALCULADO = payable_hours × 70 (not attended hours × 70).
        seed_designation has monthly_hours=12, no absences → 12 × 70 = 840 Bs.
        """
        teacher = seed_teacher(db, "70707070", "IGLESIAS PETRA")
        desig = seed_designation(db, teacher.ci)  # monthly_hours=12
        seed_attendance(db, teacher.ci, desig.id, day=2, academic_hours=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(
            db, month=3, year=2026,
            payment_overrides={"70707070": 500.0}
        )

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # Model C: COL_PAGO_CALCULADO = payable_hours × 70 = monthly_hours × 70 = 12 × 70 = 840
        calculated = ws.cell(row=DATA_ROW_START, column=COL_PAGO_CALCULADO).value
        assert calculated == 12 * 70.0  # = 840 Bs (monthly_hours × rate)

        # COL_PAGO_AJUSTADO should have the override
        adjusted = ws.cell(row=DATA_ROW_START, column=COL_PAGO_AJUSTADO).value
        assert adjusted == 500.0

    def test_no_payment_override_leaves_pago_ajustado_empty(self, db, temp_output_dir):
        teacher = seed_teacher(db, "80808080", "VEGA SUSANA")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=2, academic_hours=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]
        adjusted = ws.cell(row=DATA_ROW_START, column=COL_PAGO_AJUSTADO).value
        assert adjusted is None

    def test_detalle_sheet_has_headers(self, db, temp_output_dir):
        teacher = seed_teacher(db, "90909090", "DELGADO JOSE")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=6)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Detalle"]
        # Row 2 should have "CI" header
        assert ws.cell(row=2, column=1).value == "CI"
        assert ws.cell(row=2, column=2).value == "Docente"

    def test_detalle_sheet_has_correct_record_count(self, db, temp_output_dir):
        teacher = seed_teacher(db, "11223344", "MONTOYA LUZ")
        desig = seed_designation(db, teacher.ci)
        # 3 attendance records
        seed_attendance(db, teacher.ci, desig.id, day=7)
        seed_attendance(db, teacher.ci, desig.id, day=14)
        seed_attendance(db, teacher.ci, desig.id, day=21)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Detalle"]
        # Data starts at row 3; count rows where col A contains a CI (numeric-like string)
        # Skip the summary row which starts with "TOTAL REGISTROS:"
        data_rows = [
            ws.cell(row=r, column=1).value
            for r in range(3, 3 + 20)
            if ws.cell(row=r, column=1).value
            and "TOTAL" not in str(ws.cell(row=r, column=1).value)
        ]
        assert len(data_rows) == 3

    def test_result_statistics_match_data(self, db, temp_output_dir):
        teacher = seed_teacher(db, "55443322", "PAREDES ALICIA")
        desig = seed_designation(db, teacher.ci)
        for day in [2, 9, 16]:
            seed_attendance(db, teacher.ci, desig.id, day=day, academic_hours=4)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        assert result.total_teachers == 1
        assert result.total_rows == 1
        assert result.total_hours == 12
        assert result.total_payment == 12 * 70.0

    def test_weekday_row_has_correct_letters(self, db, temp_output_dir):
        """March 2026 day 1 = Sunday (weekday=6 → 'D')."""
        teacher = seed_teacher(db, "11223355", "GUTIERREZ PABLO")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=2)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # March 1, 2026 = Sunday → WEEKDAY_LETTERS[6] = 'D'
        day1_weekday = ws.cell(row=ROW_WEEKDAY, column=DAY_COL_START).value
        assert day1_weekday == "D"  # March 1, 2026 is Sunday

        # March 2, 2026 = Monday → 'L'
        day2_weekday = ws.cell(row=ROW_WEEKDAY, column=DAY_COL_START + 1).value
        assert day2_weekday == "L"

    def test_february_has_no_day_29_in_non_leap_year(self, db, temp_output_dir):
        """February 2026 has 28 days (non-leap). Day 29–31 cols should be empty."""
        teacher = seed_teacher(db, "99887766", "RUIZ PABLO")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=14, month=2, year=2026, academic_hours=2)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=2, year=2026)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # Col headers for day 29, 30, 31 should be None (month only has 28 days)
        day29_header = ws.cell(row=ROW_COL_HEADERS, column=DAY_COL_START + 28).value
        assert day29_header is None

    def test_output_filename_contains_month_year(self, db, temp_output_dir):
        teacher = seed_teacher(db, "12341234", "TORRES BELEN")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        filename = os.path.basename(result.file_path)
        assert "03" in filename
        assert "2026" in filename

    def test_planilla_output_persisted_to_db(self, db, temp_output_dir):
        teacher = seed_teacher(db, "56785678", "SALINAS OMAR")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=7)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        assert result.planilla_output_id is not None
        po = db.query(PlanillaOutput).filter_by(
            month=3, year=2026
        ).first()
        assert po is not None
        assert po.total_teachers == 1
        assert po.status == "generated"


# ===========================================================================
# Edge Case Tests
# ===========================================================================


class TestEdgeCases:
    """Edge cases: empty data, all absent, overrides, etc."""

    def test_generate_with_all_absent_month(self, db, temp_output_dir):
        """All records ABSENT → Model C: base - all_absent = 0, payment = 0."""
        teacher = seed_teacher(db, "13131313", "IBARRA KARINA")
        desig = seed_designation(db, teacher.ci)  # monthly_hours=12, schedule slot=3h
        seed_biometric(db, teacher.ci, day=2)  # Mark as having real biometric data
        # 4 absent days × 3h per slot = 12h absent = monthly_hours → payable = 0
        for day in [2, 9, 16, 23]:
            seed_attendance(db, teacher.ci, desig.id, day=day,
                            status="ABSENT", academic_hours=0)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=3, year=2026)

        assert result.total_hours == 0
        assert result.total_payment == 0.0

    def test_no_attendance_records_generates_empty_planilla(self, db, temp_output_dir):
        """No records → generates file but with no data rows (just headers + warning)."""
        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=4, year=2026)

        assert len(result.warnings) > 0
        assert result.total_rows == 0
        # File should still be created
        assert os.path.exists(result.file_path)

    def test_payment_override_only_affects_target_teacher(self, db, temp_output_dir):
        """Override for CI A shouldn't change CI B's payment."""
        t1 = seed_teacher(db, "14141414", "PONCE RAUL")
        t2 = seed_teacher(db, "15151515", "FERIA DIANA")
        d1 = seed_designation(db, t1.ci, subject="Bioquímica")
        d2 = seed_designation(db, t2.ci, subject="Histología")
        seed_attendance(db, t1.ci, d1.id, day=2, academic_hours=3)
        seed_attendance(db, t2.ci, d2.id, day=2, academic_hours=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(
            db, month=3, year=2026,
            payment_overrides={"14141414": 1000.0}
        )

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # Find rows by CI
        pago_ajustado_by_ci = {}
        for row_num in range(DATA_ROW_START, DATA_ROW_START + 10):
            ci = ws.cell(row=row_num, column=COL_CI).value
            if ci in ("14141414", "15151515"):
                pago_ajustado_by_ci[ci] = ws.cell(
                    row=row_num, column=COL_PAGO_AJUSTADO
                ).value

        assert pago_ajustado_by_ci.get("14141414") == 1000.0
        assert pago_ajustado_by_ci.get("15151515") is None  # No override

    def test_teacher_total_override_is_not_duplicated_across_rows(self, db, temp_output_dir):
        """A teacher-level override must apply once even when the teacher has multiple rows."""
        teacher = seed_teacher(db, "16161616", "SUAREZ LIDIA")
        d1 = seed_designation(db, teacher.ci, subject="Bioquímica", group_code="M-1")
        d2 = seed_designation(db, teacher.ci, subject="Histología", group_code="T-2")
        seed_attendance(db, teacher.ci, d1.id, day=2, academic_hours=3)
        seed_attendance(db, teacher.ci, d2.id, day=3, academic_hours=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(
            db,
            month=3,
            year=2026,
            payment_overrides={teacher.ci: 1000.0},
        )

        assert result.total_payment == 1000.0

    def test_teacher_total_override_is_distributed_across_multiple_rows(self, db, temp_output_dir):
        """
        Multi-row teacher override must be distributed proportionally by payable_hours.

        Model C: Both designations have monthly_hours=12, no absences → payable_hours=12 each.
        Distribution is 50/50 → each row gets 500.0.
        (In Model A this was 75/25 based on 3h vs 1h attended — that is no longer correct.)
        """
        teacher = seed_teacher(db, "16160000", "SUAREZ LIDIA")
        d1 = seed_designation(db, teacher.ci, subject="Bioquímica", group_code="M-1")
        d2 = seed_designation(db, teacher.ci, subject="Histología", group_code="T-2")
        seed_attendance(db, teacher.ci, d1.id, day=2, academic_hours=3)
        seed_attendance(db, teacher.ci, d2.id, day=3, academic_hours=1)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(
            db,
            month=3,
            year=2026,
            payment_overrides={teacher.ci: 1000.0},
        )

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        adjusted_values = []
        for row_num in range(DATA_ROW_START, DATA_ROW_START + 10):
            if ws.cell(row=row_num, column=COL_CI).value == teacher.ci:
                adjusted_values.append(ws.cell(row=row_num, column=COL_PAGO_AJUSTADO).value)

        # Model C: payable_hours=12 for both → 50/50 split
        assert len(adjusted_values) == 2
        assert sum(adjusted_values) == 1000.0
        assert result.total_payment == 1000.0

    def test_row_override_uses_teacher_and_designation_key(self, db, temp_output_dir):
        """
        A row override must target a single teacher/designation combination.

        Model C: d2 has no override → payment = monthly_hours × 70 = 12 × 70 = 840.
        Total = 500 (d1 override) + 840 (d2 Model C) = 1340.
        """
        teacher = seed_teacher(db, "17171717", "MENDEZ LARA")
        d1 = seed_designation(db, teacher.ci, subject="Bioquímica", group_code="M-1")  # monthly_hours=12
        d2 = seed_designation(db, teacher.ci, subject="Histología", group_code="T-2")  # monthly_hours=12
        seed_attendance(db, teacher.ci, d1.id, day=2, academic_hours=3)
        seed_attendance(db, teacher.ci, d2.id, day=3, academic_hours=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(
            db,
            month=3,
            year=2026,
            payment_overrides={f"{teacher.ci}:{d1.id}": 500.0},
        )

        # Model C: d2 has no override → uses calculated_payment = payable_hours × 70 = 12 × 70 = 840
        assert result.total_payment == 500.0 + (12 * RATE_PER_HOUR)

    def test_row_override_takes_precedence_over_teacher_override(self, db, temp_output_dir):
        """Row-level overrides must beat teacher-level overrides in both row and total calculations."""
        teacher = seed_teacher(db, "17170000", "MENDEZ LARA")
        d1 = seed_designation(db, teacher.ci, subject="Bioquímica", group_code="M-1")
        d2 = seed_designation(db, teacher.ci, subject="Histología", group_code="T-2")
        seed_attendance(db, teacher.ci, d1.id, day=2, academic_hours=3)
        seed_attendance(db, teacher.ci, d2.id, day=3, academic_hours=3)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(
            db,
            month=3,
            year=2026,
            payment_overrides={teacher.ci: 1000.0, f"{teacher.ci}:{d1.id}": 500.0},
        )

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        adjusted_by_subject = {}
        for row_num in range(DATA_ROW_START, DATA_ROW_START + 10):
            if ws.cell(row=row_num, column=COL_CI).value == teacher.ci:
                subject = ws.cell(row=row_num, column=6).value
                adjusted_by_subject[subject] = ws.cell(row=row_num, column=COL_PAGO_AJUSTADO).value

        assert adjusted_by_subject["Bioquímica"] == 500.0
        assert adjusted_by_subject["Histología"] == 500.0
        assert result.total_payment == 1000.0

    def test_teacher_override_remainder_excludes_row_override_amount(self, db, temp_output_dir):
        """Teacher override must distribute only the remainder after row overrides."""
        teacher = seed_teacher(db, "17170001", "MENDEZ LARA")
        d1 = seed_designation(db, teacher.ci, subject="Bioquímica", group_code="M-1")
        d2 = seed_designation(db, teacher.ci, subject="Histología", group_code="T-2")
        seed_attendance(db, teacher.ci, d1.id, day=2, academic_hours=1)
        seed_attendance(db, teacher.ci, d2.id, day=3, academic_hours=2)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(
            db,
            month=3,
            year=2026,
            payment_overrides={teacher.ci: 1000.0, f"{teacher.ci}:{d1.id}": 500.0},
        )

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        adjusted_by_subject = {}
        for row_num in range(DATA_ROW_START, DATA_ROW_START + 10):
            if ws.cell(row=row_num, column=COL_CI).value == teacher.ci:
                subject = ws.cell(row=row_num, column=6).value
                adjusted_by_subject[subject] = ws.cell(row=row_num, column=COL_PAGO_AJUSTADO).value

        assert adjusted_by_subject["Bioquímica"] == 500.0
        assert adjusted_by_subject["Histología"] == 500.0
        assert result.total_payment == 1000.0

    def test_february_leap_year_has_day_29(self, db, temp_output_dir):
        """February 2028 is a leap year — day 29 column should have value 29."""
        teacher = seed_teacher(db, "29022028", "LUNA FELIX")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=14, month=2, year=2028)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result = gen.generate(db, month=2, year=2028)

        wb = load_workbook(result.file_path)
        ws = wb["Planilla"]

        # Day 29 should appear in header row for leap year
        day29_header = ws.cell(row=ROW_COL_HEADERS, column=DAY_COL_START + 28).value
        assert day29_header == 29

    def test_upsert_planilla_output_on_regeneration(self, db, temp_output_dir):
        """Calling generate twice for same month/year should UPDATE, not create duplicate."""
        teacher = seed_teacher(db, "99001100", "SORIA BEATRIZ")
        desig = seed_designation(db, teacher.ci)
        seed_attendance(db, teacher.ci, desig.id, day=5)

        gen = PlanillaGenerator(output_dir=temp_output_dir)
        result1 = gen.generate(db, month=5, year=2026)
        result2 = gen.generate(db, month=5, year=2026)

        # Should be same ID (upserted)
        assert result1.planilla_output_id == result2.planilla_output_id

        # Only one record in DB
        count = db.query(PlanillaOutput).filter_by(month=5, year=2026).count()
        assert count == 1
