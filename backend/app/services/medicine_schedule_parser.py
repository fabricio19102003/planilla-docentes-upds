from __future__ import annotations

import hashlib
import re
import unicodedata
from datetime import time
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.medicine_schedule import (
    MedicineIssuePreview, MedicineMeetingPreview, MedicineOfferingPreview,
    MedicineWorkbookPreview,
)

PARSER_SCHEMA_VERSION = "medicine-v1"
DAYS = {"LUNES": "monday", "MARTES": "tuesday", "MIERCOLES": "wednesday",
        "JUEVES": "thursday", "VIERNES": "friday", "SABADO": "saturday"}
SEMESTERS = {"PRIMER": 1, "SEGUNDO": 2, "TERCER": 3, "CUARTO": 4,
             "QUINTO": 5, "SEXTO": 6, "SEPTIMO": 7, "OCTAVO": 8}
SHIFTS = {"M": "morning", "T": "afternoon", "N": "night"}
TIME_RE = re.compile(r"^\s*([0-2]?\d):([0-5]\d)\s*-\s*([0-2]?\d):([0-5]\d)\s*$")


def _fold(value: Any) -> str:
    text = unicodedata.normalize("NFKD", "" if value is None else str(value))
    return " ".join("".join(c for c in text if not unicodedata.combining(c)).upper().split())


def _clean(value: Any) -> str | None:
    text = " ".join(str(value).split()) if value is not None else ""
    return text or None

def _value(ws: Worksheet, row: int, column: int) -> Any:
    cell = ws.cell(row, column)
    return None if isinstance(cell, MergedCell) else cell.value

def _resolved_cell(ws: Worksheet, row: int, column: int):
    cell = ws.cell(row, column)
    if not isinstance(cell, MergedCell):
        return cell
    merged = next(area for area in ws.merged_cells.ranges if cell.coordinate in area)
    return ws.cell(merged.min_row, merged.min_col)

def normalize_group(value: str) -> str:
    match = re.fullmatch(r"([MTN])\s*-?\s*0*(\d+)", _fold(value))
    if not match or int(match.group(2)) < 1:
        raise ValueError("Group must be M/T/N followed by a positive number")
    return f"{match.group(1)}{int(match.group(2))}"

def _time_range(value: Any) -> tuple[time, time, bool] | None:
    raw = str(value or "")
    tokens = re.findall(r"(?<!\d)([0-2]?\d)\s*:\s*([0-5]\d)(?!\d)", raw)
    if len(tokens) != 2:
        return None
    start, end = (time(int(hour), int(minute)) for hour, minute in tokens)
    return (start, end, not bool(TIME_RE.fullmatch(raw))) if end > start else None


def _issue(severity: str, code: str, message: str, sheet: str, cell: str) -> MedicineIssuePreview:
    return MedicineIssuePreview(severity=severity, code=code, message=message,
                                location={"sheet": sheet, "cell": cell})


def _headers(ws: Worksheet) -> list[tuple[int, int, dict[int, str]]]:
    found = []
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            if _fold(_value(ws, row, column)) != "HORARIO":
                continue
            days = {col: DAYS[_fold(_value(ws, row, col))] for col in range(column + 1, min(ws.max_column, column + 7) + 1)
                    if _fold(_value(ws, row, col)) in DAYS}
            if len(days) >= 3:
                found.append((row, column, days))
    return found


def _metadata(ws: Worksheet, row: int, end_col: int) -> str:
    return " | ".join(filter(None, (_clean(_value(ws, r, c))
        for r in range(max(1, row - 8), row) for c in range(1, end_col + 1))))


def _activity(subject: str, detail: str | None) -> str | None:
    text = _fold(f"{subject} {detail or ''}")
    for markers, activity in ((('LAB', 'ANFITEATRO'), 'laboratory'), (('PRACTI',), 'practice'), (('TEORI',), 'theory')):
        if any(marker in text for marker in markers):
            return activity
    return None


def _semesters(text: str, sheet_name: bool = False) -> set[int]:
    found = {number for word, number in SEMESTERS.items() if re.search(rf"\b{word}\s+SEMESTRE\b", text)}
    found |= {int(value) for value in re.findall(r"\b([1-8])(?:RO|DO|ER|TO|MO|VO)?\s+SEMESTRE\b", text)}
    if sheet_name and not found:
        found |= {int(value) for value in re.findall(r"\b([1-8])(?:RO|DO|ER|TO|MO|VO)\b", text)}
    return found


def _layout(ws: Worksheet, first_row: int, last_row: int, column: int, days: dict[int, str]) -> str:
    rows = [row for row in range(first_row, last_row + 1) if _time_range(_value(ws, row, column))][:8]
    if any(any(re.search(r"TEORI|PRACTI|LAB|ANFITEATRO", _fold(_value(ws, row + 2, col)))
               for col in days) for row in rows if row + 2 <= last_row):
        return "subject_teacher_activity"
    if sum(any(_clean(_value(ws, row + 1, col)) for col in days) for row in rows if row < last_row) > len(rows) / 2:
        return "teacher_subject"
    return "subject_only"


def parse_medicine_workbook(content: bytes) -> MedicineWorkbookPreview:
    result = MedicineWorkbookPreview(workbook_sha256=hashlib.sha256(content).hexdigest(),
                                     parser_schema_version=PARSER_SCHEMA_VERSION)
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, keep_links=False)
    except Exception:
        result.issues.append(_issue("error", "unsupported_workbook", "File is not a readable XLSX workbook", "", ""))
        return result
    groups: dict[tuple[str, int | None, str, str], MedicineOfferingPreview] = {}
    header_count = 0
    for ws in workbook.worksheets:
        headers = _headers(ws)
        header_count += len(headers)
        for index, (row, column, days) in enumerate(headers):
            cell = ws.cell(row, column).coordinate
            metadata = _metadata(ws, row, max(days))
            block_text, sheet_text = _fold(metadata), _fold(ws.title)
            folded = f"{sheet_text} {block_text}"
            category = "convalidacion" if "CONVALIDACION" in folded else "regular"
            semesters, sheet_semesters = _semesters(block_text), _semesters(sheet_text, True)
            raw_groups = re.findall(r"\b([MTN]\s*-?\s*0*\d+)\b", block_text)
            normalized_groups = set()
            for value in raw_groups:
                try:
                    normalized_groups.add(normalize_group(value))
                except ValueError:
                    pass
            shifts = {name for word, name in (("MANANA", "morning"), ("TARDE", "afternoon"), ("NOCHE", "night")) if word in folded}
            if category == "convalidacion":
                semester, group = None, "SPECIAL"
            elif len(semesters) == 1 and len(normalized_groups) == 1:
                semester, group = next(iter(semesters)), next(iter(normalized_groups))
                if sheet_semesters and semester not in sheet_semesters:
                    result.issues.append(_issue("warning", "semester_sheet_mismatch", "Explicit block semester differs from sheet", ws.title, cell))
            elif not semesters and len(sheet_semesters) == 1 and len(normalized_groups) == 1:
                semester, group = next(iter(sheet_semesters)), next(iter(normalized_groups))
                result.issues.append(_issue("warning", "semester_from_sheet", "Semester read from explicit sheet name", ws.title, cell))
            else:
                code = "missing_metadata" if not semesters or not normalized_groups else "contradictory_metadata"
                result.issues.append(_issue("error", code, "Block requires one unambiguous semester and group", ws.title, cell))
                continue
            if semester == 7:
                result.unsupported_semesters = [7]
                result.issues.append(_issue("warning", "semester_7_unsupported", "Seventh semester matrix is detected but not extracted", ws.title, cell))
                continue
            shift = next(iter(shifts)) if len(shifts) == 1 else SHIFTS.get(group[0])
            if len(shifts) > 1 or shift is None or (category == "regular" and shifts and shift != SHIFTS[group[0]]):
                result.issues.append(_issue("error", "contradictory_shift", "Shift contradicts group", ws.title, cell))
                continue
            if not shifts:
                result.issues.append(_issue("warning", "shift_from_group", "Shift derived from explicit group", ws.title, cell))
            end_row = headers[index + 1][0] - 1 if index + 1 < len(headers) else ws.max_row
            layout = _layout(ws, row + 1, end_row, column, days)
            if layout != "subject_teacher_activity":
                result.issues.append(_issue("warning", "alternate_layout", f"Parsed recognized {layout} block", ws.title, cell))
            breaks = 0
            for data_row in range(row + 1, end_row + 1):
                raw_time = _value(ws, data_row, column)
                parsed_time = _time_range(raw_time)
                if not parsed_time:
                    if _clean(raw_time) and any(_clean(_value(ws, data_row, col)) for col in days):
                        result.issues.append(_issue("error", "invalid_time", "Schedule row has an invalid time range", ws.title,
                                                    ws.cell(data_row, column).coordinate))
                    continue
                if parsed_time[2]:
                    result.issues.append(_issue("warning", "normalized_time", "Normalized an unambiguous time separator", ws.title,
                                                ws.cell(data_row, column).coordinate))
                for day_col, day in days.items():
                    subject_offset, teacher_offset = ((1, 0) if layout == "teacher_subject" else (0, 1))
                    subject_cell = _resolved_cell(ws, data_row + subject_offset, day_col)
                    teacher_cell = _resolved_cell(ws, data_row + teacher_offset, day_col) if layout != "subject_only" else None
                    detail_cell = _resolved_cell(ws, data_row + 2, day_col) if layout == "subject_teacher_activity" else None
                    subject = _clean(subject_cell.value)
                    teacher = _clean(teacher_cell.value) if teacher_cell else None
                    detail = _clean(detail_cell.value) if detail_cell else None
                    if not subject:
                        continue
                    if _fold(subject) == "RECESO":
                        breaks += 1
                        continue
                    activity = _activity(subject, detail)
                    activity = activity or "unspecified"
                    subject_key, teacher_key = _fold(subject), _fold(teacher) or None
                    key = (category, semester, subject_key, group)
                    offering = groups.setdefault(key, MedicineOfferingPreview(
                        category=category, semester=semester, subject_raw=subject, subject_key=subject_key,
                        group_code=group, shift=shift, source_sheet=ws.title, source_row=data_row,
                        raw_payload={"metadata": metadata, "subject": subject, "subject_cell": subject_cell.coordinate}))
                    offering.meetings.append(MedicineMeetingPreview(
                        activity=activity, teacher_raw=teacher, teacher_key=teacher_key,
                        day=day, start_time=parsed_time[0], end_time=parsed_time[1],
                        source_cell=subject_cell.coordinate,
                        raw_payload={"time": _value(ws, data_row, column), "subject": subject,
                                     "teacher": teacher, "detail": detail, "time_cell": ws.cell(data_row, column).coordinate,
                                     "teacher_cell": teacher_cell.coordinate if teacher_cell else None,
                                     "detail_cell": detail_cell.coordinate if detail_cell else None}))
                    if not teacher:
                        result.issues.append(_issue("warning", "missing_teacher", "Teacher is missing; no value inferred", ws.title,
                                                    ws.cell(data_row, day_col).coordinate))
            if breaks:
                result.issues.append(_issue("warning", "break_excluded", f"Excluded {breaks} break cells", ws.title, cell))
    workbook.close()
    if not header_count:
        result.issues.append(_issue("error", "unsupported_structure", "No supported Medicine schedule block found", "", ""))
    result.offerings = list(groups.values())
    return result
