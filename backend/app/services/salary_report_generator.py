"""
Service: Salary Report Generator

Generates an Excel "Planilla Salarios" report that EXACTLY replicates the
template `PLANTILLA_PLANILLA_SALARIO_DOCENTES_MARZO_2026.xlsx`.

The report is a single sheet named "<MES> <AÑO>" (e.g. "MARZO 2026") and
contains one data row per teacher × subject × semester combination (same
rows produced by PlanillaGenerator._build_planilla_data).

Layout summary:
    Row 1   : Company name  (B1)
    Row 2   : Company NIT   (B2)
    Row 3   : spacer
    Row 4   : Merged A4:O4 title "PLANILLA HONORARIOS DOCENTES MEDICINA - MES DE <MES> <AÑO>"
    Row 5   : spacer
    Row 6   : Column headers (green fill, bold)
    Row 7+  : Data rows
    Row N+1 : Totals row (SUBTOTAL on J/K/L, merged TOTAL label on B..I)

Retention logic:
    - K column holds `=J{row}*13%` ONLY when the teacher has retention
      (row.has_retention == True). Empty cell otherwise.
    - L column is ALWAYS `=J{row}-K{row}` (Excel treats empty K as 0).

M column (NIT):
    - If teacher.nit is set, use teacher.nit.
    - Else if row.has_retention, use the literal "RETENCION".
    - Else leave empty.

Print setup mirrors the template: landscape, scale=19, fitToHeight=0.
"""
from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from sqlalchemy.orm import Session

from app.models.teacher import Teacher
from app.services.planilla_generator import PlanillaGenerator

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MONTH_NAMES_UPPER: dict[int, str] = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}

# Colors (openpyxl uses AARRGGBB or RRGGBB)
COLOR_BLACK = "FF000000"
COLOR_HEADER_GREEN = "FF92D050"

# Currency format used on J / K / L (template-exact, note trailing space)
CURRENCY_FORMAT = '#,##0.00_ ;\\-#,##0.00\\ '

# Borders
_THIN_SIDE = Side(border_style="thin", color="FF000000")
THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

# Column widths (match template exactly)
COLUMN_WIDTHS: dict[str, float] = {
    "A": 8,
    "B": 41.43,
    "C": 17,
    "D": 28,
    "E": 10.43,
    "F": 33.71,
    "G": 37.57,
    "H": 15.57,
    "I": 12.71,
    "J": 16.71,
    "K": 18.86,
    "L": 17,
    "M": 12.43,
    "N": 19.29,
    "O": 17.71,
}

# Row heights for the title block
TITLE_ROW_HEIGHTS: dict[int, float] = {
    1: 25.9,
    2: 25.9,
    3: 15.0,
    4: 25.9,
    5: 12.75,
}
HEADER_ROW_HEIGHT = 27.6
DATA_ROW_HEIGHT = 30.0

DATA_ROW_START = 7


# ---------------------------------------------------------------------------
# Main service
# ---------------------------------------------------------------------------


class SalaryReportGenerator:
    """Generate the "Planilla Salarios" Excel file for a given month/year."""

    def __init__(self, output_dir: str = "data/output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate(
        self,
        db: Session,
        month: int,
        year: int,
        company_name: Optional[str] = None,
        company_nit: Optional[str] = None,
        discount_mode: str = "attendance",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> Path:
        """
        Build the Excel file and return its path.

        Args:
            db: SQLAlchemy session.
            month: 1–12.
            year: calendar year.
            company_name: header company name. Falls back to the value stored
                in ``app_settings`` (key ``COMPANY_NAME``) when the caller
                passes None.
            company_nit: header company NIT. Falls back to the value stored
                in ``app_settings`` (key ``COMPANY_NIT``).
            discount_mode: "attendance" or "full" — forwarded to PlanillaGenerator.
            start_date / end_date: optional attendance window filter.
        """
        # Import here to avoid circulars at module load time
        from app.services import app_settings_service

        if company_name is None:
            company_name = app_settings_service.get_company_name(db)
        if company_nit is None:
            company_nit = app_settings_service.get_company_nit(db)

        month_name = MONTH_NAMES_UPPER.get(month, str(month))

        logger.info(
            "SalaryReportGenerator.generate: month=%d year=%d discount_mode=%s",
            month, year, discount_mode,
        )

        # Step 1: build planilla rows using the shared generator
        gen = PlanillaGenerator()
        rows, _details, _warnings = gen._build_planilla_data(
            db,
            month=month,
            year=year,
            start_date=start_date,
            end_date=end_date,
            discount_mode=discount_mode,
        )

        # Sort rows: teacher_name → subject → group_code (same as planilla)
        rows.sort(key=lambda r: (r.teacher_name, r.subject, r.group_code))

        # Step 2: bulk-load teachers by CI for phone/email/nit/bank/account_number
        cis = {r.teacher_ci for r in rows}
        teachers: dict[str, Teacher] = {
            t.ci: t
            for t in db.query(Teacher).filter(Teacher.ci.in_(cis)).all()
        } if cis else {}

        # Step 3: build workbook
        wb = Workbook()
        # Remove default "Sheet" and create our named one
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)
        ws = wb.create_sheet(title=f"{month_name} {year}")

        self._apply_column_widths(ws)
        self._write_title_block(ws, company_name, company_nit, month_name, year)
        self._write_header_row(ws)
        last_data_row = self._write_data_rows(ws, rows, teachers)
        total_row = last_data_row + 1 if rows else DATA_ROW_START
        self._write_totals_row(ws, total_row, last_data_row if rows else DATA_ROW_START - 1)
        self._apply_print_setup(ws, total_row)

        # Step 4: save file
        filename = f"planilla_salario_{month:02d}_{year}.xlsx"
        file_path = self.output_dir / filename
        wb.save(str(file_path))
        logger.info("Saved salary report to %s", file_path)

        return file_path

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------

    def _apply_column_widths(self, ws) -> None:
        for col_letter, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

    # ------------------------------------------------------------------
    # Title block (rows 1–5)
    # ------------------------------------------------------------------

    def _write_title_block(
        self,
        ws,
        company_name: str,
        company_nit: str,
        month_name: str,
        year: int,
    ) -> None:
        # Row heights
        for rn, height in TITLE_ROW_HEIGHTS.items():
            ws.row_dimensions[rn].height = height

        # B1 — company name
        b1 = ws["B1"]
        b1.value = company_name
        b1.font = Font(name="Calibri", size=20, bold=True, color=COLOR_BLACK)

        # B2 — NIT
        b2 = ws["B2"]
        b2.value = f"NIT: {company_nit}"
        b2.font = Font(name="Calibri", size=20, color=COLOR_BLACK)

        # Row 3 — spacer, no content

        # Row 4 — merged A4:O4 title
        ws.merge_cells("A4:O4")
        a4 = ws["A4"]
        a4.value = (
            f"PLANILLA HONORARIOS DOCENTES MEDICINA - MES DE {month_name} {year}"
        )
        a4.font = Font(name="Calibri", size=20, bold=True, italic=True, color=COLOR_BLACK)
        a4.alignment = Alignment(horizontal="center", vertical="center")

        # Row 5 — spacer

    # ------------------------------------------------------------------
    # Column headers (row 6)
    # ------------------------------------------------------------------

    def _write_header_row(self, ws) -> None:
        ws.row_dimensions[6].height = HEADER_ROW_HEIGHT

        header_fill = PatternFill("solid", fgColor=COLOR_HEADER_GREEN)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
        center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # (cell, value, font_name, size, wrap_text, number_format)
        headers: list[tuple[str, str, str, int, bool, str]] = [
            ("B6", "NOMBRE COMPLETO", "Arial", 10, False, "General"),
            ("C6", "Número de \nteléfono", "Arial", 10, True, "0"),
            ("D6", "Correo electrónico", "Arial", 10, False, "General"),
            ("E6", "Nº C.I.", "Arial", 10, False, "General"),
            ("F6", "MATERIA", "Arial", 11, False, "General"),
            ("G6", "TIPO DE CONTRATO", "Arial", 11, True, "General"),
            ("H6", "SEMESTRE", "Arial", 11, True, "General"),
            ("I6", "TOTAL HORAS", "Arial", 11, True, "General"),
            ("J6", "MONTO TOTAL", "Arial", 11, True, "General"),
            ("K6", "RETENCION RC IVA", "Arial", 11, True, "General"),
            ("L6", "LIQUIDO PAGABLE", "Arial", 11, True, "General"),
            ("M6", "NIT", "Arial", 11, False, "General"),
            ("N6", "No. CUENTA BANCARIA", "Arial", 11, True, "0"),
            ("O6", "BANCO", "Arial", 11, False, "General"),
        ]

        for coord, value, font_name, size, wrap, num_fmt in headers:
            cell = ws[coord]
            cell.value = value
            cell.font = Font(name=font_name, size=size, bold=True)
            cell.fill = header_fill
            cell.alignment = center_wrap if wrap else center_align
            cell.border = THIN_BORDER
            cell.number_format = num_fmt

    # ------------------------------------------------------------------
    # Data rows (starting row 7)
    # ------------------------------------------------------------------

    def _write_data_rows(self, ws, rows, teachers: dict[str, Teacher]) -> int:
        """Write all data rows. Returns the excel row number of the LAST data row."""
        base_font = Font(name="Arial", size=9)
        black_font = Font(name="Arial", size=9, color=COLOR_BLACK)

        align_left_center = Alignment(horizontal="general", vertical="center", wrap_text=False)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_right_wrap = Alignment(horizontal="right", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=False)

        for i, row in enumerate(rows):
            r = DATA_ROW_START + i
            ws.row_dimensions[r].height = DATA_ROW_HEIGHT
            teacher = teachers.get(row.teacher_ci)

            # A: spacer (empty, but bordered)
            self._set_cell(
                ws, f"A{r}", value=None, font=base_font,
                alignment=align_center, border=THIN_BORDER,
            )

            # B: teacher_name  (left default, no wrap, black)
            self._set_cell(
                ws, f"B{r}", value=row.teacher_name,
                font=black_font, alignment=align_left_center,
                border=THIN_BORDER, number_format="General",
            )

            # C: phone (center, format "0")
            phone = teacher.phone if teacher else None
            phone_value = self._coerce_numeric(phone)
            self._set_cell(
                ws, f"C{r}", value=phone_value,
                font=base_font, alignment=align_center,
                border=THIN_BORDER, number_format="0",
            )

            # D: email (center)
            self._set_cell(
                ws, f"D{r}", value=teacher.email if teacher else None,
                font=base_font, alignment=align_center,
                border=THIN_BORDER, number_format="General",
            )

            # E: CI (center)
            self._set_cell(
                ws, f"E{r}", value=row.teacher_ci,
                font=base_font, alignment=align_center,
                border=THIN_BORDER, number_format="General",
            )

            # F: subject (center, wrap)
            self._set_cell(
                ws, f"F{r}", value=row.subject,
                font=base_font, alignment=align_center_wrap,
                border=THIN_BORDER, number_format="General",
            )

            # G: fixed service description (center, wrap)
            self._set_cell(
                ws, f"G{r}", value="PAGO DE SERVICIOS PROFESIONALES",
                font=base_font, alignment=align_center_wrap,
                border=THIN_BORDER, number_format="General",
            )

            # H: semester (center, wrap)
            self._set_cell(
                ws, f"H{r}", value=row.semester,
                font=base_font, alignment=align_center_wrap,
                border=THIN_BORDER, number_format="General",
            )

            # I: payable hours (center, wrap)
            self._set_cell(
                ws, f"I{r}", value=row.payable_hours,
                font=base_font, alignment=align_center_wrap,
                border=THIN_BORDER, number_format="General",
            )

            # J: calculated payment (right, wrap, currency)
            self._set_cell(
                ws, f"J{r}", value=row.calculated_payment,
                font=base_font, alignment=align_right_wrap,
                border=THIN_BORDER, number_format=CURRENCY_FORMAT,
            )

            # K: retention formula only if has_retention
            k_value = f"=J{r}*13%" if row.has_retention else None
            self._set_cell(
                ws, f"K{r}", value=k_value,
                font=base_font, alignment=align_right_wrap,
                border=THIN_BORDER, number_format=CURRENCY_FORMAT,
            )

            # L: always J - K (Excel treats empty K as 0)
            self._set_cell(
                ws, f"L{r}", value=f"=J{r}-K{r}",
                font=black_font, alignment=align_right,
                border=THIN_BORDER, number_format=CURRENCY_FORMAT,
            )

            # M: NIT or "RETENCION"
            nit_value: Optional[str]
            if teacher and teacher.nit:
                nit_value = teacher.nit
            elif row.has_retention:
                nit_value = "RETENCION"
            else:
                nit_value = None
            self._set_cell(
                ws, f"M{r}", value=nit_value,
                font=base_font, alignment=align_center,
                border=THIN_BORDER, number_format="General",
            )

            # N: account number (center, wrap, format "0")
            account_raw = teacher.account_number if teacher else None
            account_value = self._coerce_numeric(account_raw)
            self._set_cell(
                ws, f"N{r}", value=account_value,
                font=base_font, alignment=align_center_wrap,
                border=THIN_BORDER, number_format="0",
            )

            # O: bank (center, wrap, black)
            self._set_cell(
                ws, f"O{r}", value=teacher.bank if teacher else None,
                font=black_font, alignment=align_center_wrap,
                border=THIN_BORDER, number_format="General",
            )

        last_row = DATA_ROW_START + len(rows) - 1
        return last_row if rows else DATA_ROW_START - 1

    # ------------------------------------------------------------------
    # Totals row
    # ------------------------------------------------------------------

    def _write_totals_row(self, ws, total_row: int, last_data_row: int) -> None:
        """
        Write the totals row.

        - Merge B{total_row}:I{total_row} → "TOTAL" (Arial 11 bold, right/center)
        - J/K/L: SUBTOTAL(9, X7:X{last_data_row}) with currency format
        - If there are no data rows, we still write a TOTAL row with zero subtotals
          referencing the same range — Excel will return 0.
        """
        ws.row_dimensions[total_row].height = DATA_ROW_HEIGHT

        # Ensure range is valid even when no data rows exist
        range_last = max(last_data_row, DATA_ROW_START)

        # Merged label B..I
        ws.merge_cells(f"B{total_row}:I{total_row}")
        label = ws[f"B{total_row}"]
        label.value = "TOTAL"
        label.font = Font(name="Arial", size=11, bold=True)
        label.alignment = Alignment(
            horizontal="right", vertical="center", wrap_text=True
        )
        label.border = THIN_BORDER
        # Apply borders to the merged cells so every cell in the range is outlined
        for col_letter in ("C", "D", "E", "F", "G", "H", "I"):
            cell = ws[f"{col_letter}{total_row}"]
            cell.border = THIN_BORDER

        # Also border A{total_row} so the row looks clean
        a_cell = ws[f"A{total_row}"]
        a_cell.border = THIN_BORDER

        total_font = Font(name="Arial", size=9, bold=True, color=COLOR_BLACK)
        total_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_letter in ("J", "K", "L"):
            cell = ws[f"{col_letter}{total_row}"]
            cell.value = f"=SUBTOTAL(9,{col_letter}{DATA_ROW_START}:{col_letter}{range_last})"
            cell.font = total_font
            cell.alignment = total_align
            cell.border = THIN_BORDER
            cell.number_format = CURRENCY_FORMAT

        # Trailing columns M, N, O also bordered so the row is continuous
        for col_letter in ("M", "N", "O"):
            cell = ws[f"{col_letter}{total_row}"]
            cell.border = THIN_BORDER

    # ------------------------------------------------------------------
    # Print setup
    # ------------------------------------------------------------------

    def _apply_print_setup(self, ws, total_row: int) -> None:
        # Print area covers A1:P{total_row+1} to match template
        ws.print_area = f"A1:P{total_row + 1}"

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.scale = 19
        ws.page_setup.fitToHeight = 0
        # Keep fitToWidth enabled for consistency with landscape fit-to-width
        ws.page_setup.fitToWidth = 1

        ws.page_margins = PageMargins(
            left=0.551,
            right=0.709,
            top=0.748,
            bottom=0.748,
            header=0.276,
            footer=0.315,
        )

    # ------------------------------------------------------------------
    # Small helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _set_cell(
        ws,
        coord: str,
        *,
        value,
        font: Font,
        alignment: Alignment,
        border: Border,
        number_format: str = "General",
    ) -> None:
        cell = ws[coord]
        cell.value = value
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        cell.number_format = number_format

    @staticmethod
    def _coerce_numeric(value):
        """
        Convert a stringy numeric value (phone, account number) into an int
        when possible so the "0" number format renders as a plain integer.
        Non-numeric strings are kept as-is; None stays None.
        """
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return value
        s = str(value).strip()
        if not s:
            return None
        # Only coerce pure digit strings — preserve things like "+591…" or
        # account numbers with dashes as text.
        if s.isdigit():
            try:
                return int(s)
            except ValueError:
                return s
        return s
